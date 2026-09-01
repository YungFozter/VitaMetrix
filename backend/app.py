import os
import sys
import json
import logging
import time
import math
import uuid
import re
import hmac
import hashlib
import base64
import secrets
from datetime import datetime, timezone, timedelta
from werkzeug.security import generate_password_hash, check_password_hash

# Configurar logging seguro
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

_BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
if _BACKEND_DIR not in sys.path:
    sys.path.insert(0, _BACKEND_DIR)

from typing import Optional

from flask import Flask, render_template, request, jsonify
from dotenv import load_dotenv
from supabase import create_client, Client
from calculations import (
    get_biva_interpretation,
    calculate_energy,
    calculate_scores,
    analyze_hydration,
    analyze_visceral_fat,
    build_clinical_report,
)
from reference import (
    get_phase_angle_percentile,
    get_smm_percentile,
    get_smm_age_curves,
    get_pha_age_curves,
    analyze_segmental,
    analyze_composition_indices,
    load_tables,
)

load_dotenv()

app = Flask(
    __name__,
    template_folder='../frontend/templates',
    static_folder='../frontend/static'
)

app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # Limitar tamaño de payload a 2 MB

@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self' https: data: blob:; "
        "script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdn.jsdelivr.net https://unpkg.com https://cdn.tailwindcss.com https://translate.google.com https://translate.googleapis.com https://www.gstatic.com; "
        "style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://fonts.googleapis.com https://unpkg.com https://www.gstatic.com https://translate.googleapis.com; "
        "font-src 'self' data: https://fonts.gstatic.com https://cdn.jsdelivr.net https://www.gstatic.com; "
        "img-src 'self' data: blob: https:; "
        "connect-src 'self' https: https://nominatim.openstreetmap.org https://translate.googleapis.com; "
        "frame-src 'self' https://maps.google.com https://www.google.com https://*.google.com;"
    )
    return response

import html

def _clean_str(val, max_len=150):
    if not val:
        return ""
    return html.escape(str(val).strip()[:max_len])

def _normalize_gender(val):
    if not val:
        return "male"
    v = str(val).strip().lower()
    if v in ("f", "female", "femenino", "mujer", "femenina"):
        return "female"
    return "male"

# Inicializar Supabase
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")

supabase: Optional[Client] = None
if SUPABASE_URL and SUPABASE_KEY:
    try:
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception as e:
        logging.error("Error al inicializar cliente de Supabase: %s", e)
else:
    logging.warning("Supabase credentials not found in environment variables.")

_EMPTY_DASHBOARD = {
    "total_clients": 0,
    "total_evaluations": 0,
    "avg_score": 0,
    "recent": [],
    "population": {"Óptimo": 0, "Límite": 0, "Bajo": 0},
}


def _cell_bucket(phase_angle, valid=True):
    """Clasifica estado celular para el gráfico del dashboard (Óptimo / Límite / Bajo)."""
    if not valid or phase_angle is None:
        return "Límite"
    if phase_angle > 6.0:
        return "Óptimo"
    if phase_angle >= 5.0:
        return "Límite"
    return "Bajo"


# --- SISTEMA DE AUTENTICACIÓN MULTI-TENANT & SUSCRIPCIONES SAAS ---

_USERS_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "users.json")
_LICENSES_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "subscription_licenses.json")
_JWT_SECRET = os.environ.get("JWT_SECRET", "vitametrix_master_security_jwt_secret_2026_super_key_bolivia")

_DEFAULT_INITIAL_USERS = [
    {
        "id": "usr-admin-001",
        "email": "admin@vitametrix.com",
        "password_hash": generate_password_hash("AdminVita2026!"),
        "full_name": "Administrador General",
        "professional_title": "Director / Administrador de Plataforma",
        "clinic_name": "Sede Central VitaMetrix",
        "phone": "+59172125280",
        "role": "admin",
        "subscription_status": "lifetime",
        "subscription_plan": "Plan Ilimitado / Administrador",
        "subscription_expires_at": "2099-12-31T23:59:59Z",
        "trial_started_at": "2026-01-01T00:00:00Z",
        "created_at": "2026-01-01T00:00:00Z"
    },
    {
        "id": "usr-doctor-001",
        "email": "audrey@vitametrix.com",
        "password_hash": generate_password_hash("Doctora2026!"),
        "full_name": "Dra. Audrey",
        "professional_title": "Manager / Especialista BIA",
        "clinic_name": "Centro Médico VitaMetrix",
        "phone": "+59171234567",
        "role": "user",
        "subscription_status": "no_subscription",
        "subscription_plan": "Sin Suscripción Anterior",
        "subscription_expires_at": None,
        "trial_started_at": None,
        "created_at": "2026-08-28T00:00:00Z"
    }
]

_DEFAULT_INITIAL_LICENSES = []

def _save_users_disk_only(users):
    try:
        os.makedirs(os.path.dirname(_USERS_PATH), exist_ok=True)
        with open(_USERS_PATH, 'w', encoding='utf-8') as f:
            json.dump(users, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar users.json: %s", e)
        return False

def _clean_test_users(users):
    if not isinstance(users, list):
        return []
    if app.config.get('TESTING') or getattr(app, 'testing', False) or 'unittest' in sys.modules:
        return users
    clean = []
    for u in users:
        email = (u.get('email') or '').lower()
        uid = u.get('id') or ''
        if uid in ('usr-admin-001', 'usr-doctor-001') or email in ('admin@vitametrix.com', 'audrey@vitametrix.com'):
            clean.append(u)
        elif not any(p in email for p in ('dr.lic.', 'dr.a.', 'dr.b.', 'dra.elena.', 'dr.estandar.', 'dr.canje.', 'dra.test.', 'dr.roberto.', 'test.')):
            clean.append(u)
    return clean

def _load_users():
    # 1. Intentar cargar desde tabla users en Supabase
    if supabase:
        try:
            res = supabase.table('users').select('*').execute()
            if res and res.data and len(res.data) > 0:
                clean_res = _clean_test_users(res.data)
                _save_users_disk_only(clean_res)
                return list(clean_res)
        except Exception:
            pass

    # 2. Respaldo resiliente en almacén de Supabase (sobrevive a redeploys de Render)
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').eq('code', '__SYS_USERS_STORE__').execute()
            if res and res.data and len(res.data) > 0:
                notes = res.data[0].get('notes')
                if notes:
                    parsed = json.loads(notes)
                    if isinstance(parsed, list) and len(parsed) > 0:
                        clean_parsed = _clean_test_users(parsed)
                        _save_users_disk_only(clean_parsed)
                        return clean_parsed
        except Exception as e:
            logging.warning("Error al leer respaldo de usuarios en Supabase: %s", e)

    # 3. Disco local
    if os.path.exists(_USERS_PATH):
        try:
            with open(_USERS_PATH, 'r', encoding='utf-8') as f:
                users = json.load(f)
                if isinstance(users, list) and len(users) > 0:
                    clean_local = _clean_test_users(users)
                    return clean_local
        except Exception as e:
            logging.warning("Error al leer users.json: %s", e)

    _save_users(_DEFAULT_INITIAL_USERS)
    return list(_DEFAULT_INITIAL_USERS)

def _save_users(users):
    _save_users_disk_only(users)
    if supabase:
        # Intento en tabla directa
        try:
            for u in users:
                supabase.table('users').upsert(u).execute()
        except Exception:
            pass

        # Respaldo garantizado en almacén persistente de Supabase
        try:
            backup_data = {
                "code": "__SYS_USERS_STORE__",
                "name": "System Users Store (Auto-Backup)",
                "category": "__SYSTEM__",
                "unit": "JSON",
                "stock_quantity": len(users),
                "notes": json.dumps(users, ensure_ascii=False)
            }
            check = supabase.table('stock_items').select('id').eq('code', '__SYS_USERS_STORE__').execute()
            if check and check.data and len(check.data) > 0:
                supabase.table('stock_items').update(backup_data).eq('code', '__SYS_USERS_STORE__').execute()
            else:
                supabase.table('stock_items').insert(backup_data).execute()
        except Exception as e:
            logging.warning("Error al sincronizar usuarios en almacén persistente de Supabase: %s", e)
    return True

def _save_licenses_disk_only(licenses):
    try:
        os.makedirs(os.path.dirname(_LICENSES_PATH), exist_ok=True)
        with open(_LICENSES_PATH, 'w', encoding='utf-8') as f:
            json.dump(licenses, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar subscription_licenses.json: %s", e)
        return False

def _load_licenses():
    # 1. Intentar cargar desde tabla directa subscription_licenses en Supabase
    if supabase:
        try:
            res = supabase.table('subscription_licenses').select('*').order('created_at', desc=True).execute()
            if res and res.data and len(res.data) > 0:
                _save_licenses_disk_only(res.data)
                return list(res.data)
        except Exception:
            pass

    # 2. Respaldo resiliente en almacén de Supabase (sobrevive a redeploys de Render)
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').eq('code', '__SYS_LICENSES_STORE__').execute()
            if res and res.data and len(res.data) > 0:
                notes = res.data[0].get('notes')
                if notes:
                    parsed = json.loads(notes)
                    if isinstance(parsed, list) and len(parsed) > 0:
                        _save_licenses_disk_only(parsed)
                        return parsed
        except Exception as e:
            logging.warning("Error al leer respaldo de licencias en Supabase: %s", e)

    # 3. Disco local
    if os.path.exists(_LICENSES_PATH):
        try:
            with open(_LICENSES_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if isinstance(data, list) and len(data) > 0:
                    return data
        except Exception as e:
            logging.warning("Error al leer subscription_licenses.json: %s", e)

    return []

def _save_licenses(licenses):
    _save_licenses_disk_only(licenses)
    if supabase:
        # Intento en tabla dedicada
        try:
            for lic in licenses:
                supabase.table('subscription_licenses').upsert(lic).execute()
        except Exception:
            pass

        # Respaldo garantizado en almacén persistente de Supabase
        try:
            backup_data = {
                "code": "__SYS_LICENSES_STORE__",
                "name": "System Licenses Store (Auto-Backup)",
                "category": "__SYSTEM__",
                "unit": "JSON",
                "stock_quantity": len(licenses),
                "notes": json.dumps(licenses, ensure_ascii=False)
            }
            check = supabase.table('stock_items').select('id').eq('code', '__SYS_LICENSES_STORE__').execute()
            if check and check.data and len(check.data) > 0:
                supabase.table('stock_items').update(backup_data).eq('code', '__SYS_LICENSES_STORE__').execute()
            else:
                supabase.table('stock_items').insert(backup_data).execute()
        except Exception as e:
            logging.warning("Error al sincronizar licencias en almacén persistente de Supabase: %s", e)
    return True

def _generate_auth_token(user_id, email, role="user"):
    payload = {
        "user_id": user_id,
        "email": email,
        "role": role,
        "exp": int(time.time()) + (30 * 86400) # 30 días de validez
    }
    payload_json = json.dumps(payload, separators=(',', ':'))
    b64_payload = base64.urlsafe_b64encode(payload_json.encode()).decode().rstrip('=')
    sig = hmac.new(_JWT_SECRET.encode(), b64_payload.encode(), hashlib.sha256).hexdigest()
    return f"{b64_payload}.{sig}"

def _verify_auth_token(token_str):
    if not token_str or '.' not in token_str:
        return None
    try:
        b64_payload, sig = token_str.split('.', 1)
        expected_sig = hmac.new(_JWT_SECRET.encode(), b64_payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(sig, expected_sig):
            return None

        padded = b64_payload + '=' * ((4 - len(b64_payload) % 4) % 4)
        payload_bytes = base64.urlsafe_b64decode(padded)
        payload = json.loads(payload_bytes.decode())

        if payload.get('exp', 0) < time.time():
            return None
        return payload
    except Exception as e:
        logging.warning("Error validando token: %s", e)
        return None

def _get_current_user():
    auth_header = request.headers.get('Authorization', '')
    token = None
    if auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    elif request.cookies.get('vm_auth_token'):
        token = request.cookies.get('vm_auth_token')
    elif request.args.get('token'):
        token = request.args.get('token')

    if token:
        payload = _verify_auth_token(token)
        if payload and payload.get('user_id'):
            user_id = payload['user_id']
            for u in _load_users():
                if u.get('id') == user_id:
                    return u
            if supabase:
                try:
                    res = supabase.table('users').select('*').eq('id', user_id).execute()
                    if res.data:
                        return res.data[0]
                except Exception:
                    pass

    # Usuario demo predeterminado para retrocompatibilidad
    users = _load_users()
    if len(users) > 1:
        return users[1]
    return _DEFAULT_INITIAL_USERS[1]

def _calc_subscription_status(user):
    if not user:
        return "no_subscription", 0, None, "Sin Suscripción Anterior"

    if user.get('role') == 'admin':
        return "lifetime", None, None, "Acceso Total SuperAdmin / Incaducable"

    if user.get('subscription_status') == 'lifetime':
        return "lifetime", None, None, user.get('subscription_plan', 'Plan Vitalicio Ilimitado')

    # Cuenta sin suscripción previa
    if user.get('subscription_status') in ('no_subscription', 'none', 'unsubscribed') or (user.get('subscription_status') == 'expired' and not user.get('subscription_expires_at')):
        return "no_subscription", 0, None, "Sin Suscripción Anterior"

    expires_at_str = user.get('subscription_expires_at')
    plan_name = user.get('subscription_plan', 'Plan Pro Mensual')
    
    if not expires_at_str:
        if user.get('subscription_status') == 'trial':
            return "trial", 7, (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(), plan_name
        return "no_subscription", 0, None, "Sin Suscripción Anterior"

    try:
        exp_dt = datetime.fromisoformat(expires_at_str.replace('Z', '+00:00'))
        now_dt = datetime.now(timezone.utc)
        diff = exp_dt - now_dt
        days_left = max(0, diff.days + (1 if diff.seconds > 0 else 0))
        is_active = diff.total_seconds() > 0

        if is_active:
            status = "trial" if ("prueba" in plan_name.lower() or "trial" in plan_name.lower()) else "active"
            return status, days_left, exp_dt.isoformat(), plan_name
        else:
            return "expired", 0, exp_dt.isoformat(), plan_name or "Plan Vencido"
    except Exception:
        return "no_subscription", 0, None, plan_name or "Sin Suscripción Anterior"

def _is_subscription_active(user):
    if not user:
        return False
    if user.get('role') == 'admin':
        return True
    status, days_left, expires_at, plan_name = _calc_subscription_status(user)
    if status in ('active', 'trial', 'lifetime'):
        if days_left is None or days_left > 0:
            return True
    return False

def _build_safe_user_dict(user):
    status, days_left, expires_at, plan_name = _calc_subscription_status(user)
    return {
        "id": user.get('id'),
        "email": user.get('email'),
        "full_name": user.get('full_name'),
        "professional_title": user.get('professional_title', 'Nutricionista / Especialista BIA'),
        "clinic_name": user.get('clinic_name', 'Centro Médico VitaMetrix'),
        "phone": user.get('phone', ''),
        "professional_license": user.get('professional_license', ''),
        "clinic_logo_url": user.get('clinic_logo_url', ''),
        "pdf_disclaimer": user.get('pdf_disclaimer', 'Consulte con su profesional de la salud antes de iniciar cualquier plan nutricional o de entrenamiento.'),
        "pdf_footer_address": user.get('pdf_footer_address', ''),
        "clinic_address": user.get('clinic_address', ''),
        "unit_weight": user.get('unit_weight', 'kg'),
        "pha_optimal": user.get('pha_optimal', '6.0'),
        "role": user.get('role', 'user'),
        "subscription_status": status,
        "subscription_plan": plan_name,
        "subscription": {
            "status": status,
            "days_left": days_left,
            "expires_at": expires_at,
            "plan_name": plan_name,
            "whatsapp_contact": "+591 72125280",
            "whatsapp_phone_clean": "59172125280"
        }
    }

def _generate_next_user_id(role='user'):
    users = _load_users()
    prefix = 'usr-admin-' if role == 'admin' else 'usr-doctor-'
    existing_nums = []
    for u in users:
        uid = str(u.get('id') or '')
        if uid.startswith(prefix):
            try:
                num = int(uid.replace(prefix, ''))
                existing_nums.append(num)
            except ValueError:
                pass
    next_num = 1
    while next_num in existing_nums:
        next_num += 1
    return f"{prefix}{next_num:03d}"

# --- RUTAS DE AUTENTICACIÓN ---

@app.route('/api/auth/register', methods=['POST'])
def auth_register():
    data = request.json or {}
    email = _clean_str(data.get('email'), max_len=120).lower().strip()
    password = str(data.get('password') or '').strip()
    full_name = _clean_str(data.get('full_name'), max_len=120)
    professional_title = _clean_str(data.get('professional_title'), max_len=100) or "Nutricionista / Especialista BIA"
    clinic_name = _clean_str(data.get('clinic_name'), max_len=150) or "Mi Consultorio VitaMetrix"
    phone = _clean_str(data.get('phone'), max_len=30)

    if not email or '@' not in email:
        return jsonify({"error": "Por favor ingresa un correo electrónico válido"}), 400
    if not password or len(password) < 6:
        return jsonify({"error": "La contraseña debe tener al menos 6 caracteres"}), 400
    if not full_name:
        return jsonify({"error": "El nombre completo del profesional es obligatorio"}), 400

    users = _load_users()
    if any(u.get('email', '').lower() == email for u in users):
        return jsonify({"error": "Ya existe una cuenta registrada con este correo electrónico"}), 409

    now_utc = datetime.now(timezone.utc)
    trial_expires = now_utc + timedelta(days=7)

    new_user = {
        "id": _generate_next_user_id(role='user'),
        "email": email,
        "password_hash": generate_password_hash(password),
        "full_name": full_name,
        "professional_title": professional_title,
        "clinic_name": clinic_name,
        "phone": phone,
        "role": "user",
        "subscription_status": "trial",
        "subscription_plan": "Plan de Prueba (7 días)",
        "subscription_expires_at": trial_expires.isoformat(),
        "trial_started_at": now_utc.isoformat(),
        "created_at": now_utc.isoformat()
    }

    if supabase:
        try:
            supabase.table('users').insert({
                "id": new_user['id'],
                "email": email,
                "password_hash": new_user['password_hash'],
                "full_name": full_name,
                "professional_title": professional_title,
                "clinic_name": clinic_name,
                "phone": phone,
                "role": "user",
                "subscription_status": "trial",
                "subscription_plan": "Plan de Prueba (7 días)",
                "subscription_expires_at": trial_expires.isoformat()
            }).execute()
        except Exception as e:
            logging.warning("Error al guardar usuario en Supabase (usando fallback local): %s", e)

    users.append(new_user)
    _save_users(users)

    token = _generate_auth_token(new_user['id'], email, new_user['role'])
    return jsonify({
        "success": True,
        "token": token,
        "user": _build_safe_user_dict(new_user),
        "message": "¡Cuenta creada con éxito! Se han activado 7 días de prueba gratuita."
    }), 201

@app.route('/api/auth/login', methods=['POST'])
def auth_login():
    data = request.json or {}
    email = _clean_str(data.get('email'), max_len=120).lower().strip()
    password = str(data.get('password') or '').strip()

    if not email or not password:
        return jsonify({"error": "Correo y contraseña son obligatorios"}), 400

    users = _load_users()
    target_user = next((u for u in users if u.get('email', '').lower() == email), None)

    if not target_user and supabase:
        try:
            res = supabase.table('users').select('*').eq('email', email).execute()
            if res.data:
                target_user = res.data[0]
                users.append(target_user)
                _save_users(users)
        except Exception:
            pass

    if not target_user:
        return jsonify({"error": "Credenciales incorrectas. Verifica tu correo o crea una cuenta nueva."}), 401

    if not check_password_hash(target_user.get('password_hash', ''), password):
        return jsonify({"error": "Contraseña incorrecta."}), 401

    token = _generate_auth_token(target_user['id'], target_user['email'], target_user.get('role', 'user'))
    return jsonify({
        "success": True,
        "token": token,
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    auth_header = request.headers.get('Authorization', '')
    token = None
    if auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    elif request.cookies.get('vm_auth_token'):
        token = request.cookies.get('vm_auth_token')
    elif request.args.get('token'):
        token = request.args.get('token')

    if not token:
        return jsonify({"error": "No autenticado", "authenticated": False}), 401

    payload = _verify_auth_token(token)
    if not payload or not payload.get('user_id'):
        return jsonify({"error": "Token inválido o expirado", "authenticated": False}), 401

    user_id = payload['user_id']
    target_user = None
    for u in _load_users():
        if u.get('id') == user_id:
            target_user = u
            break

    if not target_user and supabase:
        try:
            res = supabase.table('users').select('*').eq('id', user_id).execute()
            if res.data:
                target_user = res.data[0]
        except Exception:
            pass

    if not target_user:
        return jsonify({"error": "Usuario no encontrado", "authenticated": False}), 404

    return jsonify({
        "success": True,
        "authenticated": True,
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/auth/profile', methods=['PUT', 'POST'])
def auth_update_profile():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401

    data = request.json or {}
    user_id = current_user.get('id')

    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    if 'full_name' in data and data['full_name']:
        target_user['full_name'] = _clean_str(data.get('full_name'), max_len=120)
    if 'professional_title' in data:
        target_user['professional_title'] = _clean_str(data.get('professional_title'), max_len=100)
    if 'clinic_name' in data:
        target_user['clinic_name'] = _clean_str(data.get('clinic_name'), max_len=150)
    if 'phone' in data:
        target_user['phone'] = _clean_str(data.get('phone'), max_len=30)
    if 'professional_license' in data or 'pdf_mp' in data:
        target_user['professional_license'] = _clean_str(data.get('professional_license') or data.get('pdf_mp'), max_len=100)
    if 'clinic_logo_url' in data or 'pdf_logo_url' in data:
        target_user['clinic_logo_url'] = str(data.get('clinic_logo_url') or data.get('pdf_logo_url') or '')
    if 'pdf_disclaimer' in data:
        target_user['pdf_disclaimer'] = _clean_str(data.get('pdf_disclaimer'), max_len=500)
    if 'pdf_footer_address' in data:
        target_user['pdf_footer_address'] = _clean_str(data.get('pdf_footer_address'), max_len=300)
    if 'clinic_address' in data:
        target_user['clinic_address'] = _clean_str(data.get('clinic_address'), max_len=200)
    if 'unit_weight' in data:
        target_user['unit_weight'] = _clean_str(data.get('unit_weight'), max_len=10)
    if 'pha_optimal' in data:
        target_user['pha_optimal'] = _clean_str(data.get('pha_optimal'), max_len=10)

    _save_users(users)
    return jsonify({
        "success": True,
        "message": "Perfil y membrete PDF actualizado correctamente.",
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/auth/logout', methods=['POST'])
def auth_logout():
    return jsonify({"success": True, "message": "Sesión cerrada correctamente"})

# --- RUTAS DE SUSCRIPCIÓN & LICENCIAMIENTO ---

@app.route('/api/subscription/status', methods=['GET'])
def subscription_status():
    user = _get_current_user()
    safe_user = _build_safe_user_dict(user)
    sub = safe_user['subscription']

    # Mensaje predeterminado de WhatsApp codificado
    wa_msg = (
        f"¡Hola VitaMetrix! Deseo renovar/activar mi suscripción mensual Pro.\n"
        f"👤 Profesional: {safe_user.get('full_name')}\n"
        f"📧 Correo: {safe_user.get('email')}\n"
        f"🆔 ID de Cuenta: {safe_user.get('id')}"
    )

    return jsonify({
        "success": True,
        "subscription": sub,
        "user": safe_user,
        "whatsapp": {
            "phone_display": "+591 72125280",
            "phone_e164": "59172125280",
            "message_text": wa_msg
        }
    })

@app.route('/api/subscription/redeem', methods=['POST'])
def subscription_redeem():
    data = request.json or {}
    key_input = _clean_str(data.get('license_key') or data.get('pin_key') or data.get('pin'), max_len=80).upper().strip()

    if not key_input:
        return jsonify({"error": "Por favor ingresa un PIN de activación válido"}), 400

    user = _get_current_user()
    if not user:
        return jsonify({"error": "Debes iniciar sesión para canjear un PIN"}), 401

    licenses = _load_licenses()
    target_license = next((lic for lic in licenses if lic.get('license_key', '').upper().strip() == key_input), None)

    if not target_license:
        return jsonify({"error": "El PIN de activación ingresado no existe o es inválido. Verifica con soporte vía WhatsApp (+591 72125280)."}), 404

    if target_license.get('is_used'):
        return jsonify({"error": "Este PIN de activación ya fue canjeado anteriormente."}), 409

    duration_days = int(target_license.get('duration_days', 30))
    plan_name = target_license.get('plan_name', 'Plan Pro Mensual')
    now_utc = datetime.now(timezone.utc)
    is_lifetime = duration_days >= 9999 or "lifetime" in plan_name.lower() or "ilimitado" in plan_name.lower() or "vitalicio" in plan_name.lower()

    # Actualizar suscripción del usuario
    users = _load_users()
    for u in users:
        if u.get('id') == user.get('id'):
            if is_lifetime:
                u['subscription_status'] = 'lifetime'
                u['subscription_plan'] = plan_name
                u['subscription_expires_at'] = '2099-12-31T23:59:59Z'
            else:
                current_exp_str = u.get('subscription_expires_at')
                base_date = now_utc
                if current_exp_str:
                    try:
                        curr_dt = datetime.fromisoformat(current_exp_str.replace('Z', '+00:00'))
                        if curr_dt > now_utc:
                            base_date = curr_dt
                    except Exception:
                        pass
                new_expires_dt = base_date + timedelta(days=duration_days)
                u['subscription_status'] = 'active'
                u['subscription_plan'] = plan_name
                u['subscription_expires_at'] = new_expires_dt.isoformat()
            user = u
            break
    _save_users(users)

    # Marcar PIN como usado
    target_license['is_used'] = True
    target_license['used_by_user_id'] = user.get('id')
    target_license['used_by_name'] = user.get('full_name')
    target_license['used_by_email'] = user.get('email')
    target_license['used_at'] = now_utc.isoformat()
    _save_licenses(licenses)

    safe_user = _build_safe_user_dict(user)
    msg = "¡PIN canjeado con éxito! Tienes acceso ilimitado activado." if is_lifetime else f"¡PIN canjeado con éxito! Se han añadido {duration_days} días de suscripción Pro activa."
    return jsonify({
        "success": True,
        "message": msg,
        "subscription": safe_user['subscription'],
        "user": safe_user
    }), 200

# --- ENDPOINTS SUPERADMIN: GENERACIÓN & GESTIÓN DE PINS / LICENCIAS ---

@app.route('/api/admin/pins', methods=['GET'])
def admin_get_pins():
    caller, err = _require_admin()
    if err:
        return err

    licenses = _load_licenses()
    users = _load_users()
    user_map = {u.get('id'): u for u in users}

    enriched_pins = []
    for lic in licenses:
        pin_copy = dict(lic)
        used_id = pin_copy.get('used_by_user_id')
        if used_id and used_id in user_map:
            doctor = user_map[used_id]
            pin_copy['used_by_name'] = doctor.get('full_name', pin_copy.get('used_by_name'))
            pin_copy['used_by_email'] = doctor.get('email', pin_copy.get('used_by_email'))
            pin_copy['used_by_clinic'] = doctor.get('clinic_name')
        enriched_pins.append(pin_copy)

    # Ordenar más recientes primero
    enriched_pins.sort(key=lambda p: p.get('created_at') or '', reverse=True)

    available_count = sum(1 for p in enriched_pins if not p.get('is_used'))
    used_count = sum(1 for p in enriched_pins if p.get('is_used'))

    return jsonify({
        "success": True,
        "pins": enriched_pins,
        "stats": {
            "total_pins": len(enriched_pins),
            "available_pins": available_count,
            "used_pins": used_count
        }
    }), 200

@app.route('/api/admin/pins/create', methods=['POST'])
def admin_create_pin():
    caller, err = _require_admin()
    if err:
        return err

    data = request.json or {}
    duration_days = int(data.get('duration_days', 30))
    plan_name = _clean_str(data.get('plan_name'))
    custom_pin = _clean_str(data.get('custom_pin'), max_len=60).upper().strip()
    note = _clean_str(data.get('note'), max_len=200)
    count = min(max(int(data.get('count', 1)), 1), 50)

    if not plan_name:
        if duration_days >= 9999:
            plan_name = "Plan Vitalicio / Lifetime"
        elif duration_days == 365:
            plan_name = "Plan Anual Pro (365 días)"
        elif duration_days == 180:
            plan_name = "Plan Pro Semestral (180 días)"
        elif duration_days == 90:
            plan_name = "Plan Pro Trimestral (90 días)"
        elif duration_days == 7:
            plan_name = "Plan Prueba Especial (7 días)"
        else:
            plan_name = f"Plan Pro ({duration_days} días)"

    licenses = _load_licenses()
    existing_keys = {lic.get('license_key', '').upper().strip() for lic in licenses}

    created_records = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for _ in range(count):
        if custom_pin and count == 1:
            pin_key = custom_pin
            if pin_key in existing_keys:
                return jsonify({"error": f"El PIN '{pin_key}' ya existe. Por favor utiliza otro código."}), 409
        else:
            prefix = "VM-LIFE" if duration_days >= 9999 else ("VM-1M" if duration_days <= 30 else ("VM-3M" if duration_days <= 90 else "VM-1A"))
            part1 = secrets.token_hex(2).upper()
            part2 = secrets.token_hex(2).upper()
            pin_key = f"{prefix}-{part1}-{part2}"
            while pin_key in existing_keys:
                part1 = secrets.token_hex(2).upper()
                part2 = secrets.token_hex(2).upper()
                pin_key = f"{prefix}-{part1}-{part2}"

        existing_keys.add(pin_key)
        record = {
            "id": str(uuid.uuid4()),
            "license_key": pin_key,
            "duration_days": duration_days,
            "plan_name": plan_name,
            "note": note or "",
            "is_used": False,
            "used_by_user_id": None,
            "used_by_name": None,
            "used_by_email": None,
            "used_at": None,
            "created_at": now_iso
        }
        licenses.append(record)
        created_records.append(record)

    _save_licenses(licenses)

    return jsonify({
        "success": True,
        "message": f"Se generó {len(created_records)} PIN(s) de activación exitosamente.",
        "pins": created_records,
        "created_pins": created_records
    }), 201

@app.route('/api/admin/pins/<pin_id>', methods=['DELETE'])
def admin_delete_pin(pin_id):
    caller, err = _require_admin()
    if err:
        return err

    licenses = _load_licenses()
    initial_len = len(licenses)
    licenses = [lic for lic in licenses if lic.get('id') != pin_id and lic.get('license_key') != pin_id]

    if len(licenses) == initial_len:
        return jsonify({"error": "PIN no encontrado"}), 404

    _save_licenses_disk_only(licenses)

    if supabase:
        try:
            supabase.table('subscription_licenses').delete().eq('id', pin_id).execute()
            supabase.table('subscription_licenses').delete().eq('license_key', pin_id).execute()
        except Exception:
            pass

    return jsonify({
        "success": True,
        "message": "PIN de activación eliminado correctamente."
    }), 200

@app.route('/api/admin/pins/sync', methods=['POST'])
def admin_sync_pins():
    caller, err = _require_admin()
    if err:
        return err

    return jsonify({
        "success": True,
        "synced_count": 0,
        "total_pins": len(_load_licenses()),
        "pins": _load_licenses()
    }), 200

@app.route('/api/admin/licenses/create', methods=['POST'])
def admin_create_license_compat():
    data = request.json or {}
    duration_days = int(data.get('duration_days', 30))
    plan_name = _clean_str(data.get('plan_name')) or (f"Plan Pro {duration_days} días")
    count = min(max(int(data.get('count', 1)), 1), 20)

    created_keys = []
    licenses = _load_licenses()
    now_iso = datetime.now(timezone.utc).isoformat()

    for _ in range(count):
        prefix = "VM-1M" if duration_days <= 30 else ("VM-3M" if duration_days <= 90 else "VM-1A")
        p1 = secrets.token_hex(2).upper()
        p2 = secrets.token_hex(2).upper()
        lic_key = f"{prefix}-{p1}-{p2}"
        
        lic_record = {
            "id": str(uuid.uuid4()),
            "license_key": lic_key,
            "duration_days": duration_days,
            "plan_name": plan_name,
            "note": "",
            "is_used": False,
            "used_by_user_id": None,
            "used_by_name": None,
            "used_by_email": None,
            "used_at": None,
            "created_at": now_iso
        }
        licenses.append(lic_record)
        created_keys.append(lic_key)

    _save_licenses(licenses)
    return jsonify({
        "success": True,
        "created_count": len(created_keys),
        "license_keys": created_keys
    }), 201

def _require_admin():
    auth_header = request.headers.get('Authorization', '')
    token = None
    if auth_header.startswith('Bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    elif request.cookies.get('vm_auth_token'):
        token = request.cookies.get('vm_auth_token')
    elif request.args.get('token'):
        token = request.args.get('token')

    if not token:
        curr = _get_current_user()
        if curr and curr.get('role') == 'admin':
            return curr, None
        return None, (jsonify({"error": "No autenticado", "authenticated": False}), 401)

    payload = _verify_auth_token(token)
    if not payload or not payload.get('user_id'):
        return None, (jsonify({"error": "Token inválido o expirado", "authenticated": False}), 401)

    user_id = payload['user_id']
    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return None, (jsonify({"error": "Usuario no encontrado"}), 404)

    if target_user.get('role') != 'admin':
        return None, (jsonify({"error": "Acceso restringido: Se requieren privilegios de SuperAdmin"}), 403)

    return target_user, None

@app.route('/api/admin/licenses', methods=['GET'])
def admin_get_licenses():
    caller, err = _require_admin()
    if err:
        return err
    licenses = _load_licenses()
    return jsonify(licenses)

@app.route('/api/admin/users', methods=['GET'])
def admin_get_users():
    caller, err = _require_admin()
    if err:
        return err

    users = _load_users()
    enriched_users = []
    
    active_count = 0
    trial_count = 0
    expired_count = 0
    admin_count = 0

    for u in users:
        status, days_left, expires_at, plan_name = _calc_subscription_status(u)
        is_admin = u.get('role') == 'admin'
        
        if is_admin:
            admin_count += 1
            active_count += 1
        elif status == 'active' or status == 'lifetime':
            active_count += 1
        elif status == 'trial':
            trial_count += 1
        else:
            expired_count += 1

        enriched_users.append({
            "id": u.get('id'),
            "email": u.get('email'),
            "full_name": u.get('full_name'),
            "professional_title": u.get('professional_title', 'Nutricionista / Especialista BIA'),
            "clinic_name": u.get('clinic_name', 'Mi Consultorio'),
            "phone": u.get('phone', ''),
            "role": u.get('role', 'user'),
            "subscription_status": status,
            "subscription_plan": plan_name,
            "subscription_expires_at": expires_at,
            "trial_started_at": u.get('trial_started_at'),
            "created_at": u.get('created_at'),
            "days_left": days_left
        })

    enriched_users.sort(key=lambda x: (0 if x['role'] == 'admin' else 1, x.get('created_at') or ''), reverse=False)

    stats = {
        "total_users": len(users),
        "active_users": active_count,
        "trial_users": trial_count,
        "expired_users": expired_count,
        "admin_users": admin_count
    }

    return jsonify({
        "success": True,
        "users": enriched_users,
        "stats": stats
    }), 200

@app.route('/api/admin/users/create', methods=['POST'])
def admin_create_user():
    caller, err = _require_admin()
    if err:
        return err

    data = request.json or {}
    email = _clean_str(data.get('email'), max_len=120).lower().strip()
    password = str(data.get('password') or '').strip()
    full_name = _clean_str(data.get('full_name'), max_len=120)
    professional_title = _clean_str(data.get('professional_title'), max_len=100) or "Nutricionista / Especialista BIA"
    clinic_name = _clean_str(data.get('clinic_name'), max_len=150) or "Mi Consultorio Clínico"
    phone = _clean_str(data.get('phone'), max_len=30)
    role = "admin" if data.get('role') == 'admin' else "user"
    duration_days = int(data.get('duration_days', 30))
    plan_name = _clean_str(data.get('subscription_plan')) or (f"Plan Pro ({duration_days} días)" if duration_days < 365 else "Plan Anual Pro")

    if not email or '@' not in email:
        return jsonify({"error": "Correo electrónico inválido"}), 400
    if not password or len(password) < 6:
        return jsonify({"error": "La contraseña debe tener mínimo 6 caracteres"}), 400
    if not full_name:
        return jsonify({"error": "El nombre completo es obligatorio"}), 400

    users = _load_users()
    if any(u.get('email', '').lower() == email for u in users):
        return jsonify({"error": "Ya existe un usuario con este correo electrónico"}), 409

    now_utc = datetime.now(timezone.utc)
    expires_dt = now_utc + timedelta(days=duration_days)

    new_user = {
        "id": _generate_next_user_id(role=role),
        "email": email,
        "password_hash": generate_password_hash(password),
        "full_name": full_name,
        "professional_title": professional_title,
        "clinic_name": clinic_name,
        "phone": phone,
        "role": role,
        "subscription_status": "active" if role == 'user' else "lifetime",
        "subscription_plan": plan_name,
        "subscription_expires_at": expires_dt.isoformat() if role == 'user' else "2099-12-31T23:59:59Z",
        "trial_started_at": now_utc.isoformat(),
        "created_at": now_utc.isoformat()
    }

    users.append(new_user)
    _save_users(users)

    return jsonify({
        "success": True,
        "message": f"Usuario {full_name} creado exitosamente con {duration_days} días de suscripción.",
        "user": _build_safe_user_dict(new_user)
    }), 201

@app.route('/api/admin/users/<user_id>/extend', methods=['POST'])
def admin_extend_user(user_id):
    caller, err = _require_admin()
    if err:
        return err

    data = request.json or {}
    days = int(data.get('days', 30))
    plan_name = _clean_str(data.get('plan_name'))

    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    now_utc = datetime.now(timezone.utc)
    curr_exp = target_user.get('subscription_expires_at')
    base_date = now_utc

    if curr_exp:
        try:
            exp_dt = datetime.fromisoformat(curr_exp.replace('Z', '+00:00'))
            if exp_dt > now_utc:
                base_date = exp_dt
        except Exception:
            pass

    new_exp_dt = base_date + timedelta(days=days)
    target_user['subscription_expires_at'] = new_exp_dt.isoformat()
    target_user['subscription_status'] = 'active'
    if plan_name:
        target_user['subscription_plan'] = plan_name
    elif not target_user.get('subscription_plan') or any(kw in target_user.get('subscription_plan', '').lower() for kw in ('prueba', 'vencid', 'sin suscripci', 'ningun')):
        target_user['subscription_plan'] = 'Plan Pro Mensual (30 días)' if days == 30 else f"Plan Pro ({days} días)"

    _save_users(users)
    return jsonify({
        "success": True,
        "message": f"Suscripción extendida +{days} días exitosamente para {target_user.get('full_name')}.",
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/admin/users/<user_id>/status', methods=['POST'])
def admin_set_user_status(user_id):
    caller, err = _require_admin()
    if err:
        return err

    data = request.json or {}
    new_status = _clean_str(data.get('status'))
    new_plan = _clean_str(data.get('plan_name'))
    new_role = _clean_str(data.get('role'))

    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    if new_status in ['active', 'trial', 'expired', 'lifetime', 'no_subscription', 'none']:
        if new_status == 'lifetime':
            target_user['subscription_status'] = 'lifetime'
            target_user['subscription_expires_at'] = '2099-12-31T23:59:59Z'
            target_user['subscription_plan'] = new_plan or 'Plan Vitalicio Ilimitado'
        elif new_status == 'active':
            target_user['subscription_status'] = 'active'
            if not target_user.get('subscription_expires_at') or _calc_subscription_status(target_user)[0] in ('expired', 'no_subscription'):
                target_user['subscription_expires_at'] = (datetime.now(timezone.utc) + timedelta(days=30)).isoformat()
            if new_plan:
                target_user['subscription_plan'] = new_plan
        elif new_status == 'trial':
            target_user['subscription_status'] = 'trial'
            target_user['subscription_expires_at'] = (datetime.now(timezone.utc) + timedelta(days=7)).isoformat()
            target_user['subscription_plan'] = new_plan or 'Plan de Prueba (7 días)'
        elif new_status == 'expired':
            target_user['subscription_status'] = 'expired'
            target_user['subscription_expires_at'] = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
            target_user['subscription_plan'] = new_plan or 'Plan Vencido'
        elif new_status in ('no_subscription', 'none'):
            target_user['subscription_status'] = 'no_subscription'
            target_user['subscription_expires_at'] = None
            target_user['subscription_plan'] = 'Sin Suscripción Anterior'
            target_user['last_redeemed_pin'] = None

    if new_plan and new_status not in ('no_subscription', 'none', 'expired'):
        target_user['subscription_plan'] = new_plan
    if new_role in ['user', 'admin']:
        target_user['role'] = new_role

    _save_users(users)
    return jsonify({
        "success": True,
        "message": f"Estado de {target_user.get('full_name')} actualizado correctamente.",
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/admin/users/<user_id>/deactivate-subscription', methods=['POST'])
def admin_deactivate_user_subscription(user_id):
    caller, err = _require_admin()
    if err:
        return err

    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    target_user['subscription_status'] = 'expired'
    target_user['subscription_expires_at'] = (datetime.now(timezone.utc) - timedelta(days=1)).isoformat()
    target_user['subscription_plan'] = 'Plan Vencido'

    _save_users(users)
    return jsonify({
        "success": True,
        "message": f"Suscripción de {target_user.get('full_name')} desactivada (cuenta en 0 días / vencida).",
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/admin/users/<user_id>/remove-subscription', methods=['POST'])
def admin_remove_user_subscription(user_id):
    caller, err = _require_admin()
    if err:
        return err

    users = _load_users()
    target_user = next((u for u in users if u.get('id') == user_id), None)
    if not target_user:
        return jsonify({"error": "Usuario no encontrado"}), 404

    target_user['subscription_status'] = 'no_subscription'
    target_user['subscription_expires_at'] = None
    target_user['subscription_plan'] = 'Sin Suscripción Anterior'
    target_user['last_redeemed_pin'] = None

    _save_users(users)
    return jsonify({
        "success": True,
        "message": f"Suscripción de {target_user.get('full_name')} eliminada (restablecido a 'Sin Suscripción Anterior').",
        "user": _build_safe_user_dict(target_user)
    }), 200

@app.route('/api/admin/users/<user_id>', methods=['DELETE'])
def admin_delete_user(user_id):
    caller, err = _require_admin()
    if err:
        return err

    if caller.get('id') == user_id:
        return jsonify({"error": "No puedes eliminar tu propia cuenta de Administrador"}), 400

    users = _load_users()
    initial_len = len(users)
    users = [u for u in users if u.get('id') != user_id]

    if len(users) == initial_len:
        return jsonify({"error": "Usuario no encontrado"}), 404

    _save_users(users)

    if supabase:
        try:
            supabase.table('users').delete().eq('id', user_id).execute()
        except Exception as e:
            logging.warning("Error eliminando usuario en Supabase: %s", e)

    return jsonify({"success": True, "message": "Usuario eliminado exitosamente de la base de datos."})

@app.route('/api/admin/users/batch-delete', methods=['POST'])
def admin_batch_delete_users():
    caller, err = _require_admin()
    if err:
        return err

    data = request.json or {}
    user_ids = data.get('user_ids', [])
    if not isinstance(user_ids, list) or len(user_ids) == 0:
        return jsonify({"error": "Debes seleccionar al menos un usuario para eliminar"}), 400

    caller_id = caller.get('id')
    filtered_ids = [uid for uid in user_ids if uid != caller_id]
    if len(filtered_ids) == 0:
        return jsonify({"error": "No puedes eliminar tu propia cuenta de Administrador"}), 400

    users = _load_users()
    initial_len = len(users)
    users = [u for u in users if u.get('id') not in filtered_ids]
    deleted_count = initial_len - len(users)

    _save_users(users)

    if supabase:
        try:
            supabase.table('users').delete().in_('id', tuple(filtered_ids)).execute()
        except Exception as e:
            logging.warning("Error eliminando lote de usuarios en Supabase: %s", e)

    return jsonify({
        "success": True,
        "deleted_count": deleted_count,
        "message": f"Se eliminaron {deleted_count} usuario(s) exitosamente de la base de datos."
    }), 200

@app.route('/')
def index():
    return render_template('index.html')

_DASHBOARD_CACHE = {
    "data": None,
    "expires_at": 0
}

def _invalidate_dashboard_cache():
    global _DASHBOARD_CACHE
    _DASHBOARD_CACHE["data"] = None
    _DASHBOARD_CACHE["expires_at"] = 0

@app.route('/api/dashboard-stats', methods=['GET'])
def dashboard_stats():
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None
    is_admin = current_user and current_user.get('role') == 'admin'

    now_ts = time.time()
    tenant_cache_key = current_uid or 'global'
    if _DASHBOARD_CACHE["data"] and _DASHBOARD_CACHE.get("tenant") == tenant_cache_key and now_ts < _DASHBOARD_CACHE["expires_at"]:
        return jsonify(_DASHBOARD_CACHE["data"])

    if not supabase:
        return jsonify(_EMPTY_DASHBOARD)
        
    try:
        if not current_uid:
            return jsonify(_EMPTY_DASHBOARD), 200

        # Contar clientes de forma segura ante esquemas que aún no tengan user_id
        all_clients = []
        try:
            clients_res = supabase.table('clients').select('*').execute()
            all_clients = clients_res.data or []
        except Exception as ce:
            logging.warning("Consulta clients select(*) falló en dashboard_stats: %s", ce)
            try:
                clients_res = supabase.table('clients').select('id').execute()
                all_clients = clients_res.data or []
            except Exception:
                all_clients = []

        # Enriquecer con mapa de usuarios persistente
        for c in all_clients:
            if not c.get('user_id'):
                c_uid = _CLIENTS_USER_MAP.get(str(c.get('id'))) or _CLIENTS_USER_MAP.get(str(c.get('idp'))) or _CLIENTS_USER_MAP.get(str(c.get('code')))
                if c_uid:
                    c['user_id'] = c_uid

        # En producción, excluir pacientes transitorios generados por tests automáticos
        if not (app.config.get('TESTING') or getattr(app, 'testing', False)):
            all_clients = [c for c in all_clients if 'Paciente Exclusivo' not in (c.get('name') or '')]

        # Estricto aislamiento multi-tenant: SuperAdmin no ve datos clínicos de doctores
        if is_admin:
            all_clients = [c for c in all_clients if c.get('user_id') == current_uid]
        else:
            all_clients = [c for c in all_clients if c.get('user_id') == current_uid or (not c.get('user_id') and current_uid == 'usr-doctor-001')]
        total_clients = len(all_clients)
        
        # Obtener evaluaciones del usuario actual de forma aislada y persistente
        evaluations = _get_all_evaluations_for_user(current_user)
        total_evaluations = len(evaluations)
        
        # Calcular promedio de TRU Score
        valid_scores = [float(e['global_score']) for e in evaluations if e.get('global_score') is not None]
        avg_score = round(sum(valid_scores) / len(valid_scores), 1) if valid_scores else 0
        
        # Estadísticas poblacionales (Estado celular)
        cell_status_counts = {"Óptimo": 0, "Límite": 0, "Bajo": 0}
        
        recent = []
        for e in evaluations:
            r = float(e.get('resistance', 0) or 0)
            xc = float(e.get('reactance', 0) or 0)
            biva_info = get_biva_interpretation(r, xc)
            bucket = _cell_bucket(biva_info.get('phase_angle', 0), biva_info.get('valid', True))
            cell_status_counts[bucket] += 1
            
            if len(recent) < 5:
                recent.append({
                    "id": e.get('id'),
                    "name": e.get('patient_name', 'Unknown'),
                    "date": (e.get('created_at') or '').split('T')[0],
                    "score": e.get('global_score', 0),
                    "phase_angle": biva_info.get('phase_angle', 0)
                })
        
        result_payload = {
            "total_clients": total_clients,
            "total_evaluations": total_evaluations,
            "avg_score": avg_score,
            "recent": recent,
            "population": cell_status_counts
        }

        # Almacenar en caché en memoria por 15s por tenant
        _DASHBOARD_CACHE["data"] = result_payload
        _DASHBOARD_CACHE["tenant"] = tenant_cache_key
        _DASHBOARD_CACHE["expires_at"] = now_ts + 15

        return jsonify(result_payload)
    except Exception as e:
        logging.error("Error al obtener estadísticas del dashboard: %s", e, exc_info=True)
        if _DASHBOARD_CACHE["data"]:
            return jsonify(_DASHBOARD_CACHE["data"]), 200
        return jsonify(_EMPTY_DASHBOARD), 200

def _run_analysis(data):
    """
    FASE 5: Núcleo de cálculo unificado.
    Recibe el payload del formulario, ejecuta los Módulos 1-7 y devuelve UN diccionario completo.
    """
    # Campos opcionales del dispositivo
    smm = data.get('smm')
    tbw = data.get('tbw')
    ecw = data.get('ecw')
    fat_mass = data.get('fat_mass')
    visceral_fat = data.get('visceral_fat')
    waist = data.get('waist')
    phase_angle_dev = data.get('phase_angle_dev')
    seg_arm_r = data.get('seg_arm_r')
    seg_arm_l = data.get('seg_arm_l')
    seg_torso = data.get('seg_torso')
    seg_leg_r = data.get('seg_leg_r')
    seg_leg_l = data.get('seg_leg_l')
    
    # Índices directos del dispositivo (si se proporcionan)
    dev_imc = data.get('imc')
    dev_fmi = data.get('fmi')
    dev_ffmi = data.get('ffmi')
    dev_fm_pct = data.get('fm_pct')
    dev_smi = data.get('smi')

    # Validar numéricos opcionales
    def _num(v):
        try:
            return float(v) if v not in (None, "") else None
        except (ValueError, TypeError):
            return None

    smm = _num(smm)
    tbw = _num(tbw)
    ecw = _num(ecw)
    fat_mass = _num(fat_mass)
    visceral_fat = _num(visceral_fat)
    waist = _num(waist)
    phase_angle_dev = _num(phase_angle_dev)
    seg_arm_r = _num(seg_arm_r)
    seg_arm_l = _num(seg_arm_l)
    seg_torso = _num(seg_torso)
    seg_leg_r = _num(seg_leg_r)
    seg_leg_l = _num(seg_leg_l)
    dev_imc = _num(dev_imc)
    dev_fmi = _num(dev_fmi)
    dev_ffmi = _num(dev_ffmi)
    dev_fm_pct = _num(dev_fm_pct)
    dev_smi = _num(dev_smi)

    # Datos paciente sanitizados
    patient_idp = _clean_str(data.get('patient_idp', ''), max_len=50) or ''
    patient_name = _clean_str(data.get('patient_name', 'Paciente sin registrar'), max_len=100) or 'Paciente sin registrar'

    # Datos físicos sanitizados y acotados a rangos clínicos válidos
    r = max(0.0, min(_num(data.get('resistance', 0)) or 0.0, 2000.0))
    xc = max(0.0, min(_num(data.get('reactance', 0)) or 0.0, 500.0))
    weight = max(10.0, min(_num(data.get('weight', 0)) or 70.0, 350.0))
    height = max(40.0, min(_num(data.get('height', 0)) or 170.0, 250.0))
    raw_age = int(_num(data.get('age', 30)) or 30)
    age = max(1, min(raw_age, 120))
    gender = _normalize_gender(data.get('gender', 'male'))
    pal = max(1.0, min(_num(data.get('pal', 1.2)) or 1.2, 3.0))

    # Cálculos - Módulos Base
    biva_info = get_biva_interpretation(r, xc)
    energy_info = calculate_energy(weight, height, age, gender, pal, smm=smm, fat_mass=fat_mass)
    scores = calculate_scores(weight, height, biva_info['phase_angle'],
                              smm=smm, fat_mass=fat_mass, gender=gender)
    hydration_info = analyze_hydration(tbw=tbw, ecw=ecw, weight=weight)
    visceral_info = analyze_visceral_fat(waist_cm=waist, visceral_fat_l=visceral_fat, gender=gender)
    clinical_findings = build_clinical_report(
        biva_info, hydration_info, visceral_info, scores,
        biva_info['phase_angle'],
        ecw_tbw_ratio=hydration_info.get('ecw_tbw_ratio')
    )

    # Percentiles y Curvas Poblacionales
    phase_for_percentile = phase_angle_dev if phase_angle_dev else biva_info['phase_angle']
    phase_percentile = get_phase_angle_percentile(phase_for_percentile, age, gender)
    pha_curves = get_pha_age_curves(gender)
    smm_percentile = get_smm_percentile(smm, age, gender) if smm else None
    smm_curves = get_smm_age_curves(gender) if smm else None

    # Músculo Segmental
    segments = {
        'arm_right': seg_arm_r, 'arm_left': seg_arm_l,
        'torso': seg_torso, 'leg_right': seg_leg_r, 'leg_left': seg_leg_l
    }
    has_segments = any(v is not None for v in segments.values())
    segmental_info = analyze_segmental(segments, gender) if has_segments else {"segments": {}, "asymmetries": []}

    # [CLI-01 FIX] Índices de Composición: cálculo integral con override granular
    calculated_indices = analyze_composition_indices(weight, height, fat_mass, smm, gender)
    tables = load_tables()

    def _eval_imc_status(v):
        r_tab = tables.get("bmi_normal_ranges", {}).get(gender, {})
        if v < r_tab.get("low", 18.5): return "Bajo peso", "yellow"
        if v <= r_tab.get("normal_max", 24.9): return "Normal", "green"
        if v <= r_tab.get("overweight_max", 29.9): return "Sobrepeso", "yellow"
        return "Obesidad", "red"

    def _eval_high_status(v, key):
        r_tab = tables.get(key, {}).get(gender, {})
        return ("Normal", "green") if (v is not None and v <= r_tab.get("normal_max", 999)) else ("Alto", "red")

    def _eval_low_status(v, key):
        r_tab = tables.get(key, {}).get(gender, {})
        return ("Normal", "green") if (v is not None and v >= r_tab.get("normal_min", 0)) else ("Bajo", "yellow")

    final_imc = dev_imc if dev_imc is not None else calculated_indices.get("imc")
    final_fmi = dev_fmi if dev_fmi is not None else calculated_indices.get("fmi")
    final_ffmi = dev_ffmi if dev_ffmi is not None else calculated_indices.get("ffmi")
    final_fm_pct = dev_fm_pct if dev_fm_pct is not None else calculated_indices.get("fm_pct")
    final_smi = dev_smi if dev_smi is not None else calculated_indices.get("smi")

    composition_indices = {
        "available": any(x is not None for x in (final_imc, final_fmi, final_ffmi, final_fm_pct, final_smi)),
        "imc": final_imc,
        "imc_status": _eval_imc_status(final_imc) if final_imc is not None else None,
        "fmi": final_fmi,
        "fmi_status": _eval_high_status(final_fmi, "fmi_normal_ranges") if final_fmi is not None else None,
        "ffmi": final_ffmi,
        "ffmi_status": _eval_low_status(final_ffmi, "ffmi_normal_ranges") if final_ffmi is not None else None,
        "fm_pct": final_fm_pct,
        "fm_pct_status": _eval_high_status(final_fm_pct, "fm_percent_ranges") if final_fm_pct is not None else None,
        "smi": final_smi,
        "smi_status": _eval_low_status(final_smi, "smi_normal_ranges") if final_smi is not None else None,
        "from_device": any(x is not None for x in (dev_imc, dev_fmi, dev_ffmi, dev_fm_pct, dev_smi))
    }

    # BCC (gráfico grasa vs músculo)
    bcc = {"available": False}
    if fat_mass and smm and weight:
        bcc = {
            "available": True,
            "fat_pct": round(fat_mass / weight * 100, 1),
            "muscle_pct": round(smm / weight * 100, 1)
        }

    # Guardar en Supabase con reciclaje seguro de códigos EVA-XXX (solo si should_save es True)
    saved = False
    assigned_code = None
    should_save = data.get('save', True)
    if supabase and should_save:
        try:
            existing_codes = []
            try:
                evals_res = supabase.table('evaluations').select('code').execute()
                for row in (evals_res.data or []):
                    raw_c = row.get('code')
                    if raw_c and str(raw_c).startswith('EVA-'):
                        try:
                            existing_codes.append(int(raw_c.replace('EVA-', '')))
                        except ValueError:
                            pass
            except Exception:
                pass
            # --- CÁLCULO SEGURO Y ÚNICO DE IDP Y CÓDIGO ---
            existing_evals = []
            try:
                existing_evals_res = supabase.table('evaluations').select('patient_name, patient_idp').execute()
                existing_evals = existing_evals_res.data or []
            except Exception:
                pass

            name_to_idp_map = {}
            used_idp_nums = set()
            for ev in existing_evals:
                pn = (ev.get('patient_name') or '').strip().lower()
                pidp = (ev.get('patient_idp') or '').strip()
                if pidp and pidp.startswith("IDP-"):
                    try:
                        num = int(pidp.replace("IDP-", ""))
                        used_idp_nums.add(num)
                    except Exception:
                        pass
                if pn and pidp and pidp not in ('Auto-asignado', 'Auto', '1234567'):
                    if pn not in name_to_idp_map:
                        name_to_idp_map[pn] = pidp

            p_name_norm = (patient_name or '').strip().lower()
            if p_name_norm in name_to_idp_map:
                final_patient_idp = name_to_idp_map[p_name_norm]
            elif patient_idp and patient_idp not in ('Auto-asignado', 'Auto', '1234567', 'IDP-0001'):
                final_patient_idp = patient_idp
            else:
                next_idp_num = (max(used_idp_nums) + 1) if used_idp_nums else 1
                final_patient_idp = f"IDP-{next_idp_num:04d}"

            existing_codes.sort()
            next_num = 1
            for num in existing_codes:
                if num == next_num:
                    next_num += 1
                elif num > next_num:
                    break

            assigned_code = f"EVA-{next_num:03d}"

            current_user = _get_current_user()
            current_uid = current_user.get('id') if current_user else None

            insert_payload = {
                "patient_idp": final_patient_idp,
                "patient_name": patient_name,
                "resistance": r,
                "reactance": xc,
                "weight": weight,
                "height": height,
                "age": age,
                "gender": gender,
                "pal": pal,
                "global_score": scores['score'],
                "muscle_score": scores['muscle_score'],
                "fat_score": scores['fat_score'],
                "smm": smm,
                "tbw": tbw,
                "ecw": ecw,
                "fat_mass": fat_mass,
                "visceral_fat": visceral_fat,
                "waist": waist,
                "code": assigned_code
            }
            if current_uid:
                insert_payload["user_id"] = current_uid

            ins_res = None
            try:
                ins_res = supabase.table('evaluations').insert(insert_payload).execute()
            except Exception as ex1:
                logging.warning("Error al insertar evaluation en Supabase (payload completo): %s", ex1)
                try:
                    insert_fb1 = {k: v for k, v in insert_payload.items() if k != 'user_id'}
                    ins_res = supabase.table('evaluations').insert(insert_fb1).execute()
                except Exception as ex2:
                    logging.warning("Error al insertar evaluation en Supabase (sin user_id): %s", ex2)
                    insert_fb2 = {k: v for k, v in insert_payload.items() if k not in ('code', 'user_id')}
                    ins_res = supabase.table('evaluations').insert(insert_fb2).execute()

            if ins_res and ins_res.data and len(ins_res.data) > 0:
                created_eval = ins_res.data[0]
                created_id = created_eval.get('id')
                if created_id and current_uid:
                    eval_map = _load_evaluations_user_map()
                    eval_map[created_id] = {
                        "user_id": current_uid,
                        "code": assigned_code
                    }
                    _save_evaluations_user_map(eval_map)
                    saved = True

            # --- AUTO-REGISTRO / VINCULACIÓN SILENCIOSA DE PACIENTE EN CLIENTS ---
            try:
                clients_res = supabase.table('clients').select('*').execute()
                existing_clients = clients_res.data or []
                
                match_client = None
                p_idp_norm = (final_patient_idp or '').strip()

                for cl in existing_clients:
                    c_name = (cl.get('name') or '').strip().lower()
                    c_idp = (cl.get('idp') or cl.get('code') or '').strip()
                    if (p_idp_norm and c_idp == p_idp_norm) or (p_name_norm and c_name == p_name_norm):
                        match_client = cl
                        break

                gender_formatted = 'Femenino' if gender in ('female', 'Femenino') else 'Masculino'

                if not match_client:
                    c_codes = [row['code'] for row in existing_clients if row.get('code') is not None]
                    c_codes.sort()
                    new_c_code = 1
                    for cc in c_codes:
                        if cc == new_c_code:
                            new_c_code += 1
                        elif cc > new_c_code:
                            break

                    new_client_record = {
                        "code": new_c_code,
                        "name": patient_name,
                        "age": age,
                        "gender": gender_formatted,
                        "height": height,
                        "phone": None,
                        "email": None
                    }
                    if current_uid:
                        new_client_record["user_id"] = current_uid
                    try:
                        ins_res = supabase.table('clients').insert(new_client_record).execute()
                        if ins_res and ins_res.data and current_uid:
                            c_created_id = ins_res.data[0].get('id')
                            if c_created_id:
                                _CLIENTS_USER_MAP[str(c_created_id)] = current_uid
                    except Exception:
                        try:
                            nc_fallback = {k: v for k, v in new_client_record.items() if k != 'user_id'}
                            supabase.table('clients').insert(nc_fallback).execute()
                        except Exception:
                            fallback_c = {"code": new_c_code, "name": patient_name, "phone": None, "email": None}
                            supabase.table('clients').insert(fallback_c).execute()
                    if current_uid:
                        _CLIENTS_USER_MAP[str(new_c_code)] = current_uid
                        _CLIENTS_USER_MAP[str(final_patient_idp)] = current_uid
                        _save_clients_user_map(_CLIENTS_USER_MAP)
                else:
                    update_fields = {}
                    if age and match_client.get('age') != age:
                        update_fields['age'] = age
                    if height and match_client.get('height') != height:
                        update_fields['height'] = height
                    if gender_formatted and match_client.get('gender') != gender_formatted:
                        update_fields['gender'] = gender_formatted
                    if update_fields:
                        try:
                            supabase.table('clients').update(update_fields).eq('id', match_client['id']).execute()
                        except Exception:
                            pass
            except Exception as e_cl:
                logging.error("Error en auto-registro silencioso de paciente: %s", e_cl)

            saved = True
            _invalidate_dashboard_cache()
        except Exception as e:
            logging.error("Error al guardar evaluación en Supabase: %s", e)

    return {
        "score": scores['score'],
        "rank": scores['rank'],
        "muscle_score": scores['muscle_score'],
        "fat_score": scores['fat_score'],
        "phase_angle": biva_info['phase_angle'],
        "cell_status": biva_info['cell_status'],
        "hydration_status": biva_info['hydration'],
        "ree_kcal": energy_info['ree_kcal'],
        "tee_kcal": energy_info['tee_kcal'],
        "hydration": hydration_info,
        "visceral": visceral_info,
        "clinical_findings": clinical_findings,
        "phase_percentile": phase_percentile,
        "pha_curves": pha_curves,
        "smm_percentile": smm_percentile,
        "smm_curves": smm_curves,
        "segmental": segmental_info,
        "composition_indices": composition_indices,
        "bcc": bcc,
        "saved": saved,
        "inputs_echo": {
            "height": height,
            "gender": gender,
            "age": age
        }
    }


@app.route('/api/calculate', methods=['POST'])
def calculate():
    """Endpoint legacy (compatibilidad). Delega en el núcleo unificado."""
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN o renueva tu plan para realizar nuevos análisis de Bioimpedancia.",
            "subscription_expired": True
        }), 403
    data = request.json or {}
    return jsonify(_run_analysis(data))


@app.route('/api/dashboard-data', methods=['POST'])
def dashboard_data():
    """Endpoint canónico del manual (Pagina2 Analyzer.md)."""
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN o renueva tu plan para realizar nuevos análisis de Bioimpedancia.",
            "subscription_expired": True
        }), 403
    data = request.json or {}
    return jsonify(_run_analysis(data))

# --- RUTAS DE EVALUACIONES ---

@app.route('/api/evaluations', methods=['GET'])
def get_evaluations():
    if not supabase:
        return jsonify([]), 200
    try:
        current_user = _get_current_user()
        if not current_user:
            return jsonify([]), 200

        evals_asc = _get_all_evaluations_for_user(current_user)

        for idx, e in enumerate(evals_asc, start=1):
            if not e.get('code'):
                e['code'] = f"EVA-{idx:03d}"
            r = float(e.get('resistance') or 0)
            xc = float(e.get('reactance') or 0)
            biva_info = get_biva_interpretation(r, xc)
            e['phase_angle'] = biva_info['phase_angle']
            e['cell_status'] = biva_info['cell_status']
            e['hydration_status'] = biva_info['hydration']
            
        evals_asc.reverse()
        return jsonify(evals_asc)
    except Exception as e:
        logging.error("Error al obtener evaluaciones: %s", e, exc_info=True)
        return jsonify([]), 200

@app.route('/api/evaluations/<string:eval_id>', methods=['GET'])
def get_evaluation_by_id(eval_id):
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    try:
        res = supabase.table('evaluations').select('*').eq('id', eval_id).execute()
        if not res.data:
            return jsonify({"error": "Evaluación no encontrada"}), 404
        
        raw_eval = res.data[0]
        current_user = _get_current_user()
        if not current_user:
            return jsonify({"error": "No autorizado"}), 401

        user_evals = _get_all_evaluations_for_user(current_user)
        user_eval_ids = [str(e.get('id')) for e in user_evals if e.get('id')]
        
        if str(eval_id) not in user_eval_ids and current_user.get('role') != 'admin':
            return jsonify({"error": "No tienes permiso para ver esta evaluación clínica"}), 403

        payload = {
            "patient_idp": raw_eval.get('patient_idp'),
            "patient_name": raw_eval.get('patient_name'),
            "resistance": raw_eval.get('resistance'),
            "reactance": raw_eval.get('reactance'),
            "weight": raw_eval.get('weight'),
            "height": raw_eval.get('height'),
            "age": raw_eval.get('age'),
            "gender": raw_eval.get('gender'),
            "pal": raw_eval.get('pal'),
            "smm": raw_eval.get('smm'),
            "tbw": raw_eval.get('tbw'),
            "ecw": raw_eval.get('ecw'),
            "fat_mass": raw_eval.get('fat_mass'),
            "visceral_fat": raw_eval.get('visceral_fat'),
            "waist": raw_eval.get('waist'),
            "save": False
        }
        full_analysis = _run_analysis(payload)
        full_analysis["id"] = raw_eval.get("id")
        full_analysis["code"] = raw_eval.get("code")
        full_analysis["created_at"] = raw_eval.get("created_at")
        full_analysis["patient_idp"] = raw_eval.get("patient_idp")
        full_analysis["patient_name"] = raw_eval.get("patient_name")
        full_analysis["raw_inputs"] = payload
        return jsonify(full_analysis)
    except Exception as e:
        logging.error("Error al obtener detalle de evaluación: %s", e, exc_info=True)
        return jsonify({"error": "Error interno al recuperar la evaluación"}), 500

@app.route('/api/evaluations/<string:eval_id>', methods=['DELETE'])
def delete_evaluation(eval_id):
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    try:
        current_user = _get_current_user()
        if not current_user:
            return jsonify({"error": "No autorizado"}), 401
        
        # Validar pertenencia con resolutor multi-fuente
        user_evals = _get_all_evaluations_for_user(current_user)
        user_eval_ids = [str(e.get('id')) for e in user_evals if e.get('id')]

        if str(eval_id) not in user_eval_ids and current_user.get('role') != 'admin':
            return jsonify({"error": "No tienes permiso para eliminar esta evaluación"}), 403

        supabase.table('evaluations').delete().eq('id', eval_id).execute()
        if str(eval_id) in _EVALUATIONS_USER_MAP:
            del _EVALUATIONS_USER_MAP[str(eval_id)]
            _save_evaluations_user_map(_EVALUATIONS_USER_MAP)

        _invalidate_dashboard_cache()
        return jsonify({"success": True})
    except Exception as e:
        logging.error("Error al eliminar evaluación: %s", e, exc_info=True)
        return jsonify({"error": "Error al eliminar la evaluación"}), 500

@app.route('/api/evaluations/batch-delete', methods=['POST'])
def batch_delete_evaluations():
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    try:
        current_user = _get_current_user()
        if not current_user:
            return jsonify({"error": "No autorizado"}), 401

        payload = request.get_json() or {}
        raw_eval_ids = payload.get('ids', [])
        if not raw_eval_ids or not isinstance(raw_eval_ids, list):
            return jsonify({"error": "No se especificaron IDs válidos para eliminar"}), 400

        user_evals = _get_all_evaluations_for_user(current_user)
        user_eval_ids_set = set(str(e.get('id')) for e in user_evals if e.get('id'))
        
        # SuperAdmin elimina todos los seleccionados; Doctor elimina solo los suyos
        is_admin = (current_user.get('role') == 'admin')
        allowed_ids = [str(i) for i in raw_eval_ids if is_admin or str(i) in user_eval_ids_set]

        if not allowed_ids:
            return jsonify({"error": "No tienes permiso para eliminar las evaluaciones seleccionadas"}), 403

        supabase.table('evaluations').delete().in_('id', allowed_ids).execute()
        for e_id in allowed_ids:
            if e_id in _EVALUATIONS_USER_MAP:
                del _EVALUATIONS_USER_MAP[e_id]
        _save_evaluations_user_map(_EVALUATIONS_USER_MAP)

        _invalidate_dashboard_cache()
        return jsonify({"success": True, "deleted_count": len(allowed_ids)})
    except Exception as e:
        logging.error("Error en batch delete de evaluaciones: %s", e, exc_info=True)
        return jsonify({"error": f"Error al eliminar lote de evaluaciones: {str(e)}"}), 500

# --- PERSISTENCIA Y ASOCIACIÓN DE EVALUACIONES POR USUARIO ---

_EVALUATIONS_USER_MAP_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "evaluations_users.json")

def _load_evaluations_user_map():
    if os.path.exists(_EVALUATIONS_USER_MAP_PATH):
        try:
            with open(_EVALUATIONS_USER_MAP_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_evaluations_user_map(m):
    global _EVALUATIONS_USER_MAP
    _EVALUATIONS_USER_MAP = m
    try:
        os.makedirs(os.path.dirname(_EVALUATIONS_USER_MAP_PATH), exist_ok=True)
        with open(_EVALUATIONS_USER_MAP_PATH, 'w', encoding='utf-8') as f:
            json.dump(m, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

_EVALUATIONS_USER_MAP = _load_evaluations_user_map()

def _get_all_evaluations_for_user(current_user):
    if not supabase or not current_user:
        return []
    try:
        current_uid = current_user.get('id')
        if not current_uid:
            return []

        # 1. Obtener todas las evaluaciones almacenadas en Supabase
        raw_evals = []
        try:
            evals_res = supabase.table('evaluations').select('*').order('created_at', desc=False).execute()
            raw_evals = evals_res.data or []
        except Exception as ee:
            logging.warning("Error al consultar tabla evaluations en Supabase: %s", ee)

        # 2. Cargar mapa de clientes en Supabase (idp -> user_id, name -> user_id)
        clients_by_idp = {}
        clients_by_name = {}
        try:
            cls_res = supabase.table('clients').select('idp, name, user_id').execute()
            for c in (cls_res.data or []):
                c_uid = c.get('user_id') or _CLIENTS_USER_MAP.get(str(c.get('id'))) or _CLIENTS_USER_MAP.get(str(c.get('idp')))
                if c_uid:
                    c_idp = (c.get('idp') or '').strip()
                    c_name = (c.get('name') or '').strip().lower()
                    if c_idp:
                        clients_by_idp[c_idp] = str(c_uid)
                    if c_name:
                        clients_by_name[c_name] = str(c_uid)
        except Exception as e_cls:
            logging.warning("Error al cargar clientes para vinculación de evaluaciones: %s", e_cls)

        # 3. Resolver user_id para cada evaluación
        resolved_evals = []
        for e in raw_evals:
            e_id = str(e.get('id')) if e.get('id') else None
            assigned_uid = e.get('user_id')

            # Normalizar user_id si viene directo de Supabase
            if assigned_uid and str(assigned_uid).strip() and str(assigned_uid) != 'None':
                assigned_uid = str(assigned_uid).strip()
            else:
                assigned_uid = None

            # Fallback 1: Mapa en memoria / disco de evaluaciones
            if not assigned_uid and e_id and e_id in _EVALUATIONS_USER_MAP:
                map_item = _EVALUATIONS_USER_MAP[e_id]
                if isinstance(map_item, dict):
                    assigned_uid = map_item.get('user_id')
                elif isinstance(map_item, str):
                    assigned_uid = map_item

            # Fallback 2: Coincidencia por IDP del paciente en la tabla clients
            p_idp = (e.get('patient_idp') or '').strip()
            if not assigned_uid and p_idp:
                if p_idp in clients_by_idp:
                    assigned_uid = clients_by_idp[p_idp]
                elif p_idp in _CLIENTS_USER_MAP:
                    assigned_uid = _CLIENTS_USER_MAP[p_idp]

            # Fallback 3: Coincidencia por Nombre del paciente en la tabla clients
            p_name = (e.get('patient_name') or '').strip().lower()
            if not assigned_uid and p_name:
                if p_name in clients_by_name:
                    assigned_uid = clients_by_name[p_name]
                elif p_name in _CLIENTS_USER_MAP:
                    assigned_uid = _CLIENTS_USER_MAP[p_name]

            # Fallback 4: Retrocompatibilidad con usuario demo inicial si está huérfano
            if not assigned_uid and current_uid == 'usr-doctor-001':
                assigned_uid = 'usr-doctor-001'

            e['user_id'] = assigned_uid
            resolved_evals.append(e)

        # 4. Aislamiento multi-tenant por doctor / usuario con auto-recuperación
        is_admin = (current_user.get('role') == 'admin')
        if is_admin:
            filtered = [e for e in resolved_evals if e.get('user_id') == current_uid or (not e.get('user_id') and current_uid == 'usr-admin-001')]
        else:
            # Doctor ve sus evaluaciones o evaluaciones legacy iniciales de muestra
            filtered = [e for e in resolved_evals if e.get('user_id') == current_uid or not e.get('user_id') or e.get('user_id') in ('usr-doctor-001', 'None', 'null', '')]
            # Auto-actualizar Supabase para sincronizar propiedad al doctor activo
            for e in filtered:
                if e.get('id') and e.get('user_id') != current_uid:
                    e['user_id'] = current_uid
                    try:
                        supabase.table('evaluations').update({'user_id': current_uid}).eq('id', e['id']).execute()
                    except Exception:
                        pass

        return filtered
    except Exception as ex:
        logging.error("Error crítico al recuperar evaluaciones del usuario: %s", ex, exc_info=True)
        return []

# --- RUTAS DE CLIENTES ---

_CLIENTS_USER_MAP_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "clients_users.json")

def _load_clients_user_map():
    if os.path.exists(_CLIENTS_USER_MAP_PATH):
        try:
            with open(_CLIENTS_USER_MAP_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _save_clients_user_map(m):
    try:
        os.makedirs(os.path.dirname(_CLIENTS_USER_MAP_PATH), exist_ok=True)
        with open(_CLIENTS_USER_MAP_PATH, 'w', encoding='utf-8') as f:
            json.dump(m, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

_CLIENTS_USER_MAP = _load_clients_user_map()

@app.route('/api/clients', methods=['GET'])
def get_clients():
    if not supabase:
        return jsonify([]), 200
    try:
        current_user = _get_current_user()
        current_uid = current_user.get('id') if current_user else None
        if not current_uid:
            return jsonify([]), 200

        try:
            res = supabase.table('clients').select('*').execute()
            clients = res.data or []
        except Exception as e_supa:
            logging.warning("Error consultando clientes en Supabase: %s", e_supa)
            clients = []

        # Enriquecer con user_id desde el mapa de aislamiento
        for c in clients:
            if not c.get('user_id'):
                c_uid = _CLIENTS_USER_MAP.get(str(c.get('id'))) or _CLIENTS_USER_MAP.get(str(c.get('idp'))) or _CLIENTS_USER_MAP.get(str(c.get('code')))
                if c_uid:
                    c['user_id'] = c_uid

        # En producción, excluir pacientes transitorios generados por tests automáticos
        if not (app.config.get('TESTING') or getattr(app, 'testing', False)):
            clients = [c for c in clients if 'Paciente Exclusivo' not in (c.get('name') or '')]

        # Aislamiento multi-tenant estricto: SuperAdmin no ve pacientes de doctores
        if current_user and current_user.get('role') == 'admin':
            clients = [c for c in clients if c.get('user_id') == current_uid]
        else:
            filtered_clients = []
            for c in clients:
                c_uid = c.get('user_id')
                if not c_uid or c_uid in ('usr-doctor-001', 'None', 'null', ''):
                    c['user_id'] = current_uid
                    c_uid = current_uid
                    if supabase and c.get('id'):
                        try:
                            supabase.table('clients').update({'user_id': current_uid}).eq('id', c['id']).execute()
                        except Exception:
                            pass
                if c_uid == current_uid:
                    filtered_clients.append(c)
            clients = filtered_clients

        # AUTO-SÍNTESIS: Extraer automáticamente pacientes únicos registrados desde Evaluaciones
        try:
            user_evals = _get_all_evaluations_for_user(current_user)
            existing_names = set(_clean_str(c.get('name') or '', max_len=100).lower() for c in clients)
            existing_idps = set(_clean_str(c.get('idp') or '', max_len=50) for c in clients if c.get('idp'))

            for ev in user_evals:
                ev_name = _clean_str(ev.get('patient_name') or '', max_len=100)
                ev_idp = _clean_str(ev.get('patient_idp') or '', max_len=50)
                ev_name_norm = ev_name.lower()

                if ev_name and ev_name_norm not in existing_names and (not ev_idp or ev_idp not in existing_idps):
                    code_num = 1
                    if ev_idp and ev_idp.startswith("IDP-"):
                        try:
                            code_num = int(ev_idp.replace("IDP-", ""))
                        except Exception:
                            code_num = len(clients) + 1

                    synth = {
                        "id": f"synth-{ev.get('id')}",
                        "code": code_num,
                        "idp": ev_idp or f"IDP-{code_num:04d}",
                        "name": ev_name,
                        "age": ev.get('age'),
                        "gender": ev.get('gender'),
                        "height": ev.get('height'),
                        "user_id": current_uid
                    }
                    clients.append(synth)
                    existing_names.add(ev_name_norm)
                    if ev_idp:
                        existing_idps.add(ev_idp)
        except Exception as ex_synth:
            logging.warning("Error al auto-sintetizar pacientes desde evaluaciones: %s", ex_synth)

        # Ordenar de forma segura por código en memoria
        def _sort_code_key(c):
            cd = c.get('code')
            if isinstance(cd, int):
                return cd
            if isinstance(cd, str) and cd.isdigit():
                return int(cd)
            return 999999

        clients.sort(key=lambda x: (_sort_code_key(x), x.get('name') or ''))

        # Adjuntar resumen de última evaluación perteneciente al mismo usuario
        try:
            evals = _get_all_evaluations_for_user(current_user)
            eval_by_name = {}
            eval_by_idp = {}
            for ev in evals:
                p_name = _clean_str(ev.get('patient_name') or '', max_len=100).lower()
                p_idp = _clean_str(ev.get('patient_idp') or '', max_len=50)
                if p_name and p_name not in eval_by_name:
                    eval_by_name[p_name] = ev
                if p_idp and p_idp not in eval_by_idp:
                    eval_by_idp[p_idp] = ev

            for c in clients:
                c_name = _clean_str(c.get('name') or '', max_len=100).lower()
                c_idp = _clean_str(c.get('idp') or '', max_len=50)
                matched_ev = eval_by_idp.get(c_idp) if c_idp else None
                if not matched_ev and c_name:
                    matched_ev = eval_by_name.get(c_name)
                
                if matched_ev:
                    try:
                        r = float(matched_ev.get('resistance') or 0)
                        xc = float(matched_ev.get('reactance') or 0)
                        biva = get_biva_interpretation(r, xc)
                        c['last_eval'] = {
                            'code': matched_ev.get('code'),
                            'date': matched_ev.get('created_at'),
                            'phase_angle': biva.get('phase_angle'),
                            'hydration': biva.get('hydration'),
                            'cell_status': biva.get('cell_status')
                        }
                    except Exception:
                        pass
        except Exception as e_ev:
            logging.warning("Error asociando evaluaciones a clientes: %s", e_ev)

        return jsonify(clients)
    except Exception as e:
        logging.error("Error al obtener clientes: %s", e, exc_info=True)
        return jsonify([]), 200

@app.route('/api/clients/next-code', methods=['GET'])
def get_next_client_code():
    if not supabase:
        return jsonify({"code": 1, "idp": "IDP-0001"}), 200
    try:
        res = supabase.table('clients').select('code').execute()
        codes = [row['code'] for row in (res.data or []) if row.get('code') is not None]
        codes.sort()
        
        new_code = 1
        for code in codes:
            if code == new_code:
                new_code += 1
            elif code > new_code:
                break
                
        return jsonify({
            "code": new_code,
            "idp": f"IDP-{new_code:04d}"
        }), 200
    except Exception as e:
        logging.error("Error al calcular siguiente IDP: %s", e)
        return jsonify({"code": 1, "idp": "IDP-0001"}), 200

@app.route('/api/clients', methods=['POST'])
def add_client():
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=100)
    phone = _clean_str(data.get('phone'), max_len=30)
    email = _clean_str(data.get('email'), max_len=100)
    idp = _clean_str(data.get('idp'), max_len=50)
    gender = data.get('gender') or None
    
    age = None
    if data.get('age'):
        try:
            age = int(data.get('age'))
        except (ValueError, TypeError):
            pass

    height = None
    if data.get('height'):
        try:
            height = float(data.get('height'))
        except (ValueError, TypeError):
            pass
    
    if not name:
        return jsonify({"error": "El nombre es obligatorio"}), 400
        
    try:
        current_user = _get_current_user()
        if not _is_subscription_active(current_user):
            return jsonify({
                "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para registrar nuevos pacientes.",
                "subscription_expired": True
            }), 403

        current_uid = current_user.get('id') if current_user else None
        if not current_uid:
            return jsonify({"error": "No autorizado"}), 401

        # Lógica de asignación de código único global en Supabase
        try:
            res = supabase.table('clients').select('code').execute()
            codes = [row['code'] for row in (res.data or []) if row.get('code') is not None]
        except Exception:
            codes = []

        codes.sort()
        new_code = 1
        for code in codes:
            if code == new_code:
                new_code += 1
            elif code > new_code:
                break
                
        if not idp or idp in ("Auto-asignado", "Auto", "Auto-Asignado"):
            idp = f"IDP-{new_code:04d}"

        new_client = {
            "code": new_code,
            "name": name,
            "phone": phone,
            "email": email,
            "idp": idp,
            "age": age,
            "gender": gender,
            "height": height,
            "user_id": current_uid
        }
            
        res_data = dict(new_client)
        try:
            res_insert = supabase.table('clients').insert(new_client).execute()
            if res_insert and res_insert.data:
                res_data = res_insert.data[0]
        except Exception as e_ins:
            logging.warning("Error al insertar cliente con user_id, intentando fallback: %s", e_ins)
            try:
                nc_copy = dict(new_client)
                nc_copy.pop('user_id', None)
                res_insert = supabase.table('clients').insert(nc_copy).execute()
                if res_insert and res_insert.data:
                    res_data = res_insert.data[0]
            except Exception:
                try:
                    fallback_client = {"code": new_code, "name": name, "phone": phone, "email": email}
                    res_insert = supabase.table('clients').insert(fallback_client).execute()
                    if res_insert and res_insert.data:
                        res_data = res_insert.data[0]
                except Exception as e_fb:
                    logging.warning("No se pudo insertar en Supabase clients: %s", e_fb)
        
        # Registrar propiedad del cliente en el mapa
        if res_data.get('id'):
            _CLIENTS_USER_MAP[str(res_data.get('id'))] = current_uid
        if res_data.get('idp'):
            _CLIENTS_USER_MAP[str(res_data.get('idp'))] = current_uid
        if res_data.get('code'):
            _CLIENTS_USER_MAP[str(res_data.get('code'))] = current_uid
        _save_clients_user_map(_CLIENTS_USER_MAP)

        _invalidate_dashboard_cache()
        return jsonify({"success": True, "data": res_data}), 200
    except Exception as e:
        logging.error("Error al registrar cliente: %s", e, exc_info=True)
        return jsonify({"error": "Error al guardar cliente"}), 500

@app.route('/api/clients/<string:client_id>', methods=['PUT'])
def update_client(client_id):
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=100)
    phone = _clean_str(data.get('phone'), max_len=30)
    email = _clean_str(data.get('email'), max_len=100)
    idp = _clean_str(data.get('idp'), max_len=50)
    gender = data.get('gender') or None

    age = None
    if data.get('age'):
        try:
            age = int(data.get('age'))
        except (ValueError, TypeError):
            pass

    height = None
    if data.get('height'):
        try:
            height = float(data.get('height'))
        except (ValueError, TypeError):
            pass
    
    if not name:
        return jsonify({"error": "El nombre es obligatorio"}), 400
        
    try:
        current_user = _get_current_user()
        if not _is_subscription_active(current_user):
            return jsonify({
                "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para modificar la ficha de pacientes.",
                "subscription_expired": True
            }), 403
        current_uid = current_user.get('id') if current_user else None
        is_admin = current_user and current_user.get('role') == 'admin'
        
        # Validar pertenencia antes de actualizar
        check = None
        try:
            check = supabase.table('clients').select('id, code, name').eq('id', client_id).execute()
        except Exception:
            pass

        if not (check and check.data) and client_id.isdigit():
            try:
                check = supabase.table('clients').select('id, code, name').eq('code', int(client_id)).execute()
            except Exception:
                pass

        if check and check.data:
            client_row = check.data[0]
            real_id = client_row.get('id')
            c_code = str(client_row.get('code') or '')
            c_uid = _CLIENTS_USER_MAP.get(str(real_id)) or _CLIENTS_USER_MAP.get(c_code)
            if not is_admin and c_uid and current_uid and c_uid != current_uid:
                return jsonify({"error": "No tienes permiso para modificar este paciente"}), 403

        updated_data = {
            "name": name,
            "phone": phone,
            "email": email,
            "idp": idp or None,
            "age": age,
            "gender": gender,
            "height": height
        }
        try:
            res = supabase.table('clients').update(updated_data).eq('id', client_id).execute()
        except Exception:
            fallback_updated = {"name": name, "phone": phone, "email": email}
            res = supabase.table('clients').update(fallback_updated).eq('id', client_id).execute()
            
        _invalidate_dashboard_cache()
        return jsonify({"success": True, "data": res.data[0] if res.data else {}})
    except Exception as e:
        logging.error("Error al actualizar cliente: %s", e, exc_info=True)
        return jsonify({"error": "Error al actualizar cliente"}), 500

@app.route('/api/clients/<string:client_id>', methods=['DELETE'])
def delete_client(client_id):
    if not supabase:
        return jsonify({"error": "Base de datos no configurada"}), 503
    try:
        current_user = _get_current_user()
        if not _is_subscription_active(current_user):
            return jsonify({
                "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar pacientes.",
                "subscription_expired": True
            }), 403
        current_uid = current_user.get('id') if current_user else None
        is_admin = current_user and current_user.get('role') == 'admin'
        
        # Validar pertenencia antes de eliminar (soportando id UUID o code entero)
        check = None
        try:
            check = supabase.table('clients').select('id, code, name').eq('id', client_id).execute()
        except Exception:
            pass

        if not (check and check.data) and client_id.isdigit():
            try:
                check = supabase.table('clients').select('id, code, name').eq('code', int(client_id)).execute()
            except Exception:
                pass

        if check and check.data:
            client_row = check.data[0]
            real_id = client_row.get('id')
            c_code = str(client_row.get('code') or '')
            c_uid = _CLIENTS_USER_MAP.get(str(real_id)) or _CLIENTS_USER_MAP.get(c_code)

            # Si no es admin y el paciente pertenece a otro doctor, prohibir
            if not is_admin and c_uid and current_uid and c_uid != current_uid:
                return jsonify({"error": "No tienes permiso para eliminar este paciente"}), 403

            supabase.table('clients').delete().eq('id', real_id).execute()
            if str(real_id) in _CLIENTS_USER_MAP:
                del _CLIENTS_USER_MAP[str(real_id)]
            if c_code in _CLIENTS_USER_MAP:
                del _CLIENTS_USER_MAP[c_code]
            _save_clients_user_map(_CLIENTS_USER_MAP)
        else:
            supabase.table('clients').delete().eq('id', client_id).execute()

        _invalidate_dashboard_cache()
        return jsonify({"success": True}), 200
    except Exception as e:
        logging.error("Error al eliminar cliente: %s", e, exc_info=True)
        return jsonify({"error": "Error al eliminar cliente"}), 500

# --- RUTAS DE CITAS Y AGENDA CLÍNICA (CON PERSISTENCIA LOCAL) ---

_APPOINTMENTS_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "appointments.json")

def _load_persisted_appointments():
    if os.path.exists(_APPOINTMENTS_PATH):
        try:
            with open(_APPOINTMENTS_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.warning("Error al leer appointments.json: %s", e)
    return []

def _save_persisted_appointments(appts):
    try:
        os.makedirs(os.path.dirname(_APPOINTMENTS_PATH), exist_ok=True)
        with open(_APPOINTMENTS_PATH, 'w', encoding='utf-8') as f:
            json.dump(appts, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar appointments.json: %s", e)
        return False

_LOCAL_APPOINTMENTS = _load_persisted_appointments()

@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None
    date_filter = request.args.get('date')

    if not current_uid:
        return jsonify([]), 200

    if not supabase:
        results = _load_persisted_appointments() or _LOCAL_APPOINTMENTS
        results = [a for a in results if a.get('user_id') == current_uid]
        if date_filter:
            results = [a for a in results if a.get('date') == date_filter]
        return jsonify(results), 200

    try:
        query = supabase.table('appointments').select('*').order('date').order('time')
        if date_filter:
            query = query.eq('date', date_filter)
        res = query.execute()
        results = res.data or []
        results = [a for a in results if a.get('user_id') == current_uid]
        return jsonify(results), 200
    except Exception as e:
        logging.warning("Tabla appointments no disponible en Supabase, usando almacenamiento local: %s", e)
        results = _load_persisted_appointments() or _LOCAL_APPOINTMENTS
        results = [a for a in results if a.get('user_id') == current_uid]
        if date_filter:
            results = [a for a in results if a.get('date') == date_filter]
        return jsonify(results), 200

@app.route('/api/appointments', methods=['POST'])
def create_appointment():
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para agendar citas.",
            "subscription_expired": True
        }), 403
    current_uid = current_user.get('id') if current_user else None

    data = request.json or {}
    patient_name = _clean_str(data.get('patient_name'), max_len=100)
    patient_phone = _clean_str(data.get('patient_phone'), max_len=30)
    patient_idp = _clean_str(data.get('patient_idp'), max_len=50)
    appt_date = _clean_str(data.get('date'), max_len=20)
    appt_time = _clean_str(data.get('time'), max_len=10)
    appt_type = _clean_str(data.get('type') or 'Evaluación Inicial BIA', max_len=50)
    appt_status = _clean_str(data.get('status') or 'confirmed', max_len=20)
    notes = _clean_str(data.get('notes'), max_len=250)

    if not patient_name or not appt_date or not appt_time:
        return jsonify({"error": "Paciente, fecha y hora son obligatorios"}), 400

    new_appt = {
        "id": f"apt_{len(_LOCAL_APPOINTMENTS) + 1}_{int(time.time() if 'time' in globals() else 1000)}",
        "user_id": current_uid,
        "patient_name": patient_name,
        "patient_phone": patient_phone,
        "patient_idp": patient_idp,
        "date": appt_date,
        "time": appt_time,
        "type": appt_type,
        "status": appt_status,
        "notes": notes,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    if supabase:
        try:
            res = supabase.table('appointments').insert({
                "patient_name": patient_name,
                "patient_phone": patient_phone,
                "patient_idp": patient_idp,
                "date": appt_date,
                "time": appt_time,
                "type": appt_type,
                "status": appt_status,
                "notes": notes
            }).execute()
            if res.data:
                return jsonify({"success": True, "data": res.data[0]}), 201
        except Exception as e:
            logging.warning("No se pudo insertar en Supabase appointments, usando fallback local: %s", e)

    _LOCAL_APPOINTMENTS.append(new_appt)
    _save_persisted_appointments(_LOCAL_APPOINTMENTS)
    return jsonify({"success": True, "data": new_appt}), 201

@app.route('/api/appointments/<string:appt_id>', methods=['PUT'])
def update_appointment(appt_id):
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para modificar citas.",
            "subscription_expired": True
        }), 403
    data = request.json or {}
    updated = {
        "patient_name": _clean_str(data.get('patient_name')),
        "patient_phone": _clean_str(data.get('patient_phone')),
        "patient_idp": _clean_str(data.get('patient_idp')),
        "date": _clean_str(data.get('date')),
        "time": _clean_str(data.get('time')),
        "type": _clean_str(data.get('type')),
        "status": _clean_str(data.get('status')),
        "notes": _clean_str(data.get('notes'))
    }
    # Remover campos vacíos
    updated = {k: v for k, v in updated.items() if v}

    if supabase:
        try:
            res = supabase.table('appointments').update(updated).eq('id', appt_id).execute()
            if res.data:
                return jsonify({"success": True, "data": res.data[0]})
        except Exception as e:
            logging.warning("Error al actualizar en Supabase appointments: %s", e)

    for item in _LOCAL_APPOINTMENTS:
        if item.get('id') == appt_id:
            item.update(updated)
            _save_persisted_appointments(_LOCAL_APPOINTMENTS)
            return jsonify({"success": True, "data": item})

    return jsonify({"success": True, "message": "Actualizado"})

@app.route('/api/appointments/<string:appt_id>', methods=['DELETE'])
def delete_appointment(appt_id):
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar citas.",
            "subscription_expired": True
        }), 403
    if supabase:
        try:
            supabase.table('appointments').delete().eq('id', appt_id).execute()
            return jsonify({"success": True})
        except Exception as e:
            logging.warning("Error al eliminar en Supabase appointments: %s", e)

    global _LOCAL_APPOINTMENTS
    _LOCAL_APPOINTMENTS = [a for a in _LOCAL_APPOINTMENTS if a.get('id') != appt_id]
    _save_persisted_appointments(_LOCAL_APPOINTMENTS)
    return jsonify({"success": True})

# --- STOCK CONTROL & INVENTARIO CLÍNICO (PERSISTENCIA TOTAL EN DISCO) ---

_STOCK_ITEMS_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "stock_items.json")
_STOCK_MOVEMENTS_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "stock_movements.json")

_DEFAULT_INITIAL_STOCK_ITEMS = [
    {
        "id": "stk-001",
        "code": "SKU-BIA-001",
        "name": "Electrodos BIA Desechables (Pack x100)",
        "category": "Insumos BIA",
        "unit": "Pack",
        "stock_quantity": 48,
        "min_stock": 15,
        "cost_price": 18.50,
        "sale_price": 28.00,
        "supplier": "BIA Medical Supplies",
        "location": "Gabinete 1 - Estante A",
        "notes": "Parches adhesivos con hidrogel de baja impedancia",
        "created_at": "2026-08-01T10:00:00Z"
    },
    {
        "id": "stk-002",
        "code": "SKU-BIA-002",
        "name": "Gel Conductor BIA Hipoalergénico 250ml",
        "category": "Insumos BIA",
        "unit": "Frasco",
        "stock_quantity": 8,
        "min_stock": 10,
        "cost_price": 12.00,
        "sale_price": 19.50,
        "supplier": "Lab FarmaBolivia",
        "location": "Gabinete 1 - Estante B",
        "notes": "Alta conductividad para mediciones en piel seca",
        "created_at": "2026-08-05T11:30:00Z"
    },
    {
        "id": "stk-003",
        "code": "SKU-SUP-001",
        "name": "Proteína Whey Isolate 100% (Bote 900g Vainilla)",
        "category": "Suplementos Nutricionales",
        "unit": "Frasco",
        "stock_quantity": 14,
        "min_stock": 5,
        "cost_price": 45.00,
        "sale_price": 68.00,
        "supplier": "NutriFit Import",
        "location": "Mostrador Recepción - Vitrina",
        "notes": "Recomendada para pacientes en ganancia de masa muscular (SMM)",
        "created_at": "2026-08-10T14:00:00Z"
    },
    {
        "id": "stk-004",
        "code": "SKU-SUP-002",
        "name": "Creatina Monohidratada Creapure 300g",
        "category": "Suplementos Nutricionales",
        "unit": "Frasco",
        "stock_quantity": 3,
        "min_stock": 6,
        "cost_price": 24.00,
        "sale_price": 38.00,
        "supplier": "NutriFit Import",
        "location": "Mostrador Recepción - Vitrina",
        "notes": "Suplementación clínica celular y fuerza muscular",
        "created_at": "2026-08-12T16:20:00Z"
    },
    {
        "id": "stk-005",
        "code": "SKU-MED-001",
        "name": "Toallitas con Alcohol Isopropílico 70% (Caja x200)",
        "category": "Material Clínico e Higiene",
        "unit": "Caja",
        "stock_quantity": 22,
        "min_stock": 8,
        "cost_price": 8.50,
        "sale_price": 14.00,
        "supplier": "Droguería Médica Santa Cruz",
        "location": "Mesa de Consulta BIA",
        "notes": "Limpieza de puntos de contacto cutáneos antes de la medición",
        "created_at": "2026-08-15T09:00:00Z"
    },
    {
        "id": "stk-006",
        "code": "SKU-ACC-001",
        "name": "Cinta Métrica Antropométrica Ergonómica Seca 201",
        "category": "Accesorios y Equipos",
        "unit": "Unidad",
        "stock_quantity": 5,
        "min_stock": 2,
        "cost_price": 22.00,
        "sale_price": 35.00,
        "supplier": "Equipos Médicos La Paz",
        "location": "Gabinete 2 - Cajón Superior",
        "notes": "Medición de circunferencia de cintura y cadera",
        "created_at": "2026-08-18T12:00:00Z"
    }
]

_DEFAULT_INITIAL_STOCK_MOVEMENTS = [
    {
        "id": "mov-001",
        "stock_item_id": "stk-001",
        "item_name": "Electrodos BIA Desechables (Pack x100)",
        "type": "IN",
        "quantity": 50,
        "previous_quantity": 0,
        "new_quantity": 50,
        "reason": "Compra inicial de insumos de bioimpedancia",
        "created_at": "2026-08-01T10:05:00Z"
    },
    {
        "id": "mov-002",
        "stock_item_id": "stk-001",
        "item_name": "Electrodos BIA Desechables (Pack x100)",
        "type": "OUT",
        "quantity": 2,
        "previous_quantity": 50,
        "new_quantity": 48,
        "reason": "Consumo en consultas clínicas de la semana",
        "created_at": "2026-08-20T17:30:00Z"
    }
]

def _load_persisted_stock_items():
    if os.path.exists(_STOCK_ITEMS_PATH):
        try:
            with open(_STOCK_ITEMS_PATH, 'r', encoding='utf-8') as f:
                items = json.load(f)
                if isinstance(items, list) and len(items) > 0:
                    return items
        except Exception as e:
            logging.warning("Error al leer stock_items.json: %s", e)
    _save_persisted_stock_items(_DEFAULT_INITIAL_STOCK_ITEMS)
    return list(_DEFAULT_INITIAL_STOCK_ITEMS)

def _save_persisted_stock_items(items):
    try:
        os.makedirs(os.path.dirname(_STOCK_ITEMS_PATH), exist_ok=True)
        with open(_STOCK_ITEMS_PATH, 'w', encoding='utf-8') as f:
            json.dump(items, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar stock_items.json: %s", e)
        return False

def _load_persisted_stock_movements():
    if os.path.exists(_STOCK_MOVEMENTS_PATH):
        try:
            with open(_STOCK_MOVEMENTS_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.warning("Error al leer stock_movements.json: %s", e)
    _save_persisted_stock_movements(_DEFAULT_INITIAL_STOCK_MOVEMENTS)
    return list(_DEFAULT_INITIAL_STOCK_MOVEMENTS)

def _save_persisted_stock_movements(movements):
    try:
        os.makedirs(os.path.dirname(_STOCK_MOVEMENTS_PATH), exist_ok=True)
        with open(_STOCK_MOVEMENTS_PATH, 'w', encoding='utf-8') as f:
            json.dump(movements, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar stock_movements.json: %s", e)
        return False

_LOCAL_STOCK_ITEMS = _load_persisted_stock_items()
_LOCAL_STOCK_MOVEMENTS = _load_persisted_stock_movements()

def _safe_stock_float(val, default=0.0, min_val=None):
    if val is None:
        return default
    if isinstance(val, (int, float)):
        if math.isnan(val) or math.isinf(val):
            return default
        f = float(val)
        if min_val is not None:
            f = max(min_val, f)
        return round(f, 2)
    
    s = str(val).strip()
    if not s:
        return default
    
    # Limpiar símbolos de moneda y espacios
    for prefix in ('bs', 'bs.', 'bs', '$', '€', 'usd', 'bob'):
        if s.lower().startswith(prefix):
            s = s[len(prefix):].strip()
        if s.lower().endswith(prefix):
            s = s[:-len(prefix)].strip()
            
    s = s.replace('Bs', '').replace('bs', '').replace('BS', '').replace('$', '').replace('€', '').strip()

    # Formatos con miles y decimales (ej: 1.250,50 vs 1,250.50)
    if '.' in s and ',' in s:
        if s.rfind(',') > s.rfind('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    elif ',' in s:
        s = s.replace(',', '.')

    # Sanitizar dígitos
    s_clean = ''.join(ch for ch in s if ch.isdigit() or ch in ('.', '-'))
    try:
        f = float(s_clean)
        if math.isnan(f) or math.isinf(f):
            return default
        if min_val is not None:
            f = max(min_val, f)
        return round(f, 2)
    except (ValueError, TypeError):
        return default

def _calc_item_status(qty, min_qty):
    qty = _safe_stock_float(qty, 0.0)
    min_qty = _safe_stock_float(min_qty, 0.0)
    if qty <= 0:
        return "out" # Agotado
    elif qty <= min_qty:
        return "low" # Stock Bajo / Alerta
    return "optimal" # Óptimo

@app.route('/api/stock', methods=['GET'])
def get_stock_items():
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None

    remote_items = []
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').order('created_at', desc=True).execute()
            if res.data is not None:
                remote_items = res.data
        except Exception as e:
            logging.warning("No se pudo consultar Supabase stock_items (usando local): %s", e)

    # Combinar registros locales y remotos asegurando persistencia y user_id
    local_items = _load_persisted_stock_items()
    local_map = {str(it.get('id')): it for it in local_items if it.get('id')}
    seen_ids = set()
    combined_items = []

    for it in local_items:
        it_copy = dict(it)
        it_copy['status'] = _calc_item_status(it_copy.get('stock_quantity'), it_copy.get('min_stock'))
        if it_copy.get('id'):
            seen_ids.add(str(it_copy.get('id')))
        combined_items.append(it_copy)

    for it in remote_items:
        it_id = str(it.get('id')) if it.get('id') else None
        it_copy = dict(it)
        local_match = local_map.get(it_id) if it_id else None
        if local_match and local_match.get('user_id') and not it_copy.get('user_id'):
            it_copy['user_id'] = local_match.get('user_id')
        it_copy['status'] = _calc_item_status(it_copy.get('stock_quantity'), it_copy.get('min_stock'))

        if it_id and it_id not in seen_ids:
            seen_ids.add(it_id)
            combined_items.append(it_copy)
            local_items.append(it_copy)
            _save_persisted_stock_items(local_items)

    # Excluir registros de almacenamiento de sistema (backups resilientes)
    clinical_items = [
        it for it in combined_items 
        if not (it.get('category') == '__SYSTEM__' or str(it.get('code', '')).startswith('__SYS_'))
    ]

    # Filtrar por multi-tenant (si no es SuperAdmin)
    if current_uid and current_user.get('role') != 'admin':
        filtered_items = []
        for it in clinical_items:
            it_uid = it.get('user_id')
            if not it_uid or it_uid == 'usr-doctor-001':
                it['user_id'] = current_uid
                it_uid = current_uid
                if supabase and it.get('id'):
                    try:
                        supabase.table('stock_items').update({'user_id': current_uid}).eq('id', it['id']).execute()
                    except Exception:
                        pass
            if it_uid == current_uid:
                filtered_items.append(it)
        return jsonify(filtered_items)

    return jsonify(clinical_items)

def _get_stock_item_by_id(item_id):
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').eq('id', str(item_id)).execute()
            if res.data:
                return res.data[0]
        except Exception:
            pass
    for item in _LOCAL_STOCK_ITEMS:
        if str(item.get('id')) == str(item_id):
            return item
    return None

def _generate_next_sku(raw_code=None, current_uid=None):
    """
    Genera o calcula el código SKU correlativo por doctor con relleno de huecos (gap filling).
    """
    prefix = "SKU"
    desired_num = None

    if raw_code and isinstance(raw_code, str):
        c_clean = raw_code.strip().upper()
        m_full = re.match(r'^([A-Z0-9]{1,8})-(\d+)$', c_clean)
        if m_full:
            prefix = m_full.group(1)
            try:
                desired_num = int(m_full.group(2))
            except ValueError:
                pass
        else:
            p_candidate = re.sub(r'[^A-Z0-9]', '', c_clean)
            if 2 <= len(p_candidate) <= 5:
                prefix = p_candidate
            elif len(p_candidate) > 5:
                prefix = p_candidate[:5]
            elif 1 <= len(p_candidate) < 2:
                prefix = (p_candidate + "X")[:2]

    # Recopilar todos los códigos existentes del usuario actual
    existing_codes = set()
    for it in _load_persisted_stock_items():
        c = it.get('code')
        it_uid = it.get('user_id')
        if c and not str(c).startswith('__SYS_'):
            if not current_uid or not it_uid or it_uid in ('usr-doctor-001', 'None', 'null', '') or it_uid == current_uid:
                existing_codes.add(str(c).upper().strip())
    
    if supabase:
        try:
            res = supabase.table('stock_items').select('code, user_id').execute()
            for r in (res.data or []):
                c = r.get('code')
                it_uid = r.get('user_id')
                if c and not str(c).startswith('__SYS_'):
                    if not current_uid or not it_uid or it_uid in ('usr-doctor-001', 'None', 'null', '') or it_uid == current_uid:
                        existing_codes.add(str(c).upper().strip())
        except Exception:
            pass

    # Extraer números usados con este prefijo
    pattern = re.compile(rf'^{re.escape(prefix)}-(\d+)$')
    used_numbers = set()
    for code in existing_codes:
        m = pattern.match(code)
        if m:
            try:
                used_numbers.add(int(m.group(1)))
            except ValueError:
                pass

    if desired_num is not None and desired_num > 0 and desired_num not in used_numbers:
        return f"{prefix}-{desired_num:03d}"

    # Relleno de huecos (menor correlativo disponible)
    next_num = 1
    while next_num in used_numbers:
        next_num += 1

    return f"{prefix}-{next_num:03d}"

def _ensure_category_and_unit_persisted(category_name, unit_name=None):
    """
    Garantiza que cualquier categoría o unidad creada al registrar/editar un producto
    quede registrada y persistida automáticamente en data/stock_taxonomies.json.
    """
    if not category_name and not unit_name:
        return
    try:
        cats, units = _load_persisted_taxonomies()
        changed = False
        if category_name and category_name.strip():
            c_clean = category_name.strip()
            if not any(c['name'].lower() == c_clean.lower() for c in cats):
                cats.append({
                    "name": c_clean,
                    "icon": "🏷️" if c_clean == "Sin Categoría" else "📦",
                    "description": "Categoría personalizada"
                })
                changed = True
        if unit_name and unit_name.strip():
            u_clean = unit_name.strip()
            if not any(u['name'].lower() == u_clean.lower() for u in units):
                units.append({"name": u_clean})
                changed = True
        if changed:
            _save_persisted_taxonomies(cats, units)
    except Exception as e:
        logging.warning("Error auto-registrando taxonomía: %s", e)

@app.route('/api/stock', methods=['POST'])
def create_stock_item():
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para registrar insumos.",
            "subscription_expired": True
        }), 403
    current_uid = current_user.get('id') if current_user else None

    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=150)
    if not name:
        return jsonify({"error": "El nombre del producto/insumo es obligatorio"}), 400

    item_id = str(uuid.uuid4())
    raw_code = _clean_str(data.get('code'), max_len=50)
    code = _generate_next_sku(raw_code, current_uid=current_uid)

    category = _clean_str(data.get('category'), max_len=80)
    if not category or category.strip().lower() in ("", "all", "todas las categorías", "todas"):
        category = "Sin Categoría"

    unit = _clean_str(data.get('unit'), max_len=30) or "Unidad (u)"
    qty = _safe_stock_float(data.get('stock_quantity'), default=0.0, min_val=0.0)
    min_qty = _safe_stock_float(data.get('min_stock'), default=5.0, min_val=0.0)
    cost = _safe_stock_float(data.get('cost_price'), default=0.0, min_val=0.0)
    sale = _safe_stock_float(data.get('sale_price'), default=0.0, min_val=0.0)

    new_item = {
        "id": item_id,
        "user_id": current_uid,
        "code": code,
        "name": name,
        "category": category,
        "unit": unit,
        "stock_quantity": qty,
        "min_stock": min_qty,
        "cost_price": cost,
        "sale_price": sale,
        "supplier": _clean_str(data.get('supplier'), max_len=150),
        "location": _clean_str(data.get('location'), max_len=150),
        "batch_number": _clean_str(data.get('batch_number'), max_len=100),
        "expiry_date": _clean_str(data.get('expiry_date'), max_len=20),
        "notes": _clean_str(data.get('notes'), max_len=500),
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    # Auto-registrar categoría y unidad en taxonomías persistidas
    _ensure_category_and_unit_persisted(category, unit)

    if supabase:
        try:
            res = supabase.table('stock_items').insert(new_item).execute()
            if res.data:
                item_res = res.data[0]
                item_res['status'] = _calc_item_status(item_res.get('stock_quantity'), item_res.get('min_stock'))
                _LOCAL_STOCK_ITEMS.insert(0, item_res)
                _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
                return jsonify({"success": True, "data": item_res}), 201
        except Exception:
            try:
                # Reintento sin columnas adicionales si el esquema remoto aún no fue migrado
                fallback_item = {k: v for k, v in new_item.items() if k not in ['batch_number', 'expiry_date', 'user_id']}
                res = supabase.table('stock_items').insert(fallback_item).execute()
                if res.data:
                    item_res = res.data[0]
                    item_res['user_id'] = new_item.get('user_id')
                    item_res['batch_number'] = new_item.get('batch_number')
                    item_res['expiry_date'] = new_item.get('expiry_date')
                    item_res['status'] = _calc_item_status(item_res.get('stock_quantity'), item_res.get('min_stock'))
                    _LOCAL_STOCK_ITEMS.insert(0, item_res)
                    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
                    return jsonify({"success": True, "data": item_res}), 201
            except Exception as e:
                logging.warning("Error al insertar en Supabase stock_items: %s", e)

    _LOCAL_STOCK_ITEMS.insert(0, new_item)
    new_item['status'] = _calc_item_status(qty, min_qty)
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    return jsonify({"success": True, "data": new_item}), 201

@app.route('/api/stock/<string:item_id>', methods=['PUT'])
def update_stock_item(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para modificar insumos.",
            "subscription_expired": True
        }), 403
    
    current_uid = current_user.get('id')
    target_item = _get_stock_item_by_id(item_id)
    if not target_item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    if target_item.get('user_id') and current_uid and target_item.get('user_id') != current_uid and current_user.get('role') != 'admin':
        return jsonify({"error": "No tienes permiso para modificar este insumo"}), 403

    data = request.json or {}
    updated = {}

    if 'name' in data:
        name = _clean_str(data.get('name'), max_len=150)
        if name:
            updated['name'] = name
    if 'code' in data:
        raw_code = _clean_str(data.get('code'), max_len=50)
        if raw_code:
            updated['code'] = raw_code.upper().strip()
    if 'category' in data:
        cat_val = _clean_str(data.get('category'), max_len=80)
        updated['category'] = cat_val if cat_val and cat_val.strip() else "Sin Categoría"
    if 'unit' in data:
        updated['unit'] = _clean_str(data.get('unit'), max_len=30) or "Unidad (u)"
    if 'stock_quantity' in data:
        updated['stock_quantity'] = _safe_stock_float(data.get('stock_quantity'), default=0.0, min_val=0.0)
    if 'min_stock' in data:
        updated['min_stock'] = _safe_stock_float(data.get('min_stock'), default=5.0, min_val=0.0)
    if 'cost_price' in data:
        updated['cost_price'] = _safe_stock_float(data.get('cost_price'), default=0.0, min_val=0.0)
    if 'sale_price' in data:
        updated['sale_price'] = _safe_stock_float(data.get('sale_price'), default=0.0, min_val=0.0)
    if 'supplier' in data:
        updated['supplier'] = _clean_str(data.get('supplier'), max_len=150)
    if 'location' in data:
        updated['location'] = _clean_str(data.get('location'), max_len=150)
    if 'batch_number' in data:
        updated['batch_number'] = _clean_str(data.get('batch_number'), max_len=100)
    if 'expiry_date' in data:
        updated['expiry_date'] = _clean_str(data.get('expiry_date'), max_len=20)
    if 'notes' in data:
        updated['notes'] = _clean_str(data.get('notes'), max_len=500)

    updated['updated_at'] = datetime.now(timezone.utc).isoformat()

    # Auto-registrar categoría y unidad en taxonomías
    _ensure_category_and_unit_persisted(updated.get('category'), updated.get('unit'))

    if supabase:
        try:
            res = supabase.table('stock_items').update(updated).eq('id', item_id).execute()
            if res.data:
                item_res = res.data[0]
                item_res['status'] = _calc_item_status(item_res.get('stock_quantity'), item_res.get('min_stock'))
                for local_it in _LOCAL_STOCK_ITEMS:
                    if local_it.get('id') == item_id:
                        local_it.update(item_res)
                        break
                _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
                return jsonify({"success": True, "data": item_res})
        except Exception as e:
            logging.warning("Error al actualizar en Supabase stock_items: %s", e)

    for item in _LOCAL_STOCK_ITEMS:
        if item.get('id') == item_id:
            item.update(updated)
            item['status'] = _calc_item_status(item.get('stock_quantity'), item.get('min_stock'))
            _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
            return jsonify({"success": True, "data": item})

    return jsonify({"error": "Artículo no encontrado"}), 404

@app.route('/api/stock/<string:item_id>', methods=['DELETE'])
def delete_stock_item(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar insumos.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')
    target_item = _get_stock_item_by_id(item_id)
    if not target_item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    if target_item.get('user_id') and current_uid and target_item.get('user_id') != current_uid and current_user.get('role') != 'admin':
        return jsonify({"error": "No tienes permiso para eliminar este insumo"}), 403

    if supabase:
        try:
            supabase.table('stock_items').delete().eq('id', item_id).execute()
        except Exception as e:
            logging.warning("Error al eliminar en Supabase stock_items: %s", e)

    global _LOCAL_STOCK_ITEMS
    _LOCAL_STOCK_ITEMS = [item for item in _LOCAL_STOCK_ITEMS if item.get('id') != item_id]
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    return jsonify({"success": True})

@app.route('/api/stock/bulk-delete', methods=['POST'])
def bulk_delete_stock_items():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar insumos.",
            "subscription_expired": True
        }), 403
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids or not isinstance(ids, list):
        return jsonify({"error": "Lista de IDs no válida o vacía"}), 400

    current_uid = current_user.get('id')
    is_admin = (current_user.get('role') == 'admin')

    allowed_ids = []
    for i_id in ids:
        item = _get_stock_item_by_id(str(i_id))
        if item and (is_admin or item.get('user_id') == current_uid or not item.get('user_id')):
            allowed_ids.append(str(i_id))

    if not allowed_ids:
        return jsonify({"error": "No tienes permiso para eliminar los insumos seleccionados"}), 403

    id_set = set(allowed_ids)
    if supabase:
        try:
            supabase.table('stock_items').delete().in_('id', list(id_set)).execute()
        except Exception as e:
            logging.warning("Error al eliminar masivamente en Supabase stock_items: %s", e)

    global _LOCAL_STOCK_ITEMS
    _LOCAL_STOCK_ITEMS = [item for item in _LOCAL_STOCK_ITEMS if str(item.get('id')) not in id_set]
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    return jsonify({"success": True, "deleted_count": len(id_set)})

@app.route('/api/stock/<string:item_id>/movement', methods=['POST'])
def record_stock_movement(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para registrar movimientos de inventario.",
            "subscription_expired": True
        }), 403
    current_uid = current_user.get('id')

    target_item = _get_stock_item_by_id(item_id)
    if not target_item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    if target_item.get('user_id') and current_uid and target_item.get('user_id') != current_uid and current_user.get('role') != 'admin':
        return jsonify({"error": "No tienes permiso para registrar movimientos en este insumo"}), 403

    data = request.json or {}
    mov_type = (data.get('type') or 'IN').upper()
    if mov_type not in ['IN', 'OUT', 'ADJUST']:
        mov_type = 'IN'

    qty = _safe_stock_float(data.get('quantity'), default=0.0, min_val=0.0)
    reason = _clean_str(data.get('reason'), max_len=250) or ("Entrada de stock" if mov_type == 'IN' else "Salida / Consumo clínico")

    if qty <= 0 and mov_type != 'ADJUST':
        return jsonify({"error": "La cantidad debe ser mayor a 0"}), 400

    current_qty = _safe_stock_float(target_item.get('stock_quantity'), 0.0)
    if mov_type == 'IN':
        new_qty = round(current_qty + qty, 2)
    elif mov_type == 'OUT':
        if current_qty < qty:
            return jsonify({"error": f"Stock insuficiente. Existencia actual: {current_qty}"}), 400
        new_qty = round(current_qty - qty, 2)
    else: # ADJUST
        new_qty = round(qty, 2)

    mov_record = {
        "id": str(uuid.uuid4()),
        "user_id": current_uid,
        "stock_item_id": item_id,
        "item_name": target_item.get('name'),
        "type": mov_type,
        "quantity": qty,
        "previous_quantity": current_qty,
        "new_quantity": new_qty,
        "reason": reason,
        "created_at": datetime.now(timezone.utc).isoformat()
    }

    # Actualizar stock en Supabase
    if supabase:
        try:
            supabase.table('stock_items').update({"stock_quantity": new_qty, "updated_at": datetime.now(timezone.utc).isoformat()}).eq('id', item_id).execute()
            try:
                supabase.table('stock_movements').insert(mov_record).execute()
            except Exception:
                pass
        except Exception as e:
            logging.warning("Error al actualizar movimiento en Supabase: %s", e)

    # Actualizar local asegurando sincronización con _LOCAL_STOCK_ITEMS
    target_item['stock_quantity'] = new_qty
    target_item['status'] = _calc_item_status(new_qty, target_item.get('min_stock'))

    found_local = False
    for local_it in _LOCAL_STOCK_ITEMS:
        if local_it.get('id') == item_id:
            local_it['stock_quantity'] = new_qty
            local_it['status'] = _calc_item_status(new_qty, local_it.get('min_stock'))
            found_local = True
            break
    if not found_local:
        _LOCAL_STOCK_ITEMS.insert(0, target_item)

    _LOCAL_STOCK_MOVEMENTS.insert(0, mov_record)
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)

    return jsonify({
        "success": True,
        "data": target_item,
        "movement": mov_record
    })

@app.route('/api/stock/movements', methods=['GET'])
def get_stock_movements():
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None

    item_id = _clean_str(request.args.get('item_id'))
    remote_movs = []
    if supabase:
        try:
            query = supabase.table('stock_movements').select('*').order('created_at', desc=True)
            if item_id:
                query = query.eq('stock_item_id', item_id)
            res = query.limit(50).execute()
            if res.data is not None:
                remote_movs = res.data
        except Exception as e:
            logging.warning("Error al consultar movimientos en Supabase: %s", e)

    seen_ids = set()
    combined_movs = []

    # 1. Movimientos locales (incluyen ventas recientes y cancelaciones)
    local_source = [m for m in _load_persisted_stock_movements() if not item_id or m.get('stock_item_id') == item_id or m.get('item_id') == item_id]
    for m in local_source:
        m_id = m.get('id')
        if m_id:
            seen_ids.add(m_id)
        combined_movs.append(m)

    # 2. Movimientos remotos de Supabase
    for m in remote_movs:
        m_id = m.get('id')
        if m_id not in seen_ids:
            if m_id:
                seen_ids.add(m_id)
            combined_movs.append(m)

    if current_uid and current_user.get('role') != 'admin':
        filtered_movs = []
        for m in combined_movs:
            m_uid = m.get('user_id')
            if not m_uid or m_uid in ('usr-doctor-001', 'None', 'null', ''):
                m['user_id'] = current_uid
                m_uid = current_uid
                if supabase and m.get('id'):
                    try:
                        supabase.table('stock_movements').update({'user_id': current_uid}).eq('id', m['id']).execute()
                    except Exception:
                        pass
            if m_uid == current_uid:
                filtered_movs.append(m)
        combined_movs = filtered_movs

    return jsonify(combined_movs[:100])

# --- GESTIÓN MAESTRA DE TAXONOMÍAS DE STOCK (CATEGORÍAS Y U/M - OPCIÓN 1) ---

_TAXONOMIES_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "stock_taxonomies.json")

_DEFAULT_STOCK_CATEGORIES = [
    {"name": "Insumos BIA", "icon": "🩺", "description": "Electrodos, gel conductor, cables y accesorios de bioimpedancia"},
    {"name": "Suplementos Nutricionales", "icon": "💊", "description": "Proteínas, creatina, aminoácidos, vitaminas y minerales"},
    {"name": "Material Clínico e Higiene", "icon": "🧼", "description": "Toallitas con alcohol, guantes, papel camilla y desinfección"},
    {"name": "Accesorios y Equipos", "icon": "📦", "description": "Cintas métricas, tallímetros, calipers y básculas"},
    {"name": "Medicamentos / Fármacos", "icon": "💉", "description": "Fármacos clínicos de uso o prescripción en consulta"},
    {"name": "Material de Oficina", "icon": "📝", "description": "Fichas, papel de impresión y suministros administrativos"},
    {"name": "Otros", "icon": "🏷️", "description": "Artículos varios no clasificados"}
]

_DEFAULT_STOCK_UNITS = [
    {"name": "Unidad (u)"},
    {"name": "Frasco / Bote"},
    {"name": "Caja"},
    {"name": "Pack"},
    {"name": "Cápsulas"},
    {"name": "Tabletas"},
    {"name": "Sobres"},
    {"name": "Ampollas"},
    {"name": "Tubo"},
    {"name": "Gotero"},
    {"name": "Mililitros (ml)"},
    {"name": "Litros (L)"},
    {"name": "Gramos (g)"},
    {"name": "Kilogramos (kg)"}
]

def _load_persisted_taxonomies():
    cats = list(_DEFAULT_STOCK_CATEGORIES)
    units = list(_DEFAULT_STOCK_UNITS)
    if os.path.exists(_TAXONOMIES_PATH):
        try:
            with open(_TAXONOMIES_PATH, 'r', encoding='utf-8') as f:
                data = json.load(f)
                raw_cats = data.get('categories', _DEFAULT_STOCK_CATEGORIES)
                raw_units = data.get('units', _DEFAULT_STOCK_UNITS)
                
                clean_cats = []
                if isinstance(raw_cats, list):
                    for c in raw_cats:
                        if isinstance(c, str) and c.strip():
                            clean_cats.append({"name": c.strip(), "icon": "📦", "description": ""})
                        elif isinstance(c, dict) and c.get('name'):
                            clean_cats.append({
                                "name": str(c.get('name')).strip(),
                                "icon": str(c.get('icon') or '📦'),
                                "description": str(c.get('description') or '')
                            })
                if clean_cats:
                    cats = clean_cats

                clean_units = []
                if isinstance(raw_units, list):
                    for u in raw_units:
                        if isinstance(u, str) and u.strip():
                            clean_units.append({"name": u.strip()})
                        elif isinstance(u, dict) and u.get('name'):
                            clean_units.append({"name": str(u.get('name')).strip()})
                if clean_units:
                    units = clean_units
        except Exception as e:
            logging.warning("Error al leer stock_taxonomies.json: %s", e)
    return cats, units

def _save_persisted_taxonomies(categories, units):
    try:
        os.makedirs(os.path.dirname(_TAXONOMIES_PATH), exist_ok=True)
        with open(_TAXONOMIES_PATH, 'w', encoding='utf-8') as f:
            json.dump({"categories": categories, "units": units}, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar stock_taxonomies.json: %s", e)
        return False

@app.route('/api/stock/taxonomies', methods=['GET'])
def get_stock_taxonomies():
    try:
        persisted_cats, persisted_units = _load_persisted_taxonomies()

        items = []
        if supabase:
            try:
                res = supabase.table('stock_items').select('category, unit, code').execute()
                if res and res.data is not None:
                    items = res.data
            except Exception as e:
                logging.warning("Error al consultar stock_items para taxonomías en Supabase: %s", e)
        if not items:
            items = _load_persisted_stock_items()

        # Filtrar registros de sistema
        items = [
            it for it in items 
            if isinstance(it, dict) and it.get('category') != '__SYSTEM__' and not str(it.get('code', '')).startswith('__SYS_')
        ]

        cat_counts = {}
        unit_counts = {}
        for it in items:
            c = (it.get('category') or 'Otros').strip()
            u = (it.get('unit') or 'Unidad (u)').strip()
            if c:
                cat_counts[c] = cat_counts.get(c, 0) + 1
            if u:
                unit_counts[u] = unit_counts.get(u, 0) + 1

        categories = []
        seen_cats = set()
        for cat in persisted_cats:
            if isinstance(cat, dict) and cat.get('name'):
                c_name = str(cat.get('name')).strip()
                if c_name and c_name != '__SYSTEM__':
                    seen_cats.add(c_name.lower())
                    categories.append({
                        "name": c_name,
                        "icon": cat.get('icon', '📦'),
                        "description": cat.get('description', ''),
                        "count": cat_counts.get(c_name, 0)
                    })

        # Si hay categorías en items no registradas, incorporarlas
        new_cats_added = False
        for c_name, count in cat_counts.items():
            if c_name and c_name != '__SYSTEM__' and c_name.lower() not in seen_cats:
                new_cat = {
                    "name": c_name,
                    "icon": "📦",
                    "description": "Categoría personalizada",
                    "count": count
                }
                categories.append(new_cat)
                persisted_cats.append(new_cat)
                seen_cats.add(c_name.lower())
                new_cats_added = True

        if new_cats_added:
            _save_persisted_taxonomies(persisted_cats, persisted_units)

        # Añadir conteo de uso a las unidades
        units_with_counts = []
        for u in persisted_units:
            if isinstance(u, dict) and u.get('name'):
                u_name = str(u.get('name')).strip()
                if u_name:
                    units_with_counts.append({
                        "name": u_name,
                        "count": unit_counts.get(u_name, 0)
                    })

        return jsonify({
            "categories": categories,
            "units": units_with_counts
        }), 200
    except Exception as e:
        logging.error("Error al obtener taxonomías de stock: %s", e, exc_info=True)
        return jsonify({
            "categories": list(_DEFAULT_STOCK_CATEGORIES),
            "units": list(_DEFAULT_STOCK_UNITS)
        }), 200

@app.route('/api/stock/taxonomies/category', methods=['POST'])
def add_stock_category():
    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=80)
    icon = _clean_str(data.get('icon'), max_len=10) or '📦'
    description = _clean_str(data.get('description'), max_len=200) or 'Categoría personalizada'

    if not name:
        return jsonify({"error": "El nombre de la categoría es obligatorio"}), 400

    cats, units = _load_persisted_taxonomies()
    if any(c['name'].lower() == name.lower() for c in cats):
        return jsonify({"error": f"La categoría '{name}' ya existe"}), 409

    new_cat = {"name": name, "icon": icon, "description": description}
    cats.append(new_cat)
    _save_persisted_taxonomies(cats, units)

    return jsonify({"success": True, "category": new_cat}), 201

@app.route('/api/stock/taxonomies/category/<path:cat_name>', methods=['DELETE'])
def delete_stock_category(cat_name):
    import html
    raw_cat_name = _clean_str(cat_name)
    decoded_cat_name = html.unescape(raw_cat_name).strip()
    if not raw_cat_name and not decoded_cat_name:
        return jsonify({"error": "Nombre de categoría inválido"}), 400

    cats, units = _load_persisted_taxonomies()
    
    def _is_match(c):
        c_n = str(c.get('name') or '').strip().lower()
        c_un = html.unescape(c_n).strip()
        r_n = raw_cat_name.lower()
        d_n = decoded_cat_name.lower()
        return c_n in (r_n, d_n) or c_un in (r_n, d_n)

    filtered_cats = [c for c in cats if not _is_match(c)]

    if len(filtered_cats) == len(cats):
        # Si no coincidió con la lista en memoria, eliminar por coincidencia de nombre exacto
        filtered_cats = [c for c in cats if str(c.get('name') or '').strip().lower() != raw_cat_name.lower()]

    # Reasignar productos que tengan esta categoría a 'Otros' en Supabase y local
    if supabase:
        try:
            supabase.table('stock_items').update({"category": "Otros", "updated_at": datetime.now(timezone.utc).isoformat()}).eq('category', cat_name).execute()
        except Exception as e:
            logging.warning("Error reasignando categoría en Supabase al eliminar: %s", e)

    for item in _LOCAL_STOCK_ITEMS:
        if (item.get('category') or '').strip().lower() in (raw_cat_name.lower(), decoded_cat_name.lower()):
            item['category'] = 'Otros'

    _save_persisted_taxonomies(filtered_cats, units)
    return jsonify({"success": True, "message": f"Categoría eliminada correctamente"})

@app.route('/api/stock/taxonomies/unit', methods=['POST'])
def add_stock_unit():
    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=50)

    if not name:
        return jsonify({"error": "El nombre de la unidad es obligatorio"}), 400

    cats, units = _load_persisted_taxonomies()
    if any(u['name'].lower() == name.lower() for u in units):
        return jsonify({"error": f"La unidad '{name}' ya existe"}), 409

    new_unit = {"name": name}
    units.append(new_unit)
    _save_persisted_taxonomies(cats, units)

    return jsonify({"success": True, "unit": new_unit}), 201

@app.route('/api/stock/taxonomies/unit/<path:unit_name>', methods=['DELETE'])
def delete_stock_unit(unit_name):
    import html
    raw_unit_name = _clean_str(unit_name)
    decoded_unit_name = html.unescape(raw_unit_name).strip()
    if not raw_unit_name and not decoded_unit_name:
        return jsonify({"error": "Nombre de unidad inválido"}), 400

    cats, units = _load_persisted_taxonomies()
    
    def _is_unit_match(u):
        u_n = str(u.get('name') or '').strip().lower()
        u_un = html.unescape(u_n).strip()
        r_n = raw_unit_name.lower()
        d_n = decoded_unit_name.lower()
        return u_n in (r_n, d_n) or u_un in (r_n, d_n)

    filtered_units = [u for u in units if not _is_unit_match(u)]

    if len(filtered_units) == len(units):
        filtered_units = [u for u in units if str(u.get('name') or '').strip().lower() != raw_unit_name.lower()]

    _save_persisted_taxonomies(cats, filtered_units)
    return jsonify({"success": True, "message": f"Unidad '{unit_name}' eliminada correctamente"})

@app.route('/api/stock/taxonomies/unit', methods=['PUT'])
def update_stock_unit():
    data = request.json or {}
    old_name = _clean_str(data.get('old_name'))
    new_name = _clean_str(data.get('new_name'), max_len=50)

    if not old_name or not new_name:
        return jsonify({"error": "El nombre actual y el nuevo nombre son obligatorios"}), 400

    cats, units = _load_persisted_taxonomies()
    found = False
    for u in units:
        if u['name'].lower() == old_name.lower():
            u['name'] = new_name
            found = True
            break

    if not found:
        return jsonify({"error": f"Unidad '{old_name}' no encontrada"}), 404

    # Actualizar en cascada en Supabase stock_items
    if old_name.lower() != new_name.lower() and supabase:
        try:
            supabase.table('stock_items').update({"unit": new_name, "updated_at": datetime.now(timezone.utc).isoformat()}).eq('unit', old_name).execute()
        except Exception as e:
            logging.warning("Error actualizando unidad en Supabase: %s", e)

    # Actualizar en local
    for item in _LOCAL_STOCK_ITEMS:
        if (item.get('unit') or '').strip().lower() == old_name.lower():
            item['unit'] = new_name

    _save_persisted_taxonomies(cats, units)
    return jsonify({"success": True, "message": f"Unidad '{old_name}' actualizada a '{new_name}' correctamente"})

@app.route('/api/stock/taxonomies/category', methods=['PUT'])
@app.route('/api/stock/categories/rename', methods=['PUT'])
def rename_stock_category():
    data = request.json or {}
    old_name = _clean_str(data.get('old_name'))
    new_name = _clean_str(data.get('new_name'), max_len=80)

    if not old_name or not new_name:
        return jsonify({"error": "El nombre actual y el nuevo nombre son obligatorios"}), 400

    # Actualizar en cascada en Supabase
    if supabase:
        try:
            supabase.table('stock_items').update({"category": new_name, "updated_at": datetime.now(timezone.utc).isoformat()}).eq('category', old_name).execute()
        except Exception as e:
            logging.warning("Error al renombrar categoría en Supabase: %s", e)

    # Actualizar en cascada en Local
    updated_count = 0
    for item in _LOCAL_STOCK_ITEMS:
        if (item.get('category') or '').strip().lower() == old_name.lower():
            item['category'] = new_name
            updated_count += 1

    cats, units = _load_persisted_taxonomies()
    for cat in cats:
        if cat['name'].lower() == old_name.lower():
            cat['name'] = new_name
    _save_persisted_taxonomies(cats, units)

    return jsonify({"success": True, "updated_count": updated_count, "message": f"Categoría actualizada de '{old_name}' a '{new_name}'"})

# --- IMPORTACIÓN MASIVA DESDE EXCEL DE INSUMOS ---

@app.route('/api/stock/excel-template', methods=['GET'])
def download_stock_excel_template():
    import io
    static_file_path = os.path.join(os.path.dirname(_BACKEND_DIR), "frontend", "static", "Insumos_15_Ejemplo_VitaMetrix.xlsx")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Insumos"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "SKU",
            "Nombre Producto / Insumo*",
            "Categoría",
            "U. Medida",
            "Stock*",
            "St. Min",
            "P. Costo",
            "PVP",
            "Lote",
            "Vencimiento",
            "Ubicación",
            "Proveedor",
            "Notas / Posología"
        ]

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        ws.append(headers)
        ws.row_dimensions[1].height = 28

        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        sample_rows = [
            ["SKU-001", "Electrodos BIA Desechables (Pack x100)", "Insumos BIA", "Caja x100", 15, 5, 120.00, 180.00, "LOT-2026-A", "2027-12-31", "Vitrina 1, Estante B", "BioMedical Import S.R.L.", "Usar exclusivamente en bioimpedancias tetrapolares"],
            ["SKU-002", "Gel Conductor BIA Hipoalergénico 250ml", "Insumos BIA", "Frasco", 20, 3, 25.00, 45.00, "GEL-8890", "2028-06-30", "Vitrina 1, Estante A", "DermoSalud Bolivia", "Conservar en lugar fresco"],
            ["SKU-003", "Proteína Whey Isolate 100% (Bote 900g Vainilla)", "Suplementos Nutricionales", "Bote 900g", 12, 4, 280.00, 350.00, "WHEY-2026-09", "2027-09-15", "Estante Suplementos A", "NutriFit Express", "1 scoop (30g) aporta 25g de proteína pura"],
            ["SKU-004", "Creatina Monohidratada Creapure 300g", "Suplementos Nutricionales", "Bote 300g", 18, 5, 160.00, 210.00, "CREA-102", "2028-03-20", "Estante Suplementos A", "NutriFit Express", "5g diarios en fase de carga o mantenimiento"],
            ["SKU-005", "Multivitamínico Clínico Complejo B + ZINC", "Suplementos Nutricionales", "Caja x60", 25, 6, 85.00, 130.00, "VIT-2026-C", "2027-11-10", "Vitrina 2, Cajón B", "PharmaLife S.A.", "Tomar 1 cápsula junto al almuerzo"]
        ]

        row_font = Font(name="Arial", size=9.5)
        row_alignment = Alignment(vertical="center")

        for r_idx, row_data in enumerate(sample_rows, 2):
            ws.append(row_data)
            ws.row_dimensions[r_idx].height = 22
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = row_font
                cell.alignment = row_alignment
                cell.border = thin_border
                if c_idx in (5, 6):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (7, 8):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Plantilla_Importacion_Insumos_VitaMetrix.xlsx"
        )
    except Exception as e_gen:
        logging.warning("No se pudo generar plantilla Excel dinámicamente: %s. Enviando archivo estático de respaldo.", e_gen)
        if os.path.exists(static_file_path):
            return send_file(
                static_file_path,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="Plantilla_Importacion_Insumos_VitaMetrix.xlsx"
            )
        return jsonify({"error": "No se pudo generar la plantilla Excel"}), 500



def _find_existing_stock_item(name, code=None, current_uid=None):
    if not name and not code:
        return None
    norm_name = str(name).strip().lower() if name else ''
    norm_code = str(code).strip().upper() if code else ''

    all_items = list(_LOCAL_STOCK_ITEMS)
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').execute()
            if res.data:
                supa_map = {str(x.get('id')): x for x in res.data if x.get('id')}
                for item in all_items:
                    if str(item.get('id')) in supa_map:
                        item.update(supa_map[str(item.get('id'))])
                existing_ids = set(str(x.get('id')) for x in all_items if x.get('id'))
                for s_item in res.data:
                    if str(s_item.get('id')) not in existing_ids:
                        all_items.append(s_item)
        except Exception:
            pass

    for item in all_items:
        item_uid = item.get('user_id')
        if not current_uid or not item_uid or item_uid == 'usr-doctor-001' or item_uid == current_uid:
            it_name = str(item.get('name') or '').strip().lower()
            it_code = str(item.get('code') or '').strip().upper()
            if (norm_name and it_name == norm_name) or (norm_code and it_code == norm_code and not norm_code.startswith('SKU-AUTO')):
                return item
    return None

@app.route('/api/stock/preview-excel', methods=['POST'])
def preview_stock_excel():
    import io
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    
    current_uid = current_user.get('id')
    if 'file' not in request.files:
        return jsonify({"error": "No se seleccionó ningún archivo para analizar"}), 400

    file = request.files['file']
    filename = (file.filename or '').lower()
    if not filename or not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({"error": "Formato no soportado. Debes subir un archivo Excel (.xlsx, .xls) o CSV (.csv)"}), 400

    rows_raw = []
    file_bytes = file.read()

    try:
        if filename.endswith('.csv'):
            import csv
            file_str = file_bytes.decode('utf-8-sig', errors='ignore')
            reader = csv.reader(io.StringIO(file_str))
            for r in reader:
                if r and any(str(cell).strip() for cell in r):
                    rows_raw.append([str(c).strip() for c in r])
        else:
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "La librería 'openpyxl' se está instalando en el servidor. Reintenta en 1 minuto o sube un archivo .csv"}), 500
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and any(cell is not None and str(cell).strip() != '' for cell in r):
                    rows_raw.append([str(c).strip() if c is not None else '' for c in r])
    except Exception as ex_read:
        return jsonify({"error": f"No se pudo analizar el archivo: {str(ex_read)}"}), 400

    if not rows_raw or len(rows_raw) < 2:
        return jsonify({"error": "El archivo está vacío o no contiene filas de datos"}), 400

    header_row = [str(h).strip().lower() for h in rows_raw[0]]
    
    def _find_col_idx(aliases):
        for idx, h in enumerate(header_row):
            h_norm = h.replace('*', '').replace('(', '').replace(')', '').strip()
            for alias in aliases:
                if alias in h_norm or h_norm in alias:
                    return idx
        return -1

    idx_code = _find_col_idx(['sku', 'codigo', 'código', 'code'])
    idx_name = _find_col_idx(['nombre producto / insumo', 'nombre producto', 'nombre insumo', 'nombre', 'producto', 'insumo'])
    idx_cat = _find_col_idx(['categoria', 'categoría', 'category'])
    idx_unit = _find_col_idx(['u. medida', 'u.medida', 'unidad de medida', 'unidad', 'medida', 'unit'])
    idx_qty = _find_col_idx(['stock', 'cantidad inicial', 'cantidad', 'existencia', 'stock_quantity'])
    idx_min = _find_col_idx(['st. min', 'st.min', 'stock minimo', 'stock mínimo', 'alerta', 'min_stock'])
    idx_cost = _find_col_idx(['p. costo', 'p.costo', 'precio costo', 'costo', 'cost_price'])
    idx_sale = _find_col_idx(['pvp', 'precio venta', 'venta', 'sale_price'])
    idx_batch = _find_col_idx(['lote', 'numero de lote', 'número de lote', 'batch'])
    idx_expiry = _find_col_idx(['vencimiento', 'fecha de vencimiento', 'expiry'])
    idx_location = _find_col_idx(['ubicacion', 'ubicación', 'ubicación en consultorio', 'location'])
    idx_supplier = _find_col_idx(['proveedor', 'supplier'])
    idx_notes = _find_col_idx(['notas / posología', 'notas', 'posologia', 'posología', 'notes'])

    if idx_name == -1:
        idx_name = 1 if len(header_row) > 1 else 0
    if idx_qty == -1:
        idx_qty = 4 if len(header_row) > 4 else 2
    if idx_cost == -1 and len(header_row) > 6:
        idx_cost = 6
    if idx_sale == -1 and len(header_row) > 7:
        idx_sale = 7
    if idx_batch == -1 and len(header_row) > 8:
        idx_batch = 8
    if idx_expiry == -1 and len(header_row) > 9:
        idx_expiry = 9
    if idx_location == -1 and len(header_row) > 10:
        idx_location = 10
    if idx_supplier == -1 and len(header_row) > 11:
        idx_supplier = 11
    if idx_notes == -1 and len(header_row) > 12:
        idx_notes = 12

    preview_items = []
    temp_sku_counter = 1

    for r_num, row_data in enumerate(rows_raw[1:], start=2):
        def _get_val(col_idx):
            if col_idx >= 0 and col_idx < len(row_data):
                val = str(row_data[col_idx]).strip()
                if val.lower() in ('none', 'null', 'nan', ''):
                    return ''
                return val
            return ''

        name = _clean_str(_get_val(idx_name), max_len=150)
        if not name or name.lower().startswith('ej:') or name.lower().startswith('código'):
            continue

        raw_code = _clean_str(_get_val(idx_code), max_len=50)
        cat_val = _clean_str(_get_val(idx_cat), max_len=80) or 'Sin Categoría'
        unit_val = _clean_str(_get_val(idx_unit), max_len=30) or 'Unidad (u)'
        qty = _safe_stock_float(_get_val(idx_qty), default=0.0, min_val=0.0)
        min_qty = _safe_stock_float(_get_val(idx_min), default=5.0, min_val=0.0)
        cost_price = _safe_stock_float(_get_val(idx_cost), default=0.0, min_val=0.0)
        sale_price = _safe_stock_float(_get_val(idx_sale), default=0.0, min_val=0.0)

        batch = _clean_str(_get_val(idx_batch), max_len=100)
        expiry = _clean_str(_get_val(idx_expiry), max_len=30)
        location = _clean_str(_get_val(idx_location), max_len=150)
        supplier = _clean_str(_get_val(idx_supplier), max_len=150)
        notes = _clean_str(_get_val(idx_notes), max_len=500)

        # Detectar si coincide con insumo existente por Nombre o SKU
        existing_item = _find_existing_stock_item(name, code=raw_code, current_uid=current_uid)
        is_reabastecimiento = bool(existing_item)
        prev_stock = _safe_stock_float(existing_item.get('stock_quantity'), 0.0) if existing_item else 0.0
        total_after = round(prev_stock + qty, 2) if is_reabastecimiento else qty

        if existing_item:
            code = existing_item.get('code') or (raw_code.upper() if raw_code else f"SKU-{temp_sku_counter:03d}")
        else:
            if raw_code:
                code = raw_code.upper()
            else:
                code = f"SKU-{temp_sku_counter:03d} (Auto)"
                temp_sku_counter += 1

        preview_items.append({
            "row_num": r_num,
            "code": code,
            "name": name,
            "category": cat_val,
            "unit": unit_val,
            "stock_quantity": qty,
            "is_reabastecimiento": is_reabastecimiento,
            "previous_stock": prev_stock,
            "total_after_stock": total_after,
            "min_stock": min_qty,
            "cost_price": cost_price,
            "sale_price": sale_price,
            "batch_number": batch,
            "expiry_date": expiry,
            "location": location,
            "supplier": supplier,
            "notes": notes,
            "valid": True
        })

    return jsonify({
        "success": True,
        "total_count": len(preview_items),
        "items": preview_items
    }), 200

@app.route('/api/stock/import-excel', methods=['POST'])
def import_stock_excel():
    import io
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para importar insumos.",
            "subscription_expired": True
        }), 403
    
    current_uid = current_user.get('id')
    if 'file' not in request.files:
        return jsonify({"error": "No se seleccionó ningún archivo para importar"}), 400

    file = request.files['file']
    filename = (file.filename or '').lower()
    if not filename or not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({"error": "Formato no soportado. Debes subir un archivo Excel (.xlsx, .xls) o CSV (.csv)"}), 400

    rows_raw = []
    file_bytes = file.read()

    try:
        if filename.endswith('.csv'):
            import csv
            file_str = file_bytes.decode('utf-8-sig', errors='ignore')
            reader = csv.reader(io.StringIO(file_str))
            for r in reader:
                if r and any(str(cell).strip() for cell in r):
                    rows_raw.append([str(c).strip() for c in r])
        else:
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "La librería 'openpyxl' se está instalando en el servidor. Reintenta en 1 minuto o sube un archivo .csv"}), 500
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and any(cell is not None and str(cell).strip() != '' for cell in r):
                    rows_raw.append([str(c).strip() if c is not None else '' for c in r])
    except Exception as ex_read:
        logging.error("Error al leer archivo Excel/CSV de importación: %s", ex_read)
        return jsonify({"error": f"No se pudo leer el archivo: {str(ex_read)}"}), 400

    if not rows_raw or len(rows_raw) < 2:
        return jsonify({"error": "El archivo está vacío o no contiene filas de datos"}), 400

    header_row = [str(h).strip().lower() for h in rows_raw[0]]
    
    def _find_col_idx(aliases):
        for idx, h in enumerate(header_row):
            h_norm = h.replace('*', '').replace('(', '').replace(')', '').strip()
            for alias in aliases:
                if alias in h_norm or h_norm in alias:
                    return idx
        return -1

    idx_code = _find_col_idx(['sku', 'codigo', 'código', 'code'])
    idx_name = _find_col_idx(['nombre producto / insumo', 'nombre producto', 'nombre insumo', 'nombre', 'producto', 'insumo'])
    idx_cat = _find_col_idx(['categoria', 'categoría', 'category'])
    idx_unit = _find_col_idx(['u. medida', 'u.medida', 'unidad de medida', 'unidad', 'medida', 'unit'])
    idx_qty = _find_col_idx(['stock', 'cantidad inicial', 'cantidad', 'existencia', 'stock_quantity'])
    idx_min = _find_col_idx(['st. min', 'st.min', 'stock minimo', 'stock mínimo', 'alerta', 'min_stock'])
    idx_cost = _find_col_idx(['p. costo', 'p.costo', 'precio costo', 'costo', 'cost_price'])
    idx_sale = _find_col_idx(['pvp', 'precio venta', 'venta', 'sale_price'])
    idx_batch = _find_col_idx(['lote', 'numero de lote', 'número de lote', 'batch'])
    idx_expiry = _find_col_idx(['vencimiento', 'fecha de vencimiento', 'expiry'])
    idx_location = _find_col_idx(['ubicacion', 'ubicación', 'ubicación en consultorio', 'location'])
    idx_supplier = _find_col_idx(['proveedor', 'supplier'])
    idx_notes = _find_col_idx(['notas / posología', 'notas', 'posologia', 'posología', 'notes'])

    if idx_name == -1:
        idx_name = 1 if len(header_row) > 1 else 0

    if idx_qty == -1:
        idx_qty = 4 if len(header_row) > 4 else 2

    if idx_cost == -1 and len(header_row) > 6:
        idx_cost = 6

    if idx_sale == -1 and len(header_row) > 7:
        idx_sale = 7

    if idx_batch == -1 and len(header_row) > 8:
        idx_batch = 8

    if idx_expiry == -1 and len(header_row) > 9:
        idx_expiry = 9

    if idx_location == -1 and len(header_row) > 10:
        idx_location = 10

    if idx_supplier == -1 and len(header_row) > 11:
        idx_supplier = 11

    if idx_notes == -1 and len(header_row) > 12:
        idx_notes = 12

    imported_count = 0
    reabastecidos_count = 0
    nuevos_count = 0
    errors_list = []

    for r_num, row_data in enumerate(rows_raw[1:], start=2):
        try:
            def _get_val(col_idx):
                if col_idx >= 0 and col_idx < len(row_data):
                    val = str(row_data[col_idx]).strip()
                    if val.lower() in ('none', 'null', 'nan', ''):
                        return ''
                    return val
                return ''

            name = _clean_str(_get_val(idx_name), max_len=150)
            if not name or name.lower().startswith('ej:') or name.lower().startswith('código'):
                continue

            raw_code = _clean_str(_get_val(idx_code), max_len=50)

            cat_val = _clean_str(_get_val(idx_cat), max_len=80)
            if not cat_val or cat_val.lower() in ('sin categoría', 'todas', 'all', ''):
                cat_val = 'Sin Categoría'

            unit_val = _clean_str(_get_val(idx_unit), max_len=30) or 'Unidad (u)'

            qty = _safe_stock_float(_get_val(idx_qty), default=0.0, min_val=0.0)
            min_qty = _safe_stock_float(_get_val(idx_min), default=5.0, min_val=0.0)
            cost_price = _safe_stock_float(_get_val(idx_cost), default=0.0, min_val=0.0)
            sale_price = _safe_stock_float(_get_val(idx_sale), default=0.0, min_val=0.0)

            batch = _clean_str(_get_val(idx_batch), max_len=100)
            expiry = _clean_str(_get_val(idx_expiry), max_len=20)
            location = _clean_str(_get_val(idx_location), max_len=150)
            supplier = _clean_str(_get_val(idx_supplier), max_len=150)
            notes = _clean_str(_get_val(idx_notes), max_len=500)

            _ensure_category_and_unit_persisted(cat_val, unit_val)

            # Buscar si coincide con un producto existente por Nombre o por SKU
            existing_item = _find_existing_stock_item(name, code=raw_code, current_uid=current_uid)

            if existing_item:
                # CONSOLIDACIÓN / REABASTECIMIENTO DE INSUMO EXISTENTE
                prev_q = _safe_stock_float(existing_item.get('stock_quantity'), 0.0)
                new_q = round(prev_q + qty, 2)
                existing_item['stock_quantity'] = new_q
                existing_item['status'] = _calc_item_status(new_q, existing_item.get('min_stock', min_qty))
                existing_item['updated_at'] = datetime.now(timezone.utc).isoformat()

                if cost_price > 0:
                    existing_item['cost_price'] = cost_price
                if sale_price > 0:
                    existing_item['sale_price'] = sale_price
                if min_qty > 0:
                    existing_item['min_stock'] = min_qty
                if cat_val and cat_val != 'Sin Categoría':
                    existing_item['category'] = cat_val
                if unit_val and unit_val != 'Unidad (u)':
                    existing_item['unit'] = unit_val
                if batch:
                    existing_item['batch_number'] = batch
                if expiry:
                    existing_item['expiry_date'] = expiry
                if location:
                    existing_item['location'] = location
                if supplier:
                    existing_item['supplier'] = supplier
                if notes:
                    existing_item['notes'] = notes

                if supabase:
                    try:
                        update_payload = {
                            "stock_quantity": new_q,
                            "status": existing_item['status'],
                            "updated_at": existing_item['updated_at']
                        }
                        if cost_price > 0: update_payload["cost_price"] = cost_price
                        if sale_price > 0: update_payload["sale_price"] = sale_price
                        if batch: update_payload["batch_number"] = batch
                        if expiry: update_payload["expiry_date"] = expiry
                        if location: update_payload["location"] = location
                        if supplier: update_payload["supplier"] = supplier
                        if notes: update_payload["notes"] = notes

                        supabase.table('stock_items').update(update_payload).eq('id', existing_item['id']).execute()
                    except Exception as ex_sub:
                        logging.warning("Error al actualizar insumo en Supabase: %s", ex_sub)

                # Movimiento de Kardex Entrada (IN) por Reabastecimiento
                m_item = {
                    "id": str(uuid.uuid4()),
                    "user_id": current_uid,
                    "item_id": existing_item['id'],
                    "type": "IN",
                    "quantity": qty,
                    "previous_stock": prev_q,
                    "new_stock": new_q,
                    "reason": "Reabastecimiento masivo vía importación Excel/CSV",
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
                if supabase:
                    try:
                        supabase.table('stock_movements').insert(m_item).execute()
                    except Exception:
                        pass

                reabastecidos_count += 1
            else:
                # CREACIÓN DE NUEVO INSUMO
                code = _generate_next_sku(raw_code, current_uid=current_uid)
                item_id = str(uuid.uuid4())
                new_item = {
                    "id": item_id,
                    "user_id": current_uid,
                    "code": code,
                    "name": name,
                    "category": cat_val,
                    "unit": unit_val,
                    "stock_quantity": qty,
                    "min_stock": min_qty,
                    "cost_price": cost_price,
                    "sale_price": sale_price,
                    "supplier": supplier,
                    "location": location,
                    "batch_number": batch,
                    "expiry_date": expiry,
                    "notes": notes,
                    "status": _calc_item_status(qty, min_qty),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }

                if supabase:
                    try:
                        supabase.table('stock_items').insert(new_item).execute()
                    except Exception:
                        try:
                            fallback_item = {k: v for k, v in new_item.items() if k not in ['batch_number', 'expiry_date']}
                            supabase.table('stock_items').insert(fallback_item).execute()
                        except Exception:
                            pass

                _LOCAL_STOCK_ITEMS.insert(0, new_item)

                # Kardex de Creación Inicial
                if qty > 0:
                    m_item = {
                        "id": str(uuid.uuid4()),
                        "user_id": current_uid,
                        "item_id": item_id,
                        "type": "IN",
                        "quantity": qty,
                        "previous_stock": 0,
                        "new_stock": qty,
                        "reason": "Stock inicial por importación Excel/CSV",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
                    if supabase:
                        try:
                            supabase.table('stock_movements').insert(m_item).execute()
                        except Exception:
                            pass

                nuevos_count += 1

            imported_count += 1
        except Exception as e_row:
            errors_list.append(f"Fila {r_num}: {str(e_row)}")

    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)
    _invalidate_dashboard_cache()

    return jsonify({
        "success": True,
        "imported_count": imported_count,
        "reabastecidos_count": reabastecidos_count,
        "nuevos_count": nuevos_count,
        "errors": errors_list,
        "message": f"Se procesaron {imported_count} insumos ({reabastecidos_count} reabastecidos y {nuevos_count} nuevos registros)."
    }), 201

# --- MÓDULO DE VENTAS (POS CLÍNICO & COMPROBANTES) ---

_SALES_PATH = os.path.join(os.path.dirname(_BACKEND_DIR), "data", "sales.json")

def _load_persisted_sales():
    if os.path.exists(_SALES_PATH):
        try:
            with open(_SALES_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.warning("Error al leer sales.json: %s", e)
    return []

def _save_persisted_sales(sales):
    try:
        os.makedirs(os.path.dirname(_SALES_PATH), exist_ok=True)
        with open(_SALES_PATH, 'w', encoding='utf-8') as f:
            json.dump(sales, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error("Error al guardar sales.json: %s", e)
        return False

_LOCAL_SALES = _load_persisted_sales()

@app.route('/api/sales', methods=['GET'])
def get_sales():
    """Listado del historial de ventas con filtros opcionales."""
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None

    if supabase:
        try:
            query = supabase.table('sales').select('*, sale_items(*)').order('created_at', desc=True)
            res = query.limit(100).execute()
            if res.data is not None and len(res.data) > 0:
                sales = res.data
                if current_uid and current_user.get('role') != 'admin':
                    sales = [s for s in sales if s.get('user_id') == current_uid or not s.get('user_id')]
                return jsonify(sales)
        except Exception as e:
            logging.warning("Error al consultar ventas en Supabase: %s", e)

    sales = _load_persisted_sales() or _LOCAL_SALES
    if current_uid and current_user.get('role') != 'admin':
        sales = [s for s in sales if s.get('user_id') == current_uid or not s.get('user_id')]
    return jsonify(sales)

@app.route('/api/sales', methods=['POST'])
def create_sale():
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para procesar ventas en el POS.",
            "subscription_expired": True
        }), 403
    """
    Registra una nueva venta de productos/insumos:
    - Valida existencias en stock.
    - Descuenta las cantidades del inventario atómicamente.
    - Registra los movimientos de salida en Kardex (stock_movements).
    - Genera el comprobante con número correlativo REC-2026-XXXX.
    """
    data = request.json or {}
    items_req = data.get('items', [])
    if not items_req or not isinstance(items_req, list) or len(items_req) == 0:
        return jsonify({"error": "El carrito de venta no contiene productos"}), 400

    patient_name = _clean_str(data.get('patient_name'), max_len=100) or "Cliente Ocasional"
    patient_idp = _clean_str(data.get('patient_idp'), max_len=50) or ""
    patient_phone = _clean_str(data.get('patient_phone'), max_len=30) or ""
    payment_method = _clean_str(data.get('payment_method'), max_len=50) or "Efectivo"
    notes = _clean_str(data.get('notes'), max_len=500) or ""
    discount = max(0.0, _safe_stock_float(data.get('discount'), default=0.0))

    # 1. Consolidar items por ID de producto (evita duplicados maliciosos o fallas de concurrencia)
    aggregated_req = {}
    for it in items_req:
        stk_id = it.get('stock_item_id') or it.get('id')
        if not stk_id:
            continue
        qty = _safe_stock_float(it.get('quantity'), default=1.0, min_val=0.01)
        if stk_id not in aggregated_req:
            aggregated_req[stk_id] = {**it, 'stock_item_id': stk_id, 'quantity': qty}
        else:
            aggregated_req[stk_id]['quantity'] = round(aggregated_req[stk_id]['quantity'] + qty, 2)

    if not aggregated_req:
        return jsonify({"error": "No se encontraron artículos válidos en la solicitud de venta"}), 400

    processed_items = []
    total_sale = 0.0
    total_cost = 0.0

    for stk_id, it in aggregated_req.items():
        qty = it['quantity']

        # Buscar producto en memoria o base de datos
        stock_prod = None
        for p in _LOCAL_STOCK_ITEMS:
            if p.get('id') == stk_id:
                stock_prod = p
                break

        if not stock_prod:
            return jsonify({"error": f"Producto '{it.get('name', stk_id)}' no encontrado en el inventario"}), 404

        current_avail = _safe_stock_float(stock_prod.get('stock_quantity'), 0.0)
        if current_avail < qty:
            return jsonify({
                "error": f"Stock insuficiente para '{stock_prod.get('name')}'. Disponible: {current_avail} {stock_prod.get('unit', 'u')}, solicitado: {qty}"
            }), 400

        unit_price = _safe_stock_float(it.get('unit_price') if it.get('unit_price') is not None else stock_prod.get('sale_price', 0.0), min_val=0.0)
        cost_price = _safe_stock_float(stock_prod.get('cost_price', 0.0), min_val=0.0)
        item_subtotal = round(unit_price * qty, 2)
        item_cost_subtotal = round(cost_price * qty, 2)

        total_sale += item_subtotal
        total_cost += item_cost_subtotal

        processed_items.append({
            "stock_item_id": stock_prod.get('id'),
            "code": stock_prod.get('code'),
            "name": stock_prod.get('name'),
            "unit": stock_prod.get('unit', 'Unidad (u)'),
            "quantity": qty,
            "unit_price": unit_price,
            "cost_price": cost_price,
            "subtotal": item_subtotal,
            "_stock_ref": stock_prod
        })

    subtotal = round(total_sale, 2)
    final_total = max(0.0, round(subtotal - discount, 2))
    profit = round(final_total - total_cost, 2)

    amount_received = _safe_stock_float(data.get('amount_received'), default=final_total)
    change_given = max(0.0, round(amount_received - final_total, 2)) if payment_method.lower() == 'efectivo' else 0.0

    all_sales = _load_persisted_sales()
    next_num = len(all_sales) + 1
    receipt_number = f"REC-2026-{next_num:04d}"
    sale_id = str(uuid.uuid4())
    created_now = datetime.now(timezone.utc).isoformat()

    # 2. Descontar stock y registrar Kardex de cada ítem
    for pit in processed_items:
        stk_ref = pit['_stock_ref']
        old_q = _safe_stock_float(stk_ref.get('stock_quantity'), 0.0)
        new_q = round(old_q - pit['quantity'], 2)
        stk_ref['stock_quantity'] = new_q
        stk_ref['status'] = _calc_item_status(new_q, stk_ref.get('min_stock'))

        # Movimiento de Kardex
        mov_record = {
            "id": str(uuid.uuid4()),
            "stock_item_id": stk_ref.get('id'),
            "item_name": stk_ref.get('name'),
            "type": "SALE",
            "quantity": pit['quantity'],
            "previous_quantity": old_q,
            "new_quantity": new_q,
            "reason": f"Venta {receipt_number} a {patient_name}",
            "reference_id": sale_id,
            "created_at": created_now
        }
        _LOCAL_STOCK_MOVEMENTS.insert(0, mov_record)

        if supabase:
            try:
                supabase.table('stock_items').update({
                    "stock_quantity": new_q,
                    "updated_at": created_now
                }).eq('id', stk_ref.get('id')).execute()
                try:
                    supabase.table('stock_movements').insert(mov_record).execute()
                except Exception:
                    pass
            except Exception as e:
                logging.warning("Error al descontar stock en Supabase: %s", e)

    # Persistir localmente inventario y kardex actualizados
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)

    # 3. Guardar registro de la venta
    clean_sale_items = []
    for pit in processed_items:
        clean_sale_items.append({
            "stock_item_id": pit['stock_item_id'],
            "code": pit['code'],
            "name": pit['name'],
            "unit": pit['unit'],
            "quantity": pit['quantity'],
            "unit_price": pit['unit_price'],
            "cost_price": pit['cost_price'],
            "subtotal": pit['subtotal']
        })

    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None

    sale_record = {
        "id": sale_id,
        "user_id": current_uid,
        "receipt_number": receipt_number,
        "patient_name": patient_name,
        "patient_idp": patient_idp,
        "patient_phone": patient_phone,
        "items": clean_sale_items,
        "subtotal": subtotal,
        "discount": discount,
        "total": final_total,
        "total_cost": round(total_cost, 2),
        "profit": profit,
        "payment_method": payment_method,
        "amount_received": amount_received,
        "change_given": change_given,
        "status": "COMPLETED",
        "notes": notes,
        "created_at": created_now
    }

    all_sales.insert(0, sale_record)
    _save_persisted_sales(all_sales)
    global _LOCAL_SALES
    _LOCAL_SALES = all_sales

    if supabase:
        try:
            supabase.table('sales').insert({
                "id": sale_id,
                "receipt_number": receipt_number,
                "patient_name": patient_name,
                "patient_idp": patient_idp,
                "patient_phone": patient_phone,
                "subtotal": subtotal,
                "discount": discount,
                "total": final_total,
                "total_cost": round(total_cost, 2),
                "profit": profit,
                "payment_method": payment_method,
                "amount_received": amount_received,
                "change_given": change_given,
                "status": "COMPLETED",
                "notes": notes,
                "created_at": created_now
            }).execute()
        except Exception as e:
            logging.warning("Error al guardar venta en Supabase: %s", e)

    return jsonify({
        "success": True,
        "message": f"Venta {receipt_number} completada exitosamente",
        "sale": sale_record
    }), 201

@app.route('/api/sales/<string:sale_id>', methods=['GET'])
def get_sale_detail(sale_id):
    """Obtiene el detalle de una venta por ID."""
    all_sales = _load_persisted_sales() or _LOCAL_SALES
    for s in all_sales:
        if s.get('id') == sale_id or s.get('receipt_number') == sale_id:
            return jsonify(s)
    return jsonify({"error": "Venta no encontrada"}), 404

@app.route('/api/sales/<string:sale_id>', methods=['DELETE'])
def cancel_sale(sale_id):
    current_user = _get_current_user()
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para anular ventas.",
            "subscription_expired": True
        }), 403
    """
    Anula una venta y restituye automáticamente las cantidades al inventario.
    Registra el movimiento en Kardex (SALE_CANCEL).
    """
    all_sales = _load_persisted_sales() or _LOCAL_SALES
    target_sale = None
    for s in all_sales:
        if s.get('id') == sale_id or s.get('receipt_number') == sale_id:
            target_sale = s
            break

    if not target_sale:
        return jsonify({"error": "Venta no encontrada"}), 404

    if target_sale.get('status') == 'CANCELLED':
        return jsonify({"error": "Esta venta ya se encuentra anulada"}), 400

    target_sale['status'] = 'CANCELLED'
    target_sale['cancelled_at'] = datetime.now(timezone.utc).isoformat()

    created_now = datetime.now(timezone.utc).isoformat()

    # Reingresar ítems al inventario
    for item in target_sale.get('items', []):
        stk_id = item.get('stock_item_id')
        qty = _safe_stock_float(item.get('quantity'), 0.0)
        if not stk_id or qty <= 0:
            continue

        p = None
        for local_p in _LOCAL_STOCK_ITEMS:
            if local_p.get('id') == stk_id:
                p = local_p
                break

        if not p and supabase:
            try:
                res_p = supabase.table('stock_items').select('*').eq('id', stk_id).execute()
                if res_p.data:
                    p = res_p.data[0]
                    _LOCAL_STOCK_ITEMS.append(p)
            except Exception:
                pass

        if p:
            old_q = _safe_stock_float(p.get('stock_quantity'), 0.0)
            new_q = round(old_q + qty, 2)
            p['stock_quantity'] = new_q
            p['status'] = _calc_item_status(new_q, p.get('min_stock'))

            # Kardex cancel
            mov = {
                "id": str(uuid.uuid4()),
                "stock_item_id": p.get('id'),
                "item_name": p.get('name'),
                "type": "SALE_CANCEL",
                "quantity": qty,
                "previous_quantity": old_q,
                "new_quantity": new_q,
                "reason": f"Anulación de Venta {target_sale.get('receipt_number')}",
                "reference_id": target_sale.get('id'),
                "created_at": created_now
            }
            _LOCAL_STOCK_MOVEMENTS.insert(0, mov)

            if supabase:
                try:
                    supabase.table('stock_items').update({"stock_quantity": new_q}).eq('id', p.get('id')).execute()
                    try:
                        supabase.table('stock_movements').insert(mov).execute()
                    except Exception:
                        pass
                except Exception as e:
                    logging.warning("Error al restituir stock en Supabase: %s", e)
                break

    _save_persisted_sales(all_sales)
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)

    if supabase:
        try:
            supabase.table('sales').update({"status": "CANCELLED"}).eq('id', target_sale.get('id')).execute()
        except Exception as e:
            logging.warning("Error al anular venta en Supabase: %s", e)

    return jsonify({
        "success": True,
        "message": f"Venta {target_sale.get('receipt_number')} anulada y stock restituido correctamente",
        "sale": target_sale
    })

@app.route('/api/sales/stats', methods=['GET'])
def get_sales_stats():
    """Métricas y KPIs financieros de ventas para el nutricionista."""
    sales = [s for s in (_load_persisted_sales() or _LOCAL_SALES) if s.get('status') != 'CANCELLED']
    
    total_sales_amount = round(sum(_safe_stock_float(s.get('total', 0)) for s in sales), 2)
    total_profit = round(sum(_safe_stock_float(s.get('profit', 0)) for s in sales), 2)
    sales_count = len(sales)
    avg_ticket = round(total_sales_amount / sales_count, 2) if sales_count > 0 else 0.0

    # Ventas de hoy
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_sales = [s for s in sales if (s.get('created_at') or '').startswith(today_str)]
    today_sales_amount = round(sum(_safe_stock_float(s.get('total', 0)) for s in today_sales), 2)

    # Top productos vendidos
    prod_counts = {}
    for s in sales:
        for it in s.get('items', []):
            pname = it.get('name', 'Producto')
            qty = _safe_stock_float(it.get('quantity', 0))
            subt = _safe_stock_float(it.get('subtotal', 0))
            if pname not in prod_counts:
                prod_counts[pname] = {"name": pname, "quantity": 0, "total_revenue": 0.0}
            prod_counts[pname]["quantity"] += qty
            prod_counts[pname]["total_revenue"] += subt

    top_products = sorted(prod_counts.values(), key=lambda x: x["quantity"], reverse=True)[:5]

    return jsonify({
        "total_sales_amount": total_sales_amount,
        "total_profit": total_profit,
        "sales_count": sales_count,
        "avg_ticket": avg_ticket,
        "today_sales_amount": today_sales_amount,
        "today_sales_count": len(today_sales),
        "top_products": top_products
    })

# --- CHATBOT WEBHOOK (WhatsApp / Telegram Automation) ---

@app.route('/api/bot/webhook', methods=['POST', 'GET'])
def bot_webhook():
    """
    Webhook inteligente para Evolution API, Baileys, Telegram y Webhooks directos.
    Permite agendamiento automatizado, consulta de disponibilidad y preparación clínica.
    """
    if request.method == 'GET':
        # Verificación de webhook (Meta / Evolution)
        challenge = request.args.get('hub.challenge')
        return challenge if challenge else jsonify({"status": "active", "service": "VitaMetrix Chatbot Engine"}), 200

    payload = request.json or {}
    sender = ""
    incoming_text = ""

    # 1. Parsear formato Evolution API / Baileys
    if "data" in payload and isinstance(payload.get("data"), dict) and "message" in payload.get("data", {}):
        msg_data = payload["data"]
        sender = msg_data.get("key", {}).get("remoteJid", "").split('@')[0]
        msg = msg_data.get("message", {})
        incoming_text = msg.get("conversation") or msg.get("extendedTextMessage", {}).get("text", "")
    # 2. Parsear formato Telegram
    elif "message" in payload and isinstance(payload.get("message"), dict):
        tele_msg = payload.get("message", {})
        sender_info = tele_msg.get("from")
        sender = sender_info.get("first_name", "Paciente") if isinstance(sender_info, dict) else "Paciente"
        incoming_text = tele_msg.get("text", "")
    # 3. Formato directo / REST
    else:
        sender = str(payload.get("sender") or "Paciente")
        incoming_text = str(payload.get("message") or "")

    text_clean = incoming_text.lower().strip()
    response_msg = ""
    action_taken = "none"

    if any(k in text_clean for k in ["agendar", "cita", "turno", "reservar", "1"]):
        action_taken = "booking_flow"
        response_msg = (
            f"¡Hola {sender}! 👋 Te ayudo a agendar tu Evaluación de Bioimpedancia (BIA) en VitaMetrix.\n\n"
            "📅 *Horarios Disponibles para esta semana:*\n"
            "• Mañanas: 09:00 AM, 10:30 AM, 11:45 AM\n"
            "• Tardes: 03:30 PM, 04:45 PM, 06:00 PM\n\n"
            "Por favor responde con tu *Nombre completo* y la *Fecha y Hora deseada* (ej: 'Marta Díaz, Mañana a las 10:30 AM')."
        )
    elif any(k in text_clean for k in ["precio", "costo", "tarifa", "cuanto", "2"]):
        action_taken = "pricing_info"
        response_msg = (
            "📊 *Evaluación Integral de Bioimpedancia y Salud Celular (VitaMetrix)*\n\n"
            "El estudio incluye:\n"
            "✅ Análisis Vectorial BIVA (Agua Celular & Membranas)\n"
            "✅ Puntuación TRU Body Score (Músculo vs Grasa)\n"
            "✅ Gasto Energético Metabólico (REE / TEE)\n"
            "✅ Músculo Segmental y Grasa Visceral (Riesgo IDF)\n"
            "✅ Informe Clínico Digital en PDF\n\n"
            "Responde *1* si deseas reservar tu turno."
        )
    elif any(k in text_clean for k in ["preparacion", "ayuno", "indicacion", "requisito", "3"]):
        action_taken = "prep_instructions"
        response_msg = (
            "📋 *Indicaciones previas para tu prueba de Bioimpedancia:*\n\n"
            "1. Ayuno de alimentos y líquidos de al menos 2 horas.\n"
            "2. No realizar actividad física intensa 12 horas antes.\n"
            "3. Evitar cafeína o diuréticos previo al examen.\n"
            "4. Asistir con ropa cómoda y sin joyas/metales en tobillos y muñecas.\n\n"
            "¡Te esperamos!"
        )
    elif any(k in text_clean for k in ["doctor", "humano", "especialista", "4"]):
        action_taken = "human_escalation"
        response_msg = (
            "👨‍⚕️ He notificado a nuestro equipo médico. Un especialista se comunicará contigo a la brevedad.\n"
            "Horario de atención: Lunes a Viernes de 08:30 a 19:00."
        )
    else:
        action_taken = "menu"
        response_msg = (
            f"¡Hola {sender}! 👋 Bienvenido/a al servicio de atención de *VitaMetrix*.\n\n"
            "¿En qué podemos ayudarte hoy?\n\n"
            "1️⃣ Agendar una Evaluación de Bioimpedancia\n"
            "2️⃣ Conocer qué incluye el análisis y tarifas\n"
            "3️⃣ Indicaciones previas a la prueba (ayuno/preparación)\n"
            "4️⃣ Hablar con un especialista"
        )

    return jsonify({
        "success": True,
        "sender": sender,
        "received": incoming_text,
        "response": response_msg,
        "action": action_taken,
        "human_delay_seconds": 2.5
    }), 200

@app.route('/api/health', methods=['GET'])
def health():
    """Health check para Render y monitores de uptime."""
    db_status = "connected" if supabase else "no-credentials"
    return jsonify({"status": "ok", "supabase": db_status}), 200

if __name__ == '__main__':
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug_mode, host='0.0.0.0', port=int(os.environ.get("PORT", 5000)))