import os
import io
import json
import logging
import uuid
from datetime import datetime, timezone
from flask import Blueprint, request, jsonify, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.helpers import (
    supabase,
    _clean_str,
    _get_current_user,
    _is_subscription_active,
    _safe_stock_float,
    _clean_expiry_date,
    _calc_item_status,
    _load_persisted_stock_items,
    _save_persisted_stock_items,
    _load_persisted_stock_movements,
    _save_persisted_stock_movements,
    _load_persisted_taxonomies,
    _save_persisted_taxonomies,
    _now_bolivia,
    _supabase_insert_stock_item,
    _supabase_update_stock_item,
    _supabase_insert_stock_movement,
    _find_existing_stock_item,
    _generate_next_sku,
    _ensure_category_and_unit_persisted,
    _get_stock_item_by_id,
    _invalidate_dashboard_cache,
    _LOCAL_STOCK_ITEMS,
    _LOCAL_STOCK_MOVEMENTS
)

stock_bp = Blueprint('stock_bp', __name__)

@stock_bp.route('/api/stock', methods=['GET'])
def get_stock_items():
    current_user = _get_current_user()
    current_uid = current_user.get('id') if current_user else None

    global _LOCAL_STOCK_ITEMS

    remote_items = []
    if supabase:
        try:
            res = supabase.table('stock_items').select('*').order('created_at', desc=True).execute()
            if res.data is not None:
                remote_items = res.data
        except Exception as e:
            logging.warning("No se pudo consultar Supabase stock_items: %s", e)

    disk_items = _load_persisted_stock_items()
    unified_map = {}

    for it in _LOCAL_STOCK_ITEMS:
        if it and it.get('id'):
            unified_map[str(it.get('id'))] = dict(it)

    for it in disk_items:
        if it and it.get('id'):
            i_id = str(it.get('id'))
            if i_id not in unified_map:
                unified_map[i_id] = dict(it)
            else:
                unified_map[i_id].update({k: v for k, v in it.items() if v is not None and v != ''})

    for r_item in remote_items:
        if r_item and r_item.get('id'):
            r_id = str(r_item.get('id'))
            if r_id not in unified_map:
                unified_map[r_id] = dict(r_item)
            else:
                unified_map[r_id].update({k: v for k, v in r_item.items() if v is not None and v != ''})

    all_combined = list(unified_map.values())
    for item in all_combined:
        item['status'] = _calc_item_status(item.get('stock_quantity'), item.get('min_stock'))

    _LOCAL_STOCK_ITEMS = all_combined
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)

    clinical_items = [
        it for it in _LOCAL_STOCK_ITEMS 
        if not (it.get('category') == '__SYSTEM__' or str(it.get('code', '')).startswith('__SYS_'))
    ]

    if current_uid and current_user.get('role') != 'admin':
        filtered_items = []
        for it in clinical_items:
            it_uid = it.get('user_id')
            if not it_uid or it_uid in ('usr-doctor-001', 'None', 'null', ''):
                it['user_id'] = current_uid
                it_uid = current_uid
                _supabase_update_stock_item(it.get('id'), {'user_id': current_uid})
            if it_uid == current_uid:
                filtered_items.append(it)
        return jsonify(filtered_items)
    return jsonify(clinical_items)

@stock_bp.route('/api/stock', methods=['POST'])
def create_stock_item():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN de licencia para registrar insumos.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')
    data = request.json or {}
    name = _clean_str(data.get('name'), max_len=150)
    if not name:
        return jsonify({"error": "El nombre del insumo/producto es obligatorio"}), 400

    raw_code = _clean_str(data.get('code'), max_len=50)
    code = _generate_next_sku(raw_code, current_uid=current_uid)

    category = _clean_str(data.get('category'), max_len=80) or "Sin Categoría"
    unit = _clean_str(data.get('unit'), max_len=30) or "Unidad (u)"
    stock_qty = _safe_stock_float(data.get('stock_quantity'), default=0.0, min_val=0.0)
    min_stock = _safe_stock_float(data.get('min_stock'), default=5.0, min_val=0.0)
    cost_price = _safe_stock_float(data.get('cost_price'), default=0.0, min_val=0.0)
    sale_price = _safe_stock_float(data.get('sale_price'), default=0.0, min_val=0.0)
    supplier = _clean_str(data.get('supplier'), max_len=150)
    location = _clean_str(data.get('location'), max_len=150)
    batch_number = _clean_str(data.get('batch_number'), max_len=100)
    expiry_date = _clean_expiry_date(data.get('expiry_date'))
    notes = _clean_str(data.get('notes'), max_len=500)

    _ensure_category_and_unit_persisted(category, unit)

    new_item = {
        "id": str(uuid.uuid4()),
        "user_id": current_uid,
        "code": code,
        "name": name,
        "category": category,
        "unit": unit,
        "stock_quantity": stock_qty,
        "min_stock": min_stock,
        "cost_price": cost_price,
        "sale_price": sale_price,
        "supplier": supplier,
        "location": location,
        "batch_number": batch_number,
        "expiry_date": expiry_date,
        "notes": notes,
        "status": _calc_item_status(stock_qty, min_stock),
        "created_at": _now_bolivia().isoformat(),
        "updated_at": _now_bolivia().isoformat()
    }

    _LOCAL_STOCK_ITEMS.insert(0, new_item)
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _supabase_insert_stock_item(new_item)

    if stock_qty > 0:
        m_item = {
            "id": str(uuid.uuid4()),
            "user_id": current_uid,
            "item_id": new_item['id'],
            "type": "IN",
            "quantity": stock_qty,
            "previous_stock": 0.0,
            "new_stock": stock_qty,
            "reason": "Inventario Inicial al registrar insumo",
            "created_at": _now_bolivia().isoformat()
        }
        _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
        _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)
        _supabase_insert_stock_movement(m_item)

    _invalidate_dashboard_cache()
    return jsonify({
        "success": True,
        "item": new_item,
        "message": f"Insumo '{name}' ({code}) registrado con éxito."
    }), 201

@stock_bp.route('/api/stock/<item_id>', methods=['PUT'])
def update_stock_item(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para modificar insumos.",
            "subscription_expired": True
        }), 403
    
    current_uid = current_user.get('id')
    target_item = _get_stock_item_by_id(item_id)
    if not target_item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    if target_item.get('user_id') and target_item.get('user_id') not in (current_uid, 'usr-doctor-001', 'None', 'null', '') and current_user.get('role') != 'admin':
        return jsonify({"error": "No tienes permiso para modificar este insumo"}), 403

    data = request.json or {}
    updated = {}

    if 'name' in data:
        name = _clean_str(data.get('name'), max_len=150)
        if name:
            updated['name'] = name
    if 'code' in data:
        raw_code = _clean_str(data.get('code'), max_len=50)
        if raw_code:
            updated['code'] = raw_code.upper().strip()
    if 'category' in data:
        cat_val = _clean_str(data.get('category'), max_len=80)
        updated['category'] = cat_val if cat_val and cat_val.strip() else "Sin Categoría"
    if 'unit' in data:
        updated['unit'] = _clean_str(data.get('unit'), max_len=30) or "Unidad (u)"
    if 'stock_quantity' in data:
        updated['stock_quantity'] = _safe_stock_float(data.get('stock_quantity'), default=0.0, min_val=0.0)
    if 'min_stock' in data:
        updated['min_stock'] = _safe_stock_float(data.get('min_stock'), default=5.0, min_val=0.0)
    if 'cost_price' in data:
        updated['cost_price'] = _safe_stock_float(data.get('cost_price'), default=0.0, min_val=0.0)
    if 'sale_price' in data:
        updated['sale_price'] = _safe_stock_float(data.get('sale_price'), default=0.0, min_val=0.0)
    if 'supplier' in data:
        updated['supplier'] = _clean_str(data.get('supplier'), max_len=150)
    if 'location' in data:
        updated['location'] = _clean_str(data.get('location'), max_len=150)
    if 'batch_number' in data:
        updated['batch_number'] = _clean_str(data.get('batch_number'), max_len=100)
    if 'expiry_date' in data:
        updated['expiry_date'] = _clean_expiry_date(data.get('expiry_date'))
    if 'notes' in data:
        updated['notes'] = _clean_str(data.get('notes'), max_len=500)

    updated['user_id'] = current_uid
    updated['updated_at'] = _now_bolivia().isoformat()

    _ensure_category_and_unit_persisted(updated.get('category'), updated.get('unit'))

    for it in _LOCAL_STOCK_ITEMS:
        if str(it.get('id')) == str(item_id):
            it.update(updated)
            it['status'] = _calc_item_status(it.get('stock_quantity'), it.get('min_stock'))
            break

    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _supabase_update_stock_item(item_id, updated)
    _invalidate_dashboard_cache()

    return jsonify({
        "success": True,
        "message": "Insumo actualizado correctamente."
    })

@stock_bp.route('/api/stock/<item_id>', methods=['DELETE'])
def delete_stock_item(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar insumos.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')
    target_item = _get_stock_item_by_id(item_id)

    if target_item:
        item_uid = target_item.get('user_id')
        if item_uid and item_uid not in (current_uid, 'usr-doctor-001', 'None', 'null', '') and current_user.get('role') != 'admin':
            return jsonify({"error": "No tienes permiso para eliminar este insumo"}), 403

    global _LOCAL_STOCK_ITEMS
    _LOCAL_STOCK_ITEMS = [it for it in _LOCAL_STOCK_ITEMS if str(it.get('id')) != str(item_id)]
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)

    if supabase:
        try:
            supabase.table('stock_items').delete().eq('id', str(item_id)).execute()
        except Exception as e:
            logging.warning("No se pudo eliminar en Supabase remoto: %s", e)

    _invalidate_dashboard_cache()
    return jsonify({"success": True, "message": "Insumo eliminado del catálogo."})

@stock_bp.route('/api/stock/bulk-delete', methods=['POST'])
def bulk_delete_stock_items():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para eliminar insumos.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')
    data = request.json or {}
    ids_to_delete = data.get('ids', [])
    if not ids_to_delete or not isinstance(ids_to_delete, list):
        return jsonify({"error": "Debes proporcionar una lista de IDs para eliminar"}), 400

    str_ids = [str(i) for i in ids_to_delete]
    global _LOCAL_STOCK_ITEMS

    for it in _LOCAL_STOCK_ITEMS:
        if str(it.get('id')) in str_ids:
            item_uid = it.get('user_id')
            if item_uid and item_uid not in (current_uid, 'usr-doctor-001', 'None', 'null', '') and current_user.get('role') != 'admin':
                return jsonify({"error": "No tienes permiso para eliminar algunos insumos seleccionados"}), 403

    _LOCAL_STOCK_ITEMS = [it for it in _LOCAL_STOCK_ITEMS if str(it.get('id')) not in str_ids]
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)

    if supabase:
        try:
            for item_id in str_ids:
                supabase.table('stock_items').delete().eq('id', item_id).execute()
        except Exception as e:
            logging.warning("No se pudo completar bulk-delete en Supabase: %s", e)

    _invalidate_dashboard_cache()
    return jsonify({
        "success": True,
        "deleted_count": len(str_ids),
        "message": f"Se eliminaron {len(str_ids)} insumos del inventario."
    })

@stock_bp.route('/api/stock/<item_id>/adjust', methods=['POST'])
def adjust_stock_item(item_id):
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN para registrar movimientos de inventario.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')
    target_item = _get_stock_item_by_id(item_id)
    if not target_item:
        return jsonify({"error": "Artículo no encontrado"}), 404

    if target_item.get('user_id') and target_item.get('user_id') not in (current_uid, 'usr-doctor-001', 'None', 'null', '') and current_user.get('role') != 'admin':
        return jsonify({"error": "No tienes permiso para registrar movimientos en este insumo"}), 403

    data = request.json or {}
    mov_type = (data.get('type') or 'IN').upper()
    if mov_type not in ['IN', 'OUT', 'ADJUST']:
        mov_type = 'IN'

    qty = _safe_stock_float(data.get('quantity'), default=0.0, min_val=0.0)
    reason = _clean_str(data.get('reason'), max_len=250) or ("Entrada de stock" if mov_type == 'IN' else "Salida / Consumo clínico")

    if qty <= 0 and mov_type != 'ADJUST':
        return jsonify({"error": "La cantidad debe ser mayor a 0"}), 400

    current_qty = _safe_stock_float(target_item.get('stock_quantity'), 0.0)
    if mov_type == 'IN':
        new_qty = round(current_qty + qty, 2)
    elif mov_type == 'OUT':
        if current_qty < qty:
            return jsonify({"error": f"Stock insuficiente. Existencia actual: {current_qty}"}), 400
        new_qty = round(current_qty - qty, 2)
    else: # ADJUST
        new_qty = round(qty, 2)

    updated = {
        "user_id": current_uid,
        "stock_quantity": new_qty,
        "status": _calc_item_status(new_qty, target_item.get('min_stock', 5)),
        "updated_at": _now_bolivia().isoformat()
    }

    for it in _LOCAL_STOCK_ITEMS:
        if str(it.get('id')) == str(item_id):
            it.update(updated)
            break
    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _supabase_update_stock_item(item_id, updated)

    m_item = {
        "id": str(uuid.uuid4()),
        "user_id": current_uid,
        "item_id": str(item_id),
        "type": mov_type,
        "quantity": qty,
        "previous_stock": current_qty,
        "new_stock": new_qty,
        "reason": reason,
        "created_at": _now_bolivia().isoformat()
    }
    _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)
    _supabase_insert_stock_movement(m_item)

    _invalidate_dashboard_cache()
    return jsonify({
        "success": True,
        "new_stock": new_qty,
        "movement": m_item,
        "message": f"Stock actualizado a {new_qty} {target_item.get('unit', 'u')}."
    })

@stock_bp.route('/api/stock/taxonomies', methods=['GET'])
def get_stock_taxonomies():
    default_categories = [
        {"name": "Insumos BIA", "icon": "🩺", "description": "Electrodos tetrapolares, gel conductor, toallitas alcohólicas."},
        {"name": "Suplementos Nutricionales", "icon": "💊", "description": "Proteínas, creatina, multivitamínicos, omegas, minerales."},
        {"name": "Material Clínico e Higiene", "icon": "🧼", "description": "Guantes de nitrilo, mascarillas, batas desechables."},
        {"name": "Accesorios y Equipos", "icon": "📦", "description": "Cintas antropométricas, plicómetros, tallímetros, estuches."},
        {"name": "Medicamentos / Fármacos", "icon": "💉", "description": "Fármacos de prescripción o administración en consultorio."},
        {"name": "Material de Oficina", "icon": "📝", "description": "Hojas de impresión, etiquetas, bolígrafos clínicos."},
        {"name": "Sin Categoría", "icon": "🏷️", "description": "Insumos generales sin clasificación específica."},
        {"name": "Otros", "icon": "🏷️", "description": "Categoría abierta para otros insumos clínicos."}
    ]

    default_units = [
        {"name": "Unidad (u)", "description": "Ítem individual"},
        {"name": "Caja x100", "description": "Presentación en caja de 100 unidades"},
        {"name": "Caja x50", "description": "Presentación en caja de 50 unidades"},
        {"name": "Frasco", "description": "Envase o bote líquido/polvo"},
        {"name": "Bote 900g", "description": "Envase de proteína/suplemento 900g"},
        {"name": "Bote 300g", "description": "Envase de suplemento 300g"},
        {"name": "Pack x10", "description": "Paquete de 10 unidades"},
        {"name": "Kg", "description": "Kilogramos"},
        {"name": "Litro (L)", "description": "Litros líquidos"},
        {"name": "Ampolla", "description": "Dosis inyectable/ampolla"}
    ]

    cats_custom, units_custom = _load_persisted_taxonomies()

    cats_map = {c['name']: c for c in default_categories}
    for c in cats_custom:
        cats_map[c['name']] = c

    units_map = {u['name']: u for u in default_units}
    for u in units_custom:
        units_map[u['name']] = u

    return jsonify({
        "categories": list(cats_map.values()),
        "units": list(units_map.values())
    })

@stock_bp.route('/api/stock/template-excel', methods=['GET'])
def generate_stock_excel_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catálogo de Insumos"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            "SKU",
            "Nombre Producto / Insumo*",
            "Categoría",
            "U. Medida",
            "Stock*",
            "St. Min",
            "P. Costo (Bs)",
            "PVP (Bs)",
            "Lote",
            "Vencimiento",
            "Ubicación",
            "Proveedor",
            "Notas / Posología"
        ]

        header_fill = PatternFill(start_color="107C41", end_color="107C41", fill_type="solid")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        ws.append(headers)
        ws.row_dimensions[1].height = 28

        for col_num, h_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        sample_rows = [
            ["SKU-001", "Electrodos BIA Desechables (Pack x100)", "Insumos BIA", "Caja x100", 15, 5, 120.00, 180.00, "LOT-2026-A", "2027-12-31", "Vitrina 1, Estante B", "BioMedical Import S.R.L.", "Usar exclusivamente en bioimpedancias tetrapolares"],
            ["SKU-002", "Gel Conductor BIA Hipoalergénico 250ml", "Insumos BIA", "Frasco", 20, 3, 25.00, 45.00, "GEL-8890", "2028-06-30", "Vitrina 1, Estante A", "DermoSalud Bolivia", "Conservar en lugar fresco"],
            ["SKU-003", "Proteína Whey Isolate 100% (Bote 900g Vainilla)", "Suplementos Nutricionales", "Bote 900g", 12, 4, 280.00, 350.00, "WHEY-2026-09", "2027-09-15", "Estante Suplementos A", "NutriFit Express", "1 scoop (30g) aporta 25g de proteína pura"],
            ["SKU-004", "Creatina Monohidratada Creapure 300g", "Suplementos Nutricionales", "Bote 300g", 18, 5, 160.00, 210.00, "CREA-102", "2028-03-20", "Estante Suplementos A", "NutriFit Express", "5g diarios en fase de carga o mantenimiento"],
            ["SKU-005", "Multivitamínico Clínico Complejo B + ZINC", "Suplementos Nutricionales", "Caja x60", 25, 6, 85.00, 130.00, "VIT-2026-C", "2027-11-10", "Vitrina 2, Cajón B", "PharmaLife S.A.", "Tomar 1 cápsula junto al almuerzo"]
        ]

        row_font = Font(name="Arial", size=9.5)
        row_alignment = Alignment(vertical="center")

        for r_idx, row_data in enumerate(sample_rows, 2):
            ws.append(row_data)
            ws.row_dimensions[r_idx].height = 22
            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = row_font
                cell.alignment = row_alignment
                cell.border = thin_border
                if c_idx in (5, 6):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif c_idx in (7, 8):
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="Plantilla_Importacion_Insumos_VitaMetrix.xlsx"
        )
    except Exception as e_gen:
        logging.warning("Error al generar plantilla Excel: %s", e_gen)
        return jsonify({"error": "No se pudo generar la plantilla Excel"}), 500

@stock_bp.route('/api/stock/preview-excel', methods=['POST'])
def preview_stock_excel():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    
    current_uid = current_user.get('id')

    if 'file' not in request.files:
        return jsonify({"error": "No se subió ningún archivo"}), 400

    file_obj = request.files['file']
    filename = file_obj.filename.lower()

    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({"error": "Formato no soportado. Sube un archivo Excel (.xlsx, .xls) o CSV (.csv)"}), 400

    file_bytes = file_obj.read()
    if not file_bytes:
        return jsonify({"error": "El archivo subido está vacío"}), 400

    rows_raw = []
    try:
        if filename.endswith('.csv'):
            content_str = file_bytes.decode('utf-8-sig', errors='replace')
            lines = [l for l in content_str.splitlines() if l.strip()]
            for l in lines:
                parts = [p.strip() for p in l.split(';')] if ';' in l else [p.strip() for p in l.split(',')]
                rows_raw.append(parts)
        else:
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "La librería 'openpyxl' no está disponible en el servidor."}), 500
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and any(cell is not None and str(cell).strip() != '' for cell in r):
                    rows_raw.append([str(c).strip() if c is not None else '' for c in r])
    except Exception as ex_read:
        logging.error("Error al previsualizar Excel: %s", ex_read)
        return jsonify({"error": f"No se pudo leer el archivo: {str(ex_read)}"}), 400

    if not rows_raw or len(rows_raw) < 2:
        return jsonify({"error": "El archivo no contiene filas de datos válidas"}), 400

    header_row = [str(h).strip().lower() for h in rows_raw[0]]

    def _find_col_idx(aliases):
        for idx, h in enumerate(header_row):
            h_norm = h.replace('*', '').replace('(', '').replace(')', '').strip()
            for alias in aliases:
                if alias in h_norm or h_norm in alias:
                    return idx
        return -1

    idx_code = _find_col_idx(['sku', 'codigo', 'código', 'code'])
    idx_name = _find_col_idx(['nombre producto / insumo', 'nombre producto', 'nombre insumo', 'nombre', 'producto', 'insumo'])
    idx_cat = _find_col_idx(['categoria', 'categoría', 'category'])
    idx_unit = _find_col_idx(['u. medida', 'u.medida', 'u medida', 'unidad de medida', 'unidad', 'medida', 'unit'])
    idx_qty = _find_col_idx(['stock', 'cantidad inicial', 'cantidad', 'existencia', 'stock_quantity'])
    idx_min = _find_col_idx(['st. min', 'st.min', 'st min', 'stock minimo', 'stock mínimo', 'alerta', 'min_stock'])
    idx_cost = _find_col_idx(['p. costo', 'p.costo', 'p. coste', 'p.coste', 'precio costo', 'precio coste', 'costo', 'coste', 'cost_price'])
    idx_sale = _find_col_idx(['pvp', 'precio venta', 'venta', 'sale_price'])
    idx_batch = _find_col_idx(['lote', 'numero de lote', 'número de lote', 'batch'])
    idx_expiry = _find_col_idx(['vencimiento', 'fecha de vencimiento', 'expiry'])
    idx_location = _find_col_idx(['ubicacion', 'ubicación', 'ubicación en consultorio', 'location'])
    idx_supplier = _find_col_idx(['proveedor', 'supplier'])
    idx_notes = _find_col_idx(['notas / posología', 'notas / posologia', 'notas', 'posologia', 'posología', 'notes'])

    if idx_code == -1 and len(header_row) > 0:
        idx_code = 0
    if idx_name == -1 and len(header_row) > 1:
        idx_name = 1
    if idx_cat == -1 and len(header_row) > 2:
        idx_cat = 2
    if idx_unit == -1 and len(header_row) > 3:
        idx_unit = 3
    if idx_qty == -1 and len(header_row) > 4:
        idx_qty = 4
    if idx_min == -1 and len(header_row) > 5:
        idx_min = 5
    if idx_cost == -1 and len(header_row) > 6:
        idx_cost = 6
    if idx_sale == -1 and len(header_row) > 7:
        idx_sale = 7
    if idx_batch == -1 and len(header_row) > 8:
        idx_batch = 8
    if idx_expiry == -1 and len(header_row) > 9:
        idx_expiry = 9
    if idx_location == -1 and len(header_row) > 10:
        idx_location = 10
    if idx_supplier == -1 and len(header_row) > 11:
        idx_supplier = 11
    if idx_notes == -1 and len(header_row) > 12:
        idx_notes = 12

    preview_items = []
    temp_sku_counter = 1

    for r_num, row_data in enumerate(rows_raw[1:], start=2):
        def _get_val(col_idx):
            if col_idx >= 0 and col_idx < len(row_data):
                val = str(row_data[col_idx]).strip()
                if val.lower() in ('none', 'null', 'nan', ''):
                    return ''
                return val
            return ''

        name = _clean_str(_get_val(idx_name), max_len=150)
        if not name:
            continue

        raw_code = _clean_str(_get_val(idx_code), max_len=50)
        if not raw_code:
            raw_code = f"SKU-AUTO-{temp_sku_counter:03d}"
            temp_sku_counter += 1

        cat = _clean_str(_get_val(idx_cat), max_len=80) or "Sin Categoría"
        unit = _clean_str(_get_val(idx_unit), max_len=30) or "Unidad (u)"
        qty = _safe_stock_float(_get_val(idx_qty), default=0.0, min_val=0.0)
        min_qty = _safe_stock_float(_get_val(idx_min), default=5.0, min_val=0.0)
        cost = _safe_stock_float(_get_val(idx_cost), default=0.0, min_val=0.0)
        sale = _safe_stock_float(_get_val(idx_sale), default=0.0, min_val=0.0)
        batch = _clean_str(_get_val(idx_batch), max_len=100)
        expiry = _clean_expiry_date(_get_val(idx_expiry))
        loc = _clean_str(_get_val(idx_location), max_len=150)
        supp = _clean_str(_get_val(idx_supplier), max_len=150)
        notes = _clean_str(_get_val(idx_notes), max_len=500)

        existing = _find_existing_stock_item(name, raw_code, current_uid=current_uid)
        action_type = "REABASTECER" if existing else "NUEVO"

        preview_items.append({
            "row": r_num,
            "code": raw_code,
            "name": name,
            "category": cat,
            "unit": unit,
            "quantity": qty,
            "min_stock": min_qty,
            "cost_price": cost,
            "sale_price": sale,
            "batch_number": batch,
            "expiry_date": expiry,
            "location": loc,
            "supplier": supp,
            "notes": notes,
            "action": action_type,
            "matched_existing": bool(existing)
        })

    return jsonify({
        "success": True,
        "filename": file_obj.filename,
        "total_detected": len(preview_items),
        "items": preview_items
    })

@stock_bp.route('/api/stock/import-excel', methods=['POST'])
def import_stock_excel():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401
    if not _is_subscription_active(current_user):
        return jsonify({
            "error": "Tu suscripción ha vencido (0 días). Canjea un PIN de licencia para importar productos.",
            "subscription_expired": True
        }), 403

    current_uid = current_user.get('id')

    if 'file' not in request.files:
        return jsonify({"error": "No se subió ningún archivo"}), 400

    file_obj = request.files['file']
    filename = file_obj.filename.lower()

    if not (filename.endswith('.xlsx') or filename.endswith('.xls') or filename.endswith('.csv')):
        return jsonify({"error": "Formato no soportado. Sube un archivo Excel (.xlsx, .xls) o CSV (.csv)"}), 400

    file_bytes = file_obj.read()
    if not file_bytes:
        return jsonify({"error": "El archivo subido está vacío"}), 400

    rows_raw = []
    try:
        if filename.endswith('.csv'):
            content_str = file_bytes.decode('utf-8-sig', errors='replace')
            lines = [l for l in content_str.splitlines() if l.strip()]
            for l in lines:
                parts = [p.strip() for p in l.split(';')] if ';' in l else [p.strip() for p in l.split(',')]
                rows_raw.append(parts)
        else:
            try:
                import openpyxl
            except ImportError:
                return jsonify({"error": "La librería 'openpyxl' no está instalada en el servidor."}), 500
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb.active
            for r in ws.iter_rows(values_only=True):
                if r and any(cell is not None and str(cell).strip() != '' for cell in r):
                    rows_raw.append([str(c).strip() if c is not None else '' for c in r])
    except Exception as ex_read:
        logging.error("Error al leer archivo Excel/CSV de importación: %s", ex_read)
        return jsonify({"error": f"No se pudo leer el archivo: {str(ex_read)}"}), 400

    if not rows_raw or len(rows_raw) < 2:
        return jsonify({"error": "El archivo está vacío o no contiene filas de datos"}), 400

    header_row = [str(h).strip().lower() for h in rows_raw[0]]
    
    def _find_col_idx(aliases):
        for idx, h in enumerate(header_row):
            h_norm = h.replace('*', '').replace('(', '').replace(')', '').strip()
            for alias in aliases:
                if alias in h_norm or h_norm in alias:
                    return idx
        return -1

    idx_code = _find_col_idx(['sku', 'codigo', 'código', 'code'])
    idx_name = _find_col_idx(['nombre producto / insumo', 'nombre producto', 'nombre insumo', 'nombre', 'producto', 'insumo'])
    idx_cat = _find_col_idx(['categoria', 'categoría', 'category'])
    idx_unit = _find_col_idx(['u. medida', 'u.medida', 'u medida', 'unidad de medida', 'unidad', 'medida', 'unit'])
    idx_qty = _find_col_idx(['stock', 'cantidad inicial', 'cantidad', 'existencia', 'stock_quantity'])
    idx_min = _find_col_idx(['st. min', 'st.min', 'st min', 'stock minimo', 'stock mínimo', 'alerta', 'min_stock'])
    idx_cost = _find_col_idx(['p. costo', 'p.costo', 'p. coste', 'p.coste', 'precio costo', 'precio coste', 'costo', 'coste', 'cost_price'])
    idx_sale = _find_col_idx(['pvp', 'precio venta', 'venta', 'sale_price'])
    idx_batch = _find_col_idx(['lote', 'numero de lote', 'número de lote', 'batch'])
    idx_expiry = _find_col_idx(['vencimiento', 'fecha de vencimiento', 'expiry'])
    idx_location = _find_col_idx(['ubicacion', 'ubicación', 'ubicación en consultorio', 'location'])
    idx_supplier = _find_col_idx(['proveedor', 'supplier'])
    idx_notes = _find_col_idx(['notas / posología', 'notas / posologia', 'notas', 'posologia', 'posología', 'notes'])

    if idx_code == -1 and len(header_row) > 0:
        idx_code = 0
    if idx_name == -1 and len(header_row) > 1:
        idx_name = 1
    if idx_cat == -1 and len(header_row) > 2:
        idx_cat = 2
    if idx_unit == -1 and len(header_row) > 3:
        idx_unit = 3
    if idx_qty == -1 and len(header_row) > 4:
        idx_qty = 4
    if idx_min == -1 and len(header_row) > 5:
        idx_min = 5
    if idx_cost == -1 and len(header_row) > 6:
        idx_cost = 6
    if idx_sale == -1 and len(header_row) > 7:
        idx_sale = 7
    if idx_batch == -1 and len(header_row) > 8:
        idx_batch = 8
    if idx_expiry == -1 and len(header_row) > 9:
        idx_expiry = 9
    if idx_location == -1 and len(header_row) > 10:
        idx_location = 10
    if idx_supplier == -1 and len(header_row) > 11:
        idx_supplier = 11
    if idx_notes == -1 and len(header_row) > 12:
        idx_notes = 12

    imported_count = 0
    reabastecidos_count = 0
    nuevos_count = 0
    errors_list = []

    for r_num, row_data in enumerate(rows_raw[1:], start=2):
        try:
            def _get_val(col_idx):
                if col_idx >= 0 and col_idx < len(row_data):
                    val = str(row_data[col_idx]).strip()
                    if val.lower() in ('none', 'null', 'nan', ''):
                        return ''
                    return val
                return ''

            name = _clean_str(_get_val(idx_name), max_len=150)
            if not name:
                continue

            raw_code = _clean_str(_get_val(idx_code), max_len=50)
            cat = _clean_str(_get_val(idx_cat), max_len=80) or "Sin Categoría"
            unit = _clean_str(_get_val(idx_unit), max_len=30) or "Unidad (u)"
            qty = _safe_stock_float(_get_val(idx_qty), default=0.0, min_val=0.0)
            min_qty = _safe_stock_float(_get_val(idx_min), default=5.0, min_val=0.0)
            cost = _safe_stock_float(_get_val(idx_cost), default=0.0, min_val=0.0)
            sale = _safe_stock_float(_get_val(idx_sale), default=0.0, min_val=0.0)
            batch = _clean_str(_get_val(idx_batch), max_len=100)
            expiry = _clean_expiry_date(_get_val(idx_expiry))
            loc = _clean_str(_get_val(idx_location), max_len=150)
            supp = _clean_str(_get_val(idx_supplier), max_len=150)
            notes = _clean_str(_get_val(idx_notes), max_len=500)

            _ensure_category_and_unit_persisted(cat, unit)

            existing = _find_existing_stock_item(name, raw_code, current_uid=current_uid)
            if existing:
                item_id = existing.get('id')
                prev_qty = _safe_stock_float(existing.get('stock_quantity'), 0.0)
                new_qty = round(prev_qty + qty, 2)
                
                update_payload = {
                    "user_id": current_uid,
                    "stock_quantity": new_qty,
                    "min_stock": min_qty if min_qty > 0 else existing.get('min_stock', 5),
                    "cost_price": cost if cost > 0 else existing.get('cost_price', 0),
                    "sale_price": sale if sale > 0 else existing.get('sale_price', 0),
                    "supplier": supp or existing.get('supplier', ''),
                    "location": loc or existing.get('location', ''),
                    "batch_number": batch or existing.get('batch_number', ''),
                    "expiry_date": expiry or existing.get('expiry_date', None),
                    "notes": notes or existing.get('notes', ''),
                    "status": _calc_item_status(new_qty, min_qty),
                    "updated_at": _now_bolivia().isoformat()
                }

                for it in _LOCAL_STOCK_ITEMS:
                    if str(it.get('id')) == str(item_id):
                        it.update(update_payload)
                        break

                _supabase_update_stock_item(item_id, update_payload)

                if qty > 0:
                    m_item = {
                        "id": str(uuid.uuid4()),
                        "user_id": current_uid,
                        "item_id": str(item_id),
                        "type": "IN",
                        "quantity": qty,
                        "previous_stock": prev_qty,
                        "new_stock": new_qty,
                        "reason": "Reabastecimiento por importación Excel/CSV",
                        "created_at": _now_bolivia().isoformat()
                    }
                    _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
                    _supabase_insert_stock_movement(m_item)

                reabastecidos_count += 1
            else:
                sku_code = _generate_next_sku(raw_code, current_uid=current_uid)
                item_id = str(uuid.uuid4())

                new_item = {
                    "id": item_id,
                    "user_id": current_uid,
                    "code": sku_code,
                    "name": name,
                    "category": cat,
                    "unit": unit,
                    "stock_quantity": qty,
                    "min_stock": min_qty,
                    "cost_price": cost,
                    "sale_price": sale,
                    "supplier": supp,
                    "location": loc,
                    "batch_number": batch,
                    "expiry_date": expiry,
                    "notes": notes,
                    "status": _calc_item_status(qty, min_qty),
                    "created_at": _now_bolivia().isoformat(),
                    "updated_at": _now_bolivia().isoformat()
                }

                _LOCAL_STOCK_ITEMS.insert(0, new_item)
                _supabase_insert_stock_item(new_item)

                if qty > 0:
                    m_item = {
                        "id": str(uuid.uuid4()),
                        "user_id": current_uid,
                        "item_id": item_id,
                        "type": "IN",
                        "quantity": qty,
                        "previous_stock": 0.0,
                        "new_stock": qty,
                        "reason": "Stock inicial por importación Excel/CSV",
                        "created_at": _now_bolivia().isoformat()
                    }
                    _LOCAL_STOCK_MOVEMENTS.insert(0, m_item)
                    _supabase_insert_stock_movement(m_item)

                nuevos_count += 1

            imported_count += 1
        except Exception as e_row:
            errors_list.append(f"Fila {r_num}: {str(e_row)}")

    _save_persisted_stock_items(_LOCAL_STOCK_ITEMS)
    _save_persisted_stock_movements(_LOCAL_STOCK_MOVEMENTS)
    _invalidate_dashboard_cache()

    return jsonify({
        "success": True,
        "imported_count": imported_count,
        "reabastecidos_count": reabastecidos_count,
        "nuevos_count": nuevos_count,
        "errors": errors_list,
        "message": f"Se procesaron {imported_count} insumos ({reabastecidos_count} reabastecidos y {nuevos_count} nuevos registros)."
    }), 201

@stock_bp.route('/api/stock/movements', methods=['GET'])
def get_all_stock_movements():
    current_user = _get_current_user()
    if not current_user:
        return jsonify({"error": "No autorizado"}), 401

    movements = []
    if supabase:
        try:
            res = supabase.table('stock_movements').select('*').order('created_at', desc=True).limit(200).execute()
            if res and res.data:
                movements = res.data
        except Exception as e:
            logging.warning("Error consultando movimientos en Supabase: %s", e)

    if not movements:
        movements = _load_persisted_stock_movements()

    return jsonify(movements)
